#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
P16 PPT填充脚本 - 从Excel读取数据并填充到PPT模板
解压PPT模板，替换图表数据，重新打包生成最终PPT
"""

import os
import sys
import yaml
import pandas as pd
import zipfile
import xml.etree.ElementTree as ET
from openpyxl import Workbook, load_workbook
import logging

# 设置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class P16PPTFiller:
    def __init__(self, config_path='config.yaml'):
        """初始化PPT填充器"""
        self.config_path = config_path
        self.config = self.load_config()
        self.setup_paths()
        
    def load_config(self):
        """加载配置文件"""
        try:
            with open(self.config_path, 'r', encoding='utf-8') as f:
                return yaml.safe_load(f)
        except Exception as e:
            logger.error(f"加载配置文件失败: {e}")
            sys.exit(1)
    
    def setup_paths(self):
        """设置路径"""
        self.template_ppt = self.config['project']['template_ppt']
        self.output_dir = self.config['project']['output_dir']
        self.tmp_dir = self.config['project']['tmp_dir']
        # Excel 输入文件：确保读取页面目录下的文件名，避免指向 output 子目录
        # - 若配置为绝对路径，尊重配置；
        # - 若配置为相对路径（如 './output/p16_data.xlsx'），仅取文件名并落到页面目录。
        self.excel_file = self.config['output']['excel_file']
        self.page_dir = os.path.dirname(os.path.abspath(__file__))
        if not os.path.isabs(self.excel_file):
            self.excel_file = os.path.join(self.page_dir, os.path.basename(self.excel_file))
        # 优先使用配置中的final_ppt，未配置时退回默认路径
        self.final_ppt = self.config['project'].get(
            'final_ppt', os.path.join(self.output_dir, 'p16-final.pptx')
        )
        
        # 确保输出与临时目录存在（Excel 不在 output 下）
        os.makedirs(self.output_dir, exist_ok=True)
        os.makedirs(self.tmp_dir, exist_ok=True)

    def _col_letter_to_index(self, col: str) -> int:
        """将Excel列字母转换为列索引（A=1）。"""
        idx = 0
        for ch in col.upper():
            if not ('A' <= ch <= 'Z'):
                raise ValueError(f"非法列字母: {col}")
            idx = idx * 26 + (ord(ch) - ord('A') + 1)
        return idx

    def _parse_excel_range(self, range_str: str):
        """解析类似 'Sheet1!$A$1:$D$1' 的范围。
        返回 (sheet, start_col_idx, start_row, end_col_idx, end_row)。
        若不符合预期格式则抛错。"""
        try:
            sheet, area = range_str.split('!')
            area = area.replace('$', '')
            start, end = area.split(':')
            # 拆分列字母与行号
            import re
            m1 = re.match(r"^([A-Za-z]+)(\d+)$", start)
            m2 = re.match(r"^([A-Za-z]+)(\d+)$", end)
            if not m1 or not m2:
                raise ValueError(f"无法解析范围: {range_str}")
            s_col, s_row = m1.group(1), int(m1.group(2))
            e_col, e_row = m2.group(1), int(m2.group(2))
            s_idx = self._col_letter_to_index(s_col)
            e_idx = self._col_letter_to_index(e_col)
            return sheet, s_idx, s_row, e_idx, e_row
        except Exception as e:
            raise ValueError(f"解析Excel范围失败: {range_str} -> {e}")

    def _update_chart_rel_to_xlsx(self, chart_path: str) -> str:
        """将charts/_rels/chartX.xml.rels中的Target改为指向WorkbookX.xlsx，返回xlsx目标绝对路径。"""
        rels_path = os.path.join(os.path.dirname(chart_path), '_rels', os.path.basename(chart_path) + '.rels')
        # chart2.xml -> chart2.xml.rels
        if not os.path.exists(rels_path):
            # WPS/Office可能命名为 chart2.xml.rels 同目录的 _rels 子目录
            rels_path = os.path.join(os.path.dirname(chart_path), '_rels', f"{os.path.basename(chart_path)}.rels")
        if not os.path.exists(rels_path):
            raise FileNotFoundError(f"关系文件缺失: {rels_path}")

        # 读取关系文件并定位嵌入工作簿关系
        tree = ET.parse(rels_path)
        root = tree.getroot()
        ns = {'r': 'http://schemas.openxmlformats.org/package/2006/relationships'}
        # 默认命名空间，元素无前缀
        # 直接遍历所有子元素
        target_xlsx = None
        for rel in root:
            t = rel.attrib.get('Type')
            target = rel.attrib.get('Target')
            if t == 'http://schemas.openxmlformats.org/officeDocument/2006/relationships/package' and target:
                base = os.path.basename(target)
                if base.startswith('Workbook'):
                    # 替换扩展名为 .xlsx
                    name_no_ext = os.path.splitext(base)[0]
                    new_target = os.path.join(os.path.dirname(target), f"{name_no_ext}.xlsx")
                    rel.set('Target', new_target)
                    # 计算绝对路径
                    charts_dir = os.path.dirname(chart_path)
                    ppt_dir = os.path.dirname(charts_dir)
                    embeddings_dir = os.path.join(ppt_dir, 'embeddings')
                    target_xlsx = os.path.join(embeddings_dir, f"{name_no_ext}.xlsx")
                    break
        if not target_xlsx:
            raise RuntimeError(f"未找到嵌入工作簿关系于: {rels_path}")
        # 写回关系文件
        tree.write(rels_path, encoding='utf-8', xml_declaration=True)
        return target_xlsx

    def _write_embedded_workbook_from_chart(self, chart_path: str, root, namespaces, values, labels=None):
        """根据图表公式写入 embeddings/WorkbookX.xlsx 指定范围的数据。
        - 同步更新关系文件指向 .xlsx
        - 仅支持单行范围（如 A1:D1），若检测到多行则报错。
        - values 长度必须与范围宽度一致，否则报错。
        """
        # 读取数值引用公式
        f_node = root.find('.//c:val/c:numRef/c:f', namespaces)
        if f_node is None or not f_node.text:
            raise RuntimeError('图表缺少数值范围公式(c:val/c:numRef/c:f)')
        sheet, s_idx, s_row, e_idx, e_row = self._parse_excel_range(f_node.text)
        if s_row != e_row:
            raise RuntimeError(f"仅支持单行范围写入，检测到多行: {f_node.text}")
        width = e_idx - s_idx + 1
        if width != len(values):
            raise RuntimeError(f"值数量({len(values)})与范围宽度({width})不一致: {f_node.text}")

        # 更新关系到 .xlsx 并准备目标路径
        target_xlsx = self._update_chart_rel_to_xlsx(chart_path)
        os.makedirs(os.path.dirname(target_xlsx), exist_ok=True)

        # 写入工作簿 - 修复：保留原有数据结构，只更新指定范围
        # 首先检查是否已存在工作簿，如果存在则加载它
        if os.path.exists(target_xlsx):
            wb = load_workbook(target_xlsx)
            logger.info(f"加载现有工作簿: {target_xlsx}")
        else:
            wb = Workbook()
            logger.info(f"创建新工作簿: {target_xlsx}")
        
        # 获取或创建工作表
        if sheet in wb.sheetnames:
            ws = wb[sheet]
        else:
            ws = wb.active
            ws.title = sheet
        
        # 从 s_idx 列开始写入到 e_idx 列
        for i, val in enumerate(values):
            col_index = s_idx + i  # openpyxl 列索引为1基
            ws.cell(row=s_row, column=col_index, value=val)
        
        # 可选写入标签到下一行（若提供）
        if labels:
            for i, lab in enumerate(labels):
                col_index = s_idx + i
                ws.cell(row=s_row + 1, column=col_index, value=str(lab))
        
        wb.save(target_xlsx)
        logger.info(f"嵌入工作簿已更新: {target_xlsx}")
    
    def load_excel_data(self):
        """加载Excel数据"""
        logger.info("加载Excel数据...")
        
        try:
            # 读取主趋势数据
            main_trend_df = pd.read_excel(self.excel_file, sheet_name='main_trend')
            logger.info(f"主趋势数据: {len(main_trend_df)} 行")
            
            # 读取渠道分解数据
            channel_df = pd.read_excel(self.excel_file, sheet_name='channel_breakdown')
            logger.info(f"渠道分解数据: {len(channel_df)} 行")
            
            return main_trend_df, channel_df
            
        except Exception as e:
            logger.error(f"加载Excel数据失败: {e}")
            return None, None
    
    def extract_ppt_template(self):
        """使用已存在的 tmp/ppt 内容，不清空、不重新解压。

        要求：
        - 必须存在目录 tmp/ppt/charts 与 tmp/ppt/embeddings；否则报错返回失败（不做兜底）。
        - 保留现有临时文件以便排查与增量更新。
        """
        logger.info("跳过解压，使用已有的 tmp/ppt 内容")
        charts_dir = os.path.join(self.tmp_dir, 'ppt', 'charts')
        embeddings_dir = os.path.join(self.tmp_dir, 'ppt', 'embeddings')
        if not os.path.isdir(charts_dir):
            logger.error(f"缺少图表目录: {charts_dir}")
            return False
        if not os.path.isdir(embeddings_dir):
            logger.error(f"缺少嵌入工作簿目录: {embeddings_dir}")
            return False
        return True
    
    def find_chart_files(self):
        """查找图表文件"""
        logger.info("查找图表文件...")
        
        chart_files = {}
        chart_mappings = self.config['chart_mapping']
        
        # 在ppt/charts目录中查找图表文件
        charts_dir = os.path.join(self.tmp_dir, 'ppt', 'charts')
        
        if not os.path.exists(charts_dir):
            logger.warning(f"图表目录不存在: {charts_dir}")
            return chart_files
        
        for chart_id, chart_info in chart_mappings.items():
            chart_file = chart_info['file']
            chart_path = os.path.join(charts_dir, chart_file)
            
            if os.path.exists(chart_path):
                chart_files[chart_id] = {
                    'path': chart_path,
                    'type': chart_info['type'],
                    'description': chart_info['description']
                }
                logger.info(f"找到图表文件: {chart_id} -> {chart_file}")
            else:
                logger.warning(f"图表文件不存在: {chart_path}")
        
        return chart_files

    def _ensure_content_types_for_xlsx_embeddings(self):
        """确保首次打开PPT即自动识别刷新，[Content_Types].xml 声明所有嵌入的 .xlsx 工作簿。

        背景：
        - 我们将图表的外部数据关系更新为 ".xlsx"（embeddings/WorkbookX.xlsx）。
        - 若 [Content_Types].xml 未声明这些部件，PowerPoint 首次打开时可能无法自动识别为 Excel 工作簿，
          导致需要手动点击"编辑数据"才能刷新，影响首次打开体验。

        处理：
        - 遍历 tmp 下的 "/ppt/embeddings" 目录，收集所有以 ".xlsx" 结尾的文件。
        - 在 [Content_Types].xml 中为每一个文件添加 <Override> 节点：
          PartName="/ppt/embeddings/WorkbookX.xlsx"
          ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        要求：
        - 严格校验：文件与 XML 缺失时直接报错，不做兜底处理。
        - 确保首次打开PPT时所有嵌入工作簿都能被正确识别为Excel文件。
        """
        content_types_path = os.path.join(self.tmp_dir, '[Content_Types].xml')
        embeddings_dir = os.path.join(self.tmp_dir, 'ppt', 'embeddings')

        if not os.path.isfile(content_types_path):
            raise FileNotFoundError(f"缺少内容类型声明文件: {content_types_path}")
        if not os.path.isdir(embeddings_dir):
            raise FileNotFoundError(f"缺少嵌入目录: {embeddings_dir}")

        # 收集所有嵌入的 .xlsx 文件
        xlsx_files = [f for f in os.listdir(embeddings_dir) if f.lower().endswith('.xlsx')]
        if not xlsx_files:
            # 若没有 .xlsx，则说明不需要此步骤；但考虑到我们已切换关系到 .xlsx，通常应当存在。
            # 遵循“不兜底”的原则，这里直接报错提醒初始化缺失。
            raise RuntimeError(f"未发现任何嵌入的 .xlsx 文件于: {embeddings_dir}")

        # 解析 [Content_Types].xml，准备添加 Override 节点
        ns_ct = 'http://schemas.openxmlformats.org/package/2006/content-types'
        ET.register_namespace('', ns_ct)
        tree = ET.parse(content_types_path)
        root = tree.getroot()

        # 已存在的 Override PartName 集合，用于去重
        existing_overrides = set()
        for ov in root.findall(f"{{{ns_ct}}}Override"):
            part = ov.attrib.get('PartName')
            if part:
                existing_overrides.add(part)

        # 为每一个 .xlsx 嵌入添加 Override
        added = 0
        for fname in sorted(xlsx_files):
            # 计算以包根为起点的部件路径，必须以 "/" 开头，使用 POSIX 分隔符
            rel = os.path.relpath(os.path.join(embeddings_dir, fname), self.tmp_dir)
            part_name = '/' + rel.replace(os.sep, '/')
            if part_name in existing_overrides:
                continue
            ET.SubElement(root, f"{{{ns_ct}}}Override", {
                'PartName': part_name,
                'ContentType': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            })
            added += 1

        if added > 0:
            tree.write(content_types_path, encoding='utf-8', xml_declaration=True)
            logger.info(f"[Content_Types].xml 已添加 {added} 个 .xlsx Override 声明")
        else:
            logger.info("[Content_Types].xml 无需更新，所有 .xlsx 已声明")

    def _to_decimal(self, value):
        """将百分比整数(如28)转换为小数(0.28)。

        - 如果值为None，返回None（保持空缺）
        - 如果值>1，按百分比缩放：v/100
        - 如果值<=1，认为已经是小数，原样返回
        """
        if value is None:
            return None
        try:
            v = float(value)
        except Exception:
            return value
        return v/100.0 if v > 1 else v

    def _remove_manual_percent_labels(self):
        """删除 slide1.xml 中所有手工百分比标签，避免与自动标签重复。
        
        实现要点：
        - 在幻灯片形状树内查找所有包含数字+百分号的文本形状
        - 从 spTree 中删除这些形状节点
        - 严格错误处理：文件不存在或解析失败时抛出异常
        """
        slide_path = os.path.join(self.tmp_dir, 'ppt', 'slides', 'slide1.xml')
        if not os.path.exists(slide_path):
            # 严格错误处理：文件缺失直接报错，避免掩盖问题
            raise FileNotFoundError(f"幻灯片文件不存在: {slide_path}")

        try:
            slide_tree = ET.parse(slide_path)
            slide_root = slide_tree.getroot()

            slide_ns = {
                'ns0': 'http://schemas.openxmlformats.org/presentationml/2006/main',
                'ns1': 'http://schemas.openxmlformats.org/drawingml/2006/main'
            }

            sp_tree = slide_root.find('.//ns0:spTree', slide_ns)
            if sp_tree is None:
                # 没有形状树无法执行删除，抛出错误以便上层处理
                raise RuntimeError("未找到形状树 spTree，无法删除手工百分比标签")

            import re
            removed = 0
            # 遍历所有形状，识别含有数字+百分号的文本并删除形状
            for sp in list(sp_tree.findall('ns0:sp', slide_ns)):
                texts = sp.findall('.//ns1:t', slide_ns)
                joined = ''.join([(t.text or '').strip() for t in texts])
                if re.search(r"\d+%", joined):
                    sp_tree.remove(sp)
                    removed += 1

            slide_tree.write(slide_path, encoding='UTF-8', xml_declaration=True)
            logger.info(f"已删除 {removed} 个手工百分比标签，避免与自动标签重复")
        except Exception as e:
            logger.error(f"删除手工百分比标签失败: {e}")
            raise
    
    def _configure_auto_labels(self, root, namespaces, auto_label_settings):
        """配置自动标签标注，确保首次打开即显示正确百分比标签。
        
        关键功能：
        1. 确保自动数据标签显示已启用
        2. 配置标签格式为百分比
        3. 设置标签位置和样式
        """
        logger.debug("配置自动标签标注，确保首次打开即显示正确百分比")
        
        # 先删除所有手工百分比标签，避免与自动标签重复
        # 严格错误处理：删除失败将抛出异常并中断流程
        self._remove_manual_percent_labels()
        
        # 1. 确保数据标签显示已启用（已在_configure_external_update_and_labels中配置）
        # 2. 配置标签格式为百分比格式
        line_chart = root.find('.//c:lineChart', namespaces)
        if line_chart is not None:
            dLbls = line_chart.find('c:dLbls', namespaces)
            if dLbls is None:
                dLbls = ET.SubElement(line_chart, f"{{{namespaces['c']}}}dLbls")
            
            # 确保显示数值
            show_val = dLbls.find('c:showVal', namespaces)
            if show_val is None:
                show_val = ET.SubElement(dLbls, f"{{{namespaces['c']}}}showVal")
            show_val.set('val', '1')  # 确保显示数值
            
            # 配置标签位置为上方
            dLblPos = dLbls.find('c:dLblPos', namespaces)
            if dLblPos is None:
                dLblPos = ET.SubElement(dLbls, f"{{{namespaces['c']}}}dLblPos")
            dLblPos.set('val', 'above')  # 标签位置：上方
            
            # 配置标签格式
            numFmt = dLbls.find('c:numFmt', namespaces)
            if numFmt is None:
                numFmt = ET.SubElement(dLbls, f"{{{namespaces['c']}}}numFmt")
            numFmt.set('formatCode', '0%')  # 百分比格式
            numFmt.set('sourceLinked', '0')  # 不链接到源数据格式
        
        logger.debug("自动标签配置完成")
        line_chart = root.find('.//c:lineChart', namespaces)
        if line_chart is None:
            logger.warning("未找到折线图节点，无法配置自动标签")
            return
        
        # 获取或创建数据标签配置
        dLbls = line_chart.find('c:dLbls', namespaces)
        if dLbls is None:
            dLbls = ET.SubElement(line_chart, f"{{{namespaces['c']}}}dLbls")
        
        # 统一关闭不需要的标签项，开启数值标签
        def _set_flag(parent, tag, val):
            el = parent.find(f'c:{tag}', namespaces)
            if el is None:
                el = ET.SubElement(parent, f"{{{namespaces['c']}}}{tag}")
            el.set('val', str(val))
        
        # 配置标签显示选项
        _set_flag(dLbls, 'showLegendKey', 0)
        _set_flag(dLbls, 'showCatName', 0)
        _set_flag(dLbls, 'showSerName', 0)
        _set_flag(dLbls, 'showPercent', 0)
        _set_flag(dLbls, 'showBubbleSize', 0)
        _set_flag(dLbls, 'showLeaderLines', 0)
        _set_flag(dLbls, 'showVal', 1)  # 启用数值显示
        
        # 配置标签位置（顶部）
        dLblsPos = dLbls.find('c:dLblPos', namespaces)
        if dLblsPos is None:
            dLblsPos = ET.SubElement(dLbls, f"{{{namespaces['c']}}}dLblPos")
        dLblsPos.set('val', 't')  # 顶部位置
        
        # 配置数值格式为百分比
        numFmt = dLbls.find('c:numFmt', namespaces)
        if numFmt is None:
            numFmt = ET.SubElement(dLbls, f"{{{namespaces['c']}}}numFmt")
        numFmt.set('formatCode', '0"%"')  # 整数百分比格式
        numFmt.set('sourceLinked', '0')
        
        # 配置文本属性，避免省略号和换行
        txPr = dLbls.find('c:txPr', namespaces)
        if txPr is None:
            txPr = ET.SubElement(dLbls, f"{{{namespaces['c']}}}txPr")
        
        bodyPr = txPr.find('a:bodyPr', namespaces)
        if bodyPr is None:
            bodyPr = ET.SubElement(txPr, f"{{{namespaces['a']}}}bodyPr")
        
        # 明确设置文本渲染行为
        bodyPr.set('rot', '0')
        bodyPr.set('spcFirstLastPara', '0')
        bodyPr.set('vertOverflow', 'clip')  # 禁用省略号
        bodyPr.set('vert', 'horz')
        bodyPr.set('wrap', 'none')  # 禁用换行
        bodyPr.set('anchor', 'ctr')  # 居中对齐
        bodyPr.set('anchorCtr', '1')
        
        # 自动适配文本到容器
        if bodyPr.find('a:spAutoFit', namespaces) is None:
            ET.SubElement(bodyPr, f"{{{namespaces['a']}}}spAutoFit")
        
        # 配置段落样式
        lstStyle = txPr.find('a:lstStyle', namespaces)
        if lstStyle is None:
            lstStyle = ET.SubElement(txPr, f"{{{namespaces['a']}}}lstStyle")
        
        p = txPr.find('a:p', namespaces)
        if p is None:
            p = ET.SubElement(txPr, f"{{{namespaces['a']}}}p")
        
        pPr = p.find('a:pPr', namespaces)
        if pPr is None:
            pPr = ET.SubElement(p, f"{{{namespaces['a']}}}pPr")
        
        defRPr = pPr.find('a:defRPr', namespaces)
        if defRPr is None:
            defRPr = ET.SubElement(pPr, f"{{{namespaces['a']}}}defRPr")
        
        # 设置语言为中文
        defRPr.set('lang', 'zh-CN')
        
        # 3. 配置系列级数据标签（确保每个系列都启用标签，首次打开即显示）
        for ser in line_chart.findall('c:ser', namespaces):
            dLbls_ser = ser.find('c:dLbls', namespaces)
            if dLbls_ser is None:
                dLbls_ser = ET.SubElement(ser, f"{{{namespaces['c']}}}dLbls")
            
            # 系列级标签配置，确保首次打开即显示数值
            _set_flag(dLbls_ser, 'showVal', 1)
            _set_flag(dLbls_ser, 'showLeaderLines', 0)
            
            # 系列级位置配置
            pos_ser = dLbls_ser.find('c:dLblPos', namespaces)
            if pos_ser is None:
                pos_ser = ET.SubElement(dLbls_ser, f"{{{namespaces['c']}}}dLblPos")
            pos_ser.set('val', 't')
            
            # 系列级数值格式，确保首次打开即显示百分比
            numFmt_ser = dLbls_ser.find('c:numFmt', namespaces)
            if numFmt_ser is None:
                numFmt_ser = ET.SubElement(dLbls_ser, f"{{{namespaces['c']}}}numFmt")
            numFmt_ser.set('formatCode', '0%')
            numFmt_ser.set('sourceLinked', '0')  # 不链接到源格式，确保固定百分比显示
            
            # 系列级文本属性
            txPr_ser = dLbls_ser.find('c:txPr', namespaces)
            if txPr_ser is None:
                txPr_ser = ET.SubElement(dLbls_ser, f"{{{namespaces['c']}}}txPr")
            
            bodyPr_ser = txPr_ser.find('a:bodyPr', namespaces)
            if bodyPr_ser is None:
                bodyPr_ser = ET.SubElement(txPr_ser, f"{{{namespaces['a']}}}bodyPr")
            
            bodyPr_ser.set('rot', '0')
            bodyPr_ser.set('spcFirstLastPara', '0')
            bodyPr_ser.set('vertOverflow', 'clip')
            bodyPr_ser.set('vert', 'horz')
            bodyPr_ser.set('wrap', 'none')
            bodyPr_ser.set('anchor', 'ctr')
            bodyPr_ser.set('anchorCtr', '1')
            
            if bodyPr_ser.find('a:spAutoFit', namespaces) is None:
                ET.SubElement(bodyPr_ser, f"{{{namespaces['a']}}}spAutoFit")
        
        logger.debug("自动标签配置完成，确保首次打开即显示正确标签")
    
    

    def _remove_point_label_overrides(self, root, namespaces):
        """清理点级标签覆写，避免显示系列名/类目名造成拥挤。

        删除所有 c:ser/c:dPt 下的 c:dLbls，使标签仅由系列级或图表级控制。
        手工标签逻辑已移除，统一采用自动标签策略。
        """
        removed = 0
        
        # 1. 删除所有点级标签覆写（c:dPt下的c:dLbls）
        for dpt in root.findall('.//c:ser/c:dPt', namespaces):
            dlbls = dpt.find('c:dLbls', namespaces)
            if dlbls is not None:
                dpt.remove(dlbls)
                removed += 1
        
        # 2. 统一使用自动标签，手工标签在 _configure_auto_labels 中被清理
        
        if removed:
            logger.debug(f"移除 {removed} 个点级标签覆写")
        logger.debug("保留手工百分比标签，将在后续步骤中更新数值")

    def _configure_external_update_and_labels(self, root, namespaces, chart_path: str):
        """重写外部数据刷新与数据标签配置，确保首次打开PPT即显示最新值。
        
        关键改进：
        1. 强制启用外部数据自动刷新
        2. 确保数据连接正确配置
        3. 彻底清理可能干扰刷新的属性
        4. 统一数值格式为百分比
        5. 强制启用数据标签显示
        6. 确保缓存数据与嵌入工作簿一致
        """
        logger.debug("开始配置外部数据刷新和数据标签")
        
        # 1. 强制启用外部数据自动刷新
        external_data = root.find('.//c:externalData', namespaces)
        # 先解析关系文件以获取正确的 r:id
        rels_dir = os.path.join(os.path.dirname(chart_path), '_rels')
        rels_file = os.path.join(rels_dir, os.path.basename(chart_path) + '.rels')
        if not os.path.exists(rels_file):
            raise FileNotFoundError(f"缺少图表关系文件: {rels_file}")
        try:
            rels_tree = ET.parse(rels_file)
            rels_root = rels_tree.getroot()
            correct_rel_id = None
            for rel in rels_root.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                if rel.get('Type') == 'http://schemas.openxmlformats.org/officeDocument/2006/relationships/package':
                    correct_rel_id = rel.get('Id')
                    break
            if not correct_rel_id:
                raise RuntimeError(f"关系文件中未找到嵌入工作簿关系: {rels_file}")
        except Exception as e:
            raise RuntimeError(f"解析图表关系文件失败: {e}")

        if external_data is None:
            # 创建外部数据节点
            chart_space = root.find('.//c:chartSpace', namespaces)
            if chart_space is None:
                raise RuntimeError('缺少 c:chartSpace 节点，无法创建外部数据连接')
            external_data = ET.SubElement(chart_space, f"{{{namespaces['c']}}}externalData")
        # 设置/纠正 externalData 的 r:id 与关系文件一致
        external_data.set('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id', correct_rel_id)
        
        # 强制设置自动刷新：使用 c:autoUpdate 子元素，移除非法属性
        # 说明：c:externalData 不支持 autoUpdate 属性，必须通过子元素配置
        auto_update = external_data.find('c:autoUpdate', namespaces)
        if auto_update is None:
            auto_update = ET.SubElement(external_data, f"{{{namespaces['c']}}}autoUpdate")
        auto_update.set('val', '1')
        if 'autoUpdate' in external_data.attrib:
            del external_data.attrib['autoUpdate']

        # 调试输出：在 chart1.xml 上打印完整 XML，便于确认 externalData 结构
        try:
            if os.path.basename(chart_path) == 'chart1.xml':
                logger.debug('调试：打印 chart1.xml 修改后的 XML 树')
                ET.dump(root)
        except Exception as _:
            pass
        
        # 清理可能干扰刷新的属性
        for attr in ['refreshError']:
            if attr in external_data.attrib:
                del external_data.attrib[attr]

        # 2. 强制设置数值缓存格式为百分比
        num_cache = root.find('.//c:val/c:numRef/c:numCache', namespaces)
        if num_cache is not None:
            fmt_code = num_cache.find('c:formatCode', namespaces)
            if fmt_code is None:
                fmt_code = ET.SubElement(num_cache, f"{{{namespaces['c']}}}formatCode")
            fmt_code.text = '0%'

        # 3. 强制设置坐标轴为百分比格式
        for val_ax in root.findall('.//c:valAx', namespaces):
            numfmt = val_ax.find('c:numFmt', namespaces)
            if numfmt is None:
                numfmt = ET.SubElement(val_ax, f"{{{namespaces['c']}}}numFmt")
            numfmt.set('formatCode', '0%')
            numfmt.set('sourceLinked', '1')

        # 4. 强制启用数据标签显示（图表级）
        line_chart = root.find('.//c:lineChart', namespaces)
        if line_chart is not None:
            dLbls = line_chart.find('c:dLbls', namespaces)
            if dLbls is None:
                dLbls = ET.SubElement(line_chart, f"{{{namespaces['c']}}}dLbls")

            # 强制启用数值显示
            for tag_name, val in [
                ('showLegendKey', '0'),
                ('showVal', '1'),  # 强制启用自动数值显示
                ('showCatName', '0'),
                ('showSerName', '0'),
                ('showPercent', '0'),
                ('showBubbleSize', '0'),
                ('showLeaderLines', '0')
            ]:
                el = dLbls.find(f'c:{tag_name}', namespaces)
                if el is None:
                    el = ET.SubElement(dLbls, f"{{{namespaces['c']}}}{tag_name}")
                el.set('val', val)

            # 强制设置百分比格式
            numfmt = dLbls.find('c:numFmt', namespaces)
            if numfmt is None:
                numfmt = ET.SubElement(dLbls, f"{{{namespaces['c']}}}numFmt")
            numfmt.set('formatCode', '0%')
            numfmt.set('sourceLinked', '0')  # 不链接到源格式

            # 强制设置标签位置
            dLblsPos = dLbls.find('c:dLblPos', namespaces)
            if dLblsPos is None:
                dLblsPos = ET.SubElement(dLbls, f"{{{namespaces['c']}}}dLblPos")
            dLblsPos.set('val', 't')  # 上方位置

        # 5. 系列级设置（强制启用显示，确保首次打开即显示）
        for s in root.findall('.//c:ser', namespaces):
            s_dlbls = s.find('c:dLbls', namespaces)
            if s_dlbls is None:
                s_dlbls = ET.SubElement(s, f"{{{namespaces['c']}}}dLbls")

            # 强制启用数值显示
            for tag_name, val in [
                ('showVal', '1'),  # 强制启用自动数值显示
                ('showSerName', '0'),
                ('showCatName', '0'),
                ('showPercent', '0'),
                ('showLegendKey', '0'),
                ('showBubbleSize', '0'),
                ('showLeaderLines', '0')
            ]:
                el = s_dlbls.find(f'c:{tag_name}', namespaces)
                if el is None:
                    el = ET.SubElement(s_dlbls, f"{{{namespaces['c']}}}{tag_name}")
                el.set('val', val)

            # 强制设置百分比格式，确保首次打开即显示正确值
            s_numfmt = s_dlbls.find('c:numFmt', namespaces)
            if s_numfmt is None:
                s_numfmt = ET.SubElement(s_dlbls, f"{{{namespaces['c']}}}numFmt")
            s_numfmt.set('formatCode', '0%')
            s_numfmt.set('sourceLinked', '0')  # 不链接到源格式，确保固定百分比显示

            # 强制设置标签位置
            s_dlblsPos = s_dlbls.find('c:dLblPos', namespaces)
            if s_dlblsPos is None:
                s_dlblsPos = ET.SubElement(s_dlbls, f"{{{namespaces['c']}}}dLblPos")
            s_dlblsPos.set('val', 't')  # 上方位置

        logger.debug("外部数据刷新和数据标签配置完成，确保首次打开显示最新值")

    def _remove_chart_extensions(self, root, namespaces):
        """重写图表扩展移除，确保彻底清理干扰扩展。

        关键改进：
        1. 递归查找并移除所有扩展列表
        2. 确保彻底清理第三方扩展缓存
        3. 添加详细调试日志
        """
        removed_count = 0
        
        def remove_element(element):
            """安全移除元素"""
            try:
                # 使用ElementTree的通用方法查找父元素
                for parent in root.iter():
                    if element in list(parent):
                        parent.remove(element)
                        return True
                return False
            except Exception as e:
                logger.debug(f"移除元素失败: {e}")
                return False
        
        # 递归查找所有扩展列表
        for ext_lst in root.findall('.//c:extLst', namespaces):
            if remove_element(ext_lst):
                removed_count += 1
                logger.debug(f"已移除图表扩展列表")
            else:
                logger.warning("扩展列表无父节点，无法移除")
        
        # 同时查找并移除可能的扩展元素
        for ext in root.findall('.//c:ext', namespaces):
            if remove_element(ext):
                removed_count += 1
                logger.debug(f"已移除扩展元素")
        
        if removed_count > 0:
            logger.debug(f"共移除 {removed_count} 个扩展相关元素")
        else:
            logger.debug("未找到图表扩展列表或扩展元素")

    def _overwrite_num_cache(self, root, namespaces, data_values, series_idx=0):
        """重写数值缓存覆盖，确保首次打开PPT即显示最新值。

        关键改进：
        1. 强制更新所有系列的数值缓存
        2. 确保ptCount与实际数据点数量一致
        3. 彻底清理旧缓存数据
        4. 添加调试日志便于排查
        5. 确保缓存数据与嵌入工作簿一致
        """
        
        # 查找所有系列
        series = root.findall('.//c:ser', namespaces)
        if not series:
            logger.warning("未找到任何系列")
            return
        
        # 如果指定了系列索引，只处理该系列
        if series_idx < len(series):
            target_series = [series[series_idx]]
        else:
            logger.warning(f"系列索引 {series_idx} 超出范围，共有 {len(series)} 个系列")
            target_series = series

        for ser_idx, ser in enumerate(target_series):
            # 移除可能存在的直接数值缓存 numLit，避免覆盖效果被旧值干扰
            val_node = ser.find('c:val', namespaces)
            if val_node is not None:
                num_lit = val_node.find('c:numLit', namespaces)
                if num_lit is not None:
                    val_node.remove(num_lit)
                    logger.debug(f"系列 {ser_idx} 已移除 numLit 以强制使用 numRef/numCache")
            
            # 查找数值缓存
            num_cache = ser.find('.//c:val/c:numRef/c:numCache', namespaces)
            if num_cache is None:
                logger.warning(f"系列 {ser_idx} 未找到数值缓存，尝试创建")
                # 尝试创建数值缓存结构
                val = ser.find('c:val', namespaces)
                if val is None:
                    val = ET.SubElement(ser, f"{{{namespaces['c']}}}val")

                num_ref = val.find('c:numRef', namespaces)
                if num_ref is None:
                    num_ref = ET.SubElement(val, f"{{{namespaces['c']}}}numRef")
                
                num_cache = ET.SubElement(num_ref, f"{{{namespaces['c']}}}numCache")
                # 添加格式代码
                fmt_code = ET.SubElement(num_cache, f"{{{namespaces['c']}}}formatCode")
                fmt_code.text = '0%'

            # 强制更新ptCount，确保与实际数据一致
            pt_count = num_cache.find('c:ptCount', namespaces)
            if pt_count is None:
                pt_count = ET.SubElement(num_cache, f"{{{namespaces['c']}}}ptCount")
            pt_count.set('val', str(len(data_values)))

            # 彻底删除所有现有pt元素，确保无旧数据残留
            for pt in num_cache.findall('c:pt', namespaces):
                num_cache.remove(pt)

            # 添加新的pt元素，确保与嵌入工作簿数据一致
            for idx, value in enumerate(data_values):
                pt = ET.SubElement(num_cache, f"{{{namespaces['c']}}}pt")
                pt.set('idx', str(idx))
                
                v = ET.SubElement(pt, f"{{{namespaces['c']}}}v")
                # 确保数值格式正确，与嵌入工作簿保持一致
                v.text = str(float(value)) if value is not None else "0"

            logger.debug(f"已覆盖系列 {ser_idx} 的数值缓存，共 {len(data_values)} 个数据点，确保首次打开显示最新值")

        logger.debug("数值缓存覆盖完成，确保缓存与嵌入工作簿一致")

    def _overwrite_str_cache(self, root, namespaces, labels, series_idx=0):
        """重写字符串缓存覆盖，确保首次打开PPT即显示最新标签。

        关键改进：
        1. 强制更新所有系列的字符串缓存
        2. 确保ptCount与实际标签数量一致
        3. 彻底清理旧缓存数据
        4. 添加调试日志便于排查
        5. 确保标签数据与嵌入工作簿一致
        """
        
        # 查找所有系列
        series = root.findall('.//c:ser', namespaces)
        if not series:
            logger.warning("未找到任何系列")
            return
        
        # 如果指定了系列索引，只处理该系列
        if series_idx < len(series):
            target_series = [series[series_idx]]
        else:
            logger.warning(f"系列索引 {series_idx} 超出范围，共有 {len(series)} 个系列")
            target_series = series

        for ser_idx, ser in enumerate(target_series):
            # 查找字符串缓存；若存在 strLit，先移除以避免旧值干扰
            cat_node = ser.find('c:cat', namespaces)
            if cat_node is not None:
                str_lit = cat_node.find('c:strLit', namespaces)
                if str_lit is not None:
                    cat_node.remove(str_lit)
                    logger.debug(f"系列 {ser_idx} 已移除 strLit 以强制使用 strRef/strCache")
            
            # 查找字符串缓存
            str_cache = ser.find('.//c:cat/c:strRef/c:strCache', namespaces)
            if str_cache is None:
                logger.warning(f"系列 {ser_idx} 未找到字符串缓存，尝试创建")
                # 尝试创建字符串缓存结构
                cat = ser.find('c:cat', namespaces)
                if cat is None:
                    cat = ET.SubElement(ser, f"{{{namespaces['c']}}}cat")
                
                str_ref = cat.find('c:strRef', namespaces)
                if str_ref is None:
                    str_ref = ET.SubElement(cat, f"{{{namespaces['c']}}}strRef")
                
                str_cache = ET.SubElement(str_ref, f"{{{namespaces['c']}}}strCache")

            # 强制更新ptCount，确保与实际标签一致
            pt_count = str_cache.find('c:ptCount', namespaces)
            if pt_count is None:
                pt_count = ET.SubElement(str_cache, f"{{{namespaces['c']}}}ptCount")
            pt_count.set('val', str(len(labels)))

            # 彻底删除所有现有pt元素，确保无旧标签残留
            for pt in str_cache.findall('c:pt', namespaces):
                str_cache.remove(pt)

            # 添加新的pt元素，确保与嵌入工作簿标签一致
            for idx, label in enumerate(labels):
                pt = ET.SubElement(str_cache, f"{{{namespaces['c']}}}pt")
                pt.set('idx', str(idx))
                
                v = ET.SubElement(pt, f"{{{namespaces['c']}}}v")
                # 确保标签格式正确，与嵌入工作簿保持一致
                v.text = str(label) if label is not None else ""

            logger.debug(f"已覆盖系列 {ser_idx} 的字符串缓存，共 {len(labels)} 个标签，确保首次打开显示最新标签")

            logger.debug(f"已覆盖系列 {ser_idx} 的字符串缓存，共 {len(labels)} 个标签")

        logger.debug("字符串缓存覆盖完成")

    def _set_value_axis_auto_scale(self, root, namespaces):
        """移除数值轴上的显式 min/max，让轴根据数据自动缩放到小数比例。"""
        for val_ax in root.findall('.//c:valAx', namespaces):
            scaling = val_ax.find('c:scaling', namespaces)
            if scaling is None:
                continue
            # 删除 min/max，避免沿用模板的整数百分比范围（导致小数数据被压扁）
            for tag in ['min', 'max']:
                node = scaling.find(f"c:{tag}", namespaces)
                if node is not None:
                    scaling.remove(node)
            # 若存在固定主刻度，删除以便自动选择
            major = val_ax.find('c:majorUnit', namespaces)
            if major is not None:
                val_ax.remove(major)
    
    def update_main_trend_chart(self, chart_path, main_trend_data):
        """重写主趋势图表更新，确保首次打开PPT即显示最新值。

        关键改进：
        1. 强制覆盖所有缓存数据
        2. 确保数据一致性
        3. 统一错误处理
        4. 添加详细调试日志
        """
        logger.info(f"更新主趋势图表: {chart_path}")
        
        try:
            # 解析XML文件
            tree = ET.parse(chart_path)
            root = tree.getroot()
            
            namespaces = {
                'c': 'http://schemas.openxmlformats.org/drawingml/2006/chart',
                'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'
            }
            
            # 准备数据（严格：任何缺失值直接报错，不做兜底）
            values_for_cache = []
            values_for_embedded = []  # 嵌入工作簿使用小数值
            labels_for_cache = []
            
            for _, row in main_trend_data.iterrows():
                val = self._to_decimal(row['sov_percent'])
                if val is None:
                    raise RuntimeError(f"主趋势图表存在缺失值，月份: {row.get('month_display', row.get('month'))}")
                values_for_cache.append(val)
                # 嵌入工作簿也写入小数，确保首次打开即正确显示百分比
                values_for_embedded.append(val)
                labels_for_cache.append(row['month_display'])
            
            logger.debug(f"准备更新主趋势图表数据: {len(values_for_cache)} 个数据点，确保首次打开显示最新值")
            
            # 1. 强制覆盖数值缓存
            try:
                self._overwrite_num_cache(root, namespaces, values_for_cache)
                logger.debug("数值缓存覆盖成功")
            except Exception as e:
                logger.error(f"数值缓存覆盖失败: {e}")
                return False
            
            # 2. 强制覆盖字符串缓存
            try:
                self._overwrite_str_cache(root, namespaces, labels_for_cache)
                logger.debug("字符串缓存覆盖成功")
            except Exception as e:
                logger.error(f"字符串缓存覆盖失败: {e}")
                return False
            
            # 3. 清理点级标签覆写
            try:
                self._remove_point_label_overrides(root, namespaces)
                logger.debug("点级标签清理成功")
            except Exception as e:
                logger.error(f"点级标签清理失败: {e}")
                return False
            
            # 4. 配置外部刷新与标签（传入 chart_path 以确保 r:id 正确）
            try:
                self._configure_external_update_and_labels(root, namespaces, chart_path)
                logger.debug("外部刷新配置成功")
            except Exception as e:
                logger.error(f"外部刷新配置失败: {e}")
                return False
            
            # 5. 根据标签模式配置选择标签更新方式
            label_mode = self.config.get('label_mode', {})
            auto_labels = label_mode.get('auto_labels', False)
            
            if auto_labels:
                # 自动标签标注：隐藏/删除所有手工标签，启用自动标签
                try:
                    self._configure_auto_labels(root, namespaces, label_mode.get('auto_label_settings', {}))
                    logger.debug("主趋势图表自动标签配置成功")
                except Exception as e:
                    logger.error(f"主趋势图表自动标签配置失败: {e}")
                    return False
            
            # 6. 设置轴自动缩放
            try:
                self._set_value_axis_auto_scale(root, namespaces)
                logger.debug("轴自动缩放设置成功")
            except Exception as e:
                logger.error(f"轴自动缩放设置失败: {e}")
                return False
            
            # 7. 移除图表扩展
            try:
                self._remove_chart_extensions(root, namespaces)
                logger.debug("图表扩展移除成功")
            except Exception as e:
                logger.error(f"图表扩展移除失败: {e}")
                return False
            
            # 保存更新后的XML
            tree.write(chart_path, encoding='utf-8', xml_declaration=True)
            
            # 8. 写入嵌入工作簿
            try:
                self._write_embedded_workbook_from_chart(chart_path, root, namespaces, values_for_embedded)
                logger.debug("嵌入工作簿写入成功")
            except Exception as e:
                logger.error(f"嵌入工作簿写入失败: {e}")
                return False
            
            logger.info("主趋势图表更新完成")
            return True
            
        except Exception as e:
            logger.error(f"更新主趋势图表失败: {e}")
            return False
    
    def update_channel_chart(self, chart_path, channel_data, channel_name):
        """重写渠道图表更新，确保首次打开PPT即显示最新值。

        关键改进：
        1. 强制覆盖所有缓存数据
        2. 确保数据一致性
        3. 统一错误处理
        4. 添加详细调试日志
        """
        logger.info(f"更新渠道图表 {channel_name}: {chart_path}")
        
        try:
            # 解析XML文件
            tree = ET.parse(chart_path)
            root = tree.getroot()
            
            namespaces = {
                'c': 'http://schemas.openxmlformats.org/drawingml/2006/chart',
                'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'
            }
            
            # 过滤当前渠道的数据
            channel_specific_data = channel_data[channel_data['channel'] == channel_name]
            
            if channel_specific_data.empty:
                raise RuntimeError(f"渠道 {channel_name} 数据为空")
            
            # 准备数据（严格：任何缺失值直接报错，不做兜底）
            values_for_cache = []
            values_for_embedded = []  # 嵌入工作簿使用小数值
            labels_for_cache = []
            
            for _, row in channel_specific_data.iterrows():
                val = self._to_decimal(row['sov_percent'])
                if val is None:
                    raise RuntimeError(f"渠道 {channel_name} 存在缺失值，月份: {row.get('month_display', row.get('month'))}")
                values_for_cache.append(val)
                # 嵌入工作簿也写入小数，确保打开即正确显示百分比
                values_for_embedded.append(val)
                labels_for_cache.append(row['month_display'])
            
            logger.debug(f"准备更新渠道 {channel_name} 图表数据: {len(values_for_cache)} 个数据点")
            
            # 1. 强制覆盖数值缓存
            try:
                self._overwrite_num_cache(root, namespaces, values_for_cache)
                logger.debug(f"渠道 {channel_name} 数值缓存覆盖成功")
            except Exception as e:
                logger.error(f"渠道 {channel_name} 数值缓存覆盖失败: {e}")
                return False
            
            # 2. 强制覆盖字符串缓存
            try:
                self._overwrite_str_cache(root, namespaces, labels_for_cache)
                logger.debug(f"渠道 {channel_name} 字符串缓存覆盖成功")
            except Exception as e:
                logger.error(f"渠道 {channel_name} 字符串缓存覆盖失败: {e}")
                return False
            
            # 3. 清理点级标签覆写
            try:
                self._remove_point_label_overrides(root, namespaces)
                logger.debug(f"渠道 {channel_name} 点级标签清理成功")
            except Exception as e:
                logger.error(f"渠道 {channel_name} 点级标签清理失败: {e}")
                return False
            
            # 4. 配置外部刷新与标签（传入 chart_path 以确保 r:id 正确）
            try:
                self._configure_external_update_and_labels(root, namespaces, chart_path)
                logger.debug(f"渠道 {channel_name} 外部刷新配置成功")
            except Exception as e:
                logger.error(f"渠道 {channel_name} 外部刷新配置失败: {e}")
                return False
            
            # 5. 根据标签模式配置选择标签更新方式
            label_mode = self.config.get('label_mode', {})
            auto_labels = label_mode.get('auto_labels', False)
            
            if auto_labels:
                # 自动标签标注：隐藏/删除所有手工标签，启用自动标签
                try:
                    self._configure_auto_labels(root, namespaces, label_mode.get('auto_label_settings', {}))
                    logger.debug(f"渠道 {channel_name} 自动标签配置成功")
                except Exception as e:
                    logger.error(f"渠道 {channel_name} 自动标签配置失败: {e}")
                    return False
            
            # 6. 设置轴自动缩放
            try:
                self._set_value_axis_auto_scale(root, namespaces)
                logger.debug(f"渠道 {channel_name} 轴自动缩放设置成功")
            except Exception as e:
                logger.error(f"渠道 {channel_name} 轴自动缩放设置失败: {e}")
                return False
            
            # 7. 移除图表扩展
            try:
                self._remove_chart_extensions(root, namespaces)
                logger.debug(f"渠道 {channel_name} 图表扩展移除成功")
            except Exception as e:
                logger.error(f"渠道 {channel_name} 图表扩展移除失败: {e}")
                return False
            
            # 保存更新后的XML
            tree.write(chart_path, encoding='utf-8', xml_declaration=True)
            
            # 8. 写入嵌入工作簿
            try:
                self._write_embedded_workbook_from_chart(chart_path, root, namespaces, values_for_embedded)
                logger.debug(f"渠道 {channel_name} 嵌入工作簿写入成功")
            except Exception as e:
                logger.error(f"渠道 {channel_name} 嵌入工作簿写入失败: {e}")
                return False
            
            logger.info(f"渠道图表 {channel_name} 更新完成")
            return True
            
        except Exception as e:
            logger.error(f"更新渠道图表 {channel_name} 失败: {e}")
            return False
    
    def update_all_charts(self, main_trend_data, channel_data):
        """更新所有图表"""
        logger.info("开始更新所有图表...")
        
        # 查找图表文件
        chart_files = self.find_chart_files()
        
        if not chart_files:
            logger.error("未找到任何图表文件")
            return False
        
        success_count = 0
        total = len(chart_files)
        
        # 更新每个图表
        for chart_id, chart_info in chart_files.items():
            chart_type = chart_info['type']
            chart_path = chart_info['path']
            
            try:
                if chart_type == 'main_trend':
                    # 更新主趋势图表
                    if self.update_main_trend_chart(chart_path, main_trend_data):
                        success_count += 1
                
                elif chart_type == 'channel_breakdown':
                    # 从图表ID中提取渠道名称
                    channel_name = self.extract_channel_from_chart_id(chart_id)
                    if channel_name:
                        if self.update_channel_chart(chart_path, channel_data, channel_name):
                            success_count += 1
                    else:
                        logger.warning(f"无法从图表ID {chart_id} 提取渠道名称")
                
            except Exception as e:
                logger.error(f"更新图表 {chart_id} 时出错: {e}")
        
        logger.info(f"图表更新完成，成功更新 {success_count}/{total} 个图表")
        # 任一失败则整体失败
        return success_count == total
    
    def extract_channel_from_chart_id(self, chart_id):
        """从图表ID提取渠道名称"""
        # 根据配置文件中的映射关系
        chart_mappings = self.config['chart_mapping']
        
        if chart_id in chart_mappings:
            description = chart_mappings[chart_id]['description']
            
            # 从描述中提取渠道名称
            if 'Forum' in description:
                return 'Forum'
            elif 'Online News' in description:
                return 'Online News'
            elif 'Blog' in description:
                return 'Blog'
            elif 'X' in description or 'Twitter' in description:
                return 'X'
            elif 'Instagram' in description:
                return 'Instagram'
            elif 'YouTube' in description:
                return 'YouTube'
        
        return None
    
    def repackage_ppt(self):
        """重新打包PPT文件"""
        logger.info("重新打包PPT文件...")
        
        try:
            # 删除已存在的最终文件
            if os.path.exists(self.final_ppt):
                os.remove(self.final_ppt)
            
            # 创建新的ZIP文件
            with zipfile.ZipFile(self.final_ppt, 'w', zipfile.ZIP_DEFLATED) as zip_ref:
                # 遍历临时目录中的所有文件
                for root, dirs, files in os.walk(self.tmp_dir):
                    for file in files:
                        file_path = os.path.join(root, file)
                        # 计算相对路径
                        arcname = os.path.relpath(file_path, self.tmp_dir)
                        zip_ref.write(file_path, arcname)
            
            logger.info(f"PPT文件重新打包完成: {self.final_ppt}")
            return True
            
        except Exception as e:
            logger.error(f"重新打包PPT文件失败: {e}")
            return False
    
    def cleanup_temp_files(self):
        """清理临时文件（保留tmp目录以便排查）"""
        logger.info("保持临时目录以便排查: %s", self.tmp_dir)
    
    def fill_ppt(self):
        """填充PPT的主流程"""
        logger.info("开始PPT填充流程...")
        
        # 1. 加载Excel数据
        main_trend_data, channel_data = self.load_excel_data()
        if main_trend_data is None or channel_data is None:
            return False
        
        # 2. 解压PPT模板
        if not self.extract_ppt_template():
            return False
        
        # 3. 更新图表数据（任一失败直接中止）
        if not self.update_all_charts(main_trend_data, channel_data):
            logger.error("图表更新失败，终止打包流程")
            return False
        
        # 3.1 更新内容类型以声明 .xlsx 嵌入，确保打开即自动识别刷新
        try:
            self._ensure_content_types_for_xlsx_embeddings()
        except Exception as e:
            logger.error(f"更新 [Content_Types].xml 失败: {e}")
            return False

        # 3.2 清理旧的 .xlsb 嵌入对象，避免与 .xlsx 干扰导致文件损坏
        self._cleanup_old_xlsb_embeddings()
        
        # 4. 重新打包PPT
        if not self.repackage_ppt():
            return False
        
        # 5. 清理临时文件
        self.cleanup_temp_files()
        
        logger.info("PPT填充流程完成")
        return True

    def _cleanup_old_xlsb_embeddings(self):
        """在 tmp/ppt/embeddings 中移除所有 .xlsb 文件，防止与 .xlsx 冲突。

        说明：
        - 模板通常包含 `WorkbookN.xlsb`，本流程改为写入并引用 `WorkbookN.xlsx`；
        - 两者并存会使 PowerPoint 内容校验混淆，提示“文件已损坏”；
        - 严格策略：若嵌入目录缺失则直接报错，不做兜底。
        """
        embeddings_dir = os.path.join(self.tmp_dir, 'ppt', 'embeddings')
        if not os.path.isdir(embeddings_dir):
            raise FileNotFoundError(f"缺少嵌入工作簿目录: {embeddings_dir}")
        for fname in os.listdir(embeddings_dir):
            if fname.lower().endswith('.xlsb'):
                fpath = os.path.join(embeddings_dir, fname)
                os.remove(fpath)
                logger.info(f"已删除旧的 .xlsb 嵌入对象: {fname}")

def main():
    """主函数"""
    logger.info("=== P16 PPT填充开始 ===")
    
    # 切换到脚本目录
    script_dir = os.path.dirname(os.path.abspath(__file__))
    os.chdir(script_dir)
    
    # 创建PPT填充器
    filler = P16PPTFiller()
    
    # 执行填充
    success = filler.fill_ppt()
    
    if success:
        logger.info("=== P16 PPT填充完成 ===")
        return 0
    else:
        logger.error("=== P16 PPT填充失败 ===")
        return 1

if __name__ == "__main__":
    sys.exit(main())
