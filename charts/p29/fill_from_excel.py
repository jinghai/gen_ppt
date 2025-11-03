#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
P29页面PPT数据填充器
从Excel文件读取数据并填充到PPT模板中
"""

import json
import re
import pandas as pd
import yaml
from pathlib import Path
import shutil
import zipfile
import tempfile
from lxml import etree
import openpyxl

# 项目路径配置
ROOT = Path(__file__).resolve().parent

# 页面级临时目录（遵循“页面级代码使用页面级 tmp”）
# 说明：所有 P29 相关的临时文件均放置在 charts/p29/tmp 下，
# 避免污染项目根目录并便于页面内自洽管理。
TMP_DIR = ROOT / 'tmp'

def load_config():
    """加载页面级配置文件"""
    config_path = ROOT / 'config.yaml'
    with open(config_path, 'r', encoding='utf-8') as f:
        config = yaml.safe_load(f)
    return config

def read_excel_data(excel_path):
    """从Excel文件读取数据"""
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel文件不存在：{excel_path}")
    
    # 读取各个工作表，跳过标题行
    sov_data = pd.read_excel(excel_path, sheet_name='渠道SOV数据', skiprows=1)
    pie_data = pd.read_excel(excel_path, sheet_name='品牌总体SOV', skiprows=1)
    
    return sov_data, pie_data

def prepare_chart_data(sov_data, pie_data, config, excel_path):
    """
    准备图表数据格式
    - 左侧堆叠柱状图：优先从 Excel 的 Sheet1 读取（保留你手动修改的值）；
      若不存在 Sheet1 或布局非法，将抛错，不做兜底。
    - 右侧饼图：从“品牌总体SOV”读取（用于饼图和右侧文本）。
    """
    channels = config['fill_policy']['channel_order']
    brands_display = config['filters']['brands_display']

    # 优先用 Sheet1（B..列为品牌，A 列为渠道；第1行为表头，第1列为渠道名）
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if 'Sheet1' not in wb.sheetnames:
        raise RuntimeError('Excel 缺少 Sheet1，请先使用 generate_excel.py 生成标准布局')
    ws = wb['Sheet1']

    # 按配置顺序读取各品牌、各渠道的整数/小数百分比，统一转为 float
    chart_labels = channels
    chart_series = []
    for c_idx, brand in enumerate(brands_display, start=2):
        header_val = ws.cell(row=1, column=c_idx).value
        if str(header_val) != brand:
            raise RuntimeError(f'Sheet1 第 {c_idx} 列表头应为 {brand}，实际为 {header_val}')
        values = []
        for r_idx, channel in enumerate(channels, start=2):
            row_label = ws.cell(row=r_idx, column=1).value
            if str(row_label) != channel:
                raise RuntimeError(f'Sheet1 第 {r_idx} 行渠道应为 {channel}，实际为 {row_label}')
            v = ws.cell(row=r_idx, column=c_idx).value
            try:
                values.append(float(v or 0))
            except Exception:
                raise RuntimeError(f'Sheet1 值不可解析为数字：行 {r_idx} 列 {c_idx}，值 {v}')
        chart_series.append({'name': brand, 'values': values})

    # 右侧饼图数据
    pie_labels = []
    pie_values = []
    for _, row in pie_data.iterrows():
        if row['Percentage'] > 0:  # 只包含有数据的品牌
            pie_labels.append(row['Brand'])
            pie_values.append(float(row['Percentage']))

    print('柱状图数据来源：Sheet1（保留手动修改）')
    return {
        'bar_chart': {
            'labels': chart_labels,
            'series': chart_series
        },
        'pie_chart': {
            'labels': pie_labels,
            'values': pie_values
        }
    }

def update_chart_data_json(chart_data):
    """准备图表数据（不再依赖chart115目录）"""
    # 写入堆叠柱状图数据
    bar_data = {
        'labels': chart_data['bar_chart']['labels'],
        'series': chart_data['bar_chart']['series']
    }
    
    # 临时数据写入页面级 tmp 目录，遵循“页面级 tmp”规则
    TMP_DIR.mkdir(parents=True, exist_ok=True)
    
    # 可选：保存数据到临时文件用于调试
    temp_data_path = TMP_DIR / 'chart_data.json'
    with open(temp_data_path, 'w', encoding='utf-8') as f:
        json.dump(bar_data, f, ensure_ascii=False, indent=2)
    
    print(f"图表数据已准备完成，临时保存到：{temp_data_path}")
    return bar_data

def extract_ppt_template():
    """解压PPT模板到临时目录"""
    template_ppt = ROOT / 'p29.pptx'
    if not template_ppt.exists():
        raise FileNotFoundError(f"PPT模板不存在：{template_ppt}")
    
    # 创建临时目录（页面级 tmp 下）
    TMP_DIR.mkdir(parents=True, exist_ok=True)
    
    # 解压PPT
    extract_dir = TMP_DIR / 'ppt_extracted'
    if extract_dir.exists():
        shutil.rmtree(extract_dir)
    
    with zipfile.ZipFile(template_ppt, 'r') as zip_ref:
        zip_ref.extractall(extract_dir)
    
    print(f"PPT模板已解压到：{extract_dir}")
    return extract_dir

def assert_sheet1_layout(excel_path, config):
    """校验 Excel 中的 Sheet1 是否为标准布局，否则抛错（不兜底）。"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if 'Sheet1' not in wb.sheetnames:
        raise RuntimeError('Excel 缺少 Sheet1，请先运行 generate_excel.py 生成标准布局')
    ws = wb['Sheet1']
    brands_display = config['filters']['brands_display']
    channels = config['fill_policy']['channel_order']

    # 检查表头品牌
    for idx, brand in enumerate(brands_display, start=2):
        val = ws.cell(row=1, column=idx).value
        if str(val) != brand:
            raise RuntimeError(f'Sheet1 表头第 {idx} 列应为 {brand}，实际为 {val}')
    # 检查渠道行
    for r_idx, channel in enumerate(channels, start=2):
        val = ws.cell(row=r_idx, column=1).value
        if str(val) != channel:
            raise RuntimeError(f'Sheet1 第 {r_idx} 行渠道应为 {channel}，实际为 {val}')
    return True

def update_right_total_sov_texts(extract_dir, pie_data, config):
    """
    填充右侧总体 SOV 文本标签：
    - 在 slide1.xml 中寻找包含“%”的纯文本形状；按垂直位置自上而下排序；
    - 依次写入 brands_display 顺序对应的整数百分比。
    - 若未找到足够的候选形状则抛错（遵循“不兜底”原则）。
    """
    slides_dir = extract_dir / 'ppt' / 'slides'
    target_slide = slides_dir / 'slide1.xml'
    if not target_slide.exists():
        raise FileNotFoundError(f'未找到 {target_slide}，无法填充右侧总体SOV文本')

    # 构建品牌 -> 百分比（整数）映射
    brand_order = config['filters']['brands_display']
    pct_map = {str(row['Brand']): int(row['Percentage']) for _, row in pie_data.iterrows()}

    ns = {
        'p': 'http://schemas.openxmlformats.org/presentationml/2006/main',
        'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
    }
    tree = etree.parse(str(target_slide))
    root = tree.getroot()

    candidates = []
    for sp in root.findall('.//p:sp', ns):
        # 获取文本内容
        texts = sp.findall('.//a:t', ns)
        joined = ''.join([(t.text or '').strip() for t in texts])
        if not joined:
            continue
        # 仅匹配形如 “23%” 的标签
        if re.fullmatch(r"\d+%", joined):
            off = sp.find('.//p:spPr/a:xfrm/a:off', ns)
            y = int(off.get('y')) if off is not None and off.get('y') else 0
            candidates.append((y, sp, texts))

    if len(candidates) < len(brand_order):
        raise RuntimeError(f'右侧文本候选仅 {len(candidates)} 个，少于品牌数量 {len(brand_order)}，无法安全填充')

    candidates.sort(key=lambda x: x[0])  # 按 y 从上到下
    for idx, brand in enumerate(brand_order):
        pct = pct_map.get(brand, 0)
        t_nodes = candidates[idx][2]
        if not t_nodes:
            raise RuntimeError('发现空文本节点，无法写入百分比')
        t_nodes[0].text = f"{pct}%"

    tree.write(str(target_slide), xml_declaration=True, encoding='UTF-8', standalone='yes')
    print('已填充右侧总体SOV文本标签（slide1.xml）')

def update_chart_xml_caches(extract_dir, chart_data):
    """
    刷新 PPT 中 chart*.xml 的缓存数据（numCache/strCache），避免打开后仍显示旧数据。
    不修改嵌入工作簿与关系，保证“可编辑”能力仍在，需要时 PowerPoint 可继续引用嵌入数据。
    """
    charts_dir = extract_dir / 'ppt' / 'charts'
    if not charts_dir.exists():
        print("警告：未找到 charts 目录，无法更新缓存")
        return

    ns = {
        'c': 'http://schemas.openxmlformats.org/drawingml/2006/chart',
        'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
    }

    labels = chart_data['bar_chart']['labels']
    series_list = chart_data['bar_chart']['series']
    # 使用 chart_data 序列顺序作为 Sheet1 表头（B..H）顺序的权威来源
    brands_order = [s['name'] for s in series_list]
    series_by_name = {s['name']: s['values'] for s in series_list}

    pie_labels = chart_data['pie_chart']['labels']
    pie_values = chart_data['pie_chart']['values']

    def _write_str_cache(cache_el, values):
        # 清空并写入 strCache
        for e in list(cache_el):
            cache_el.remove(e)
        pt_count = etree.SubElement(cache_el, f"{{{ns['c']}}}ptCount")
        pt_count.set('val', str(len(values)))
        for idx, v in enumerate(values):
            pt = etree.SubElement(cache_el, f"{{{ns['c']}}}pt")
            pt.set('idx', str(idx))
            v_el = etree.SubElement(pt, f"{{{ns['c']}}}v")
            v_el.text = str(v)

    def _write_num_cache(cache_el, numbers):
        # 清空并写入 numCache
        for e in list(cache_el):
            cache_el.remove(e)
        pt_count = etree.SubElement(cache_el, f"{{{ns['c']}}}ptCount")
        pt_count.set('val', str(len(numbers)))
        for idx, num in enumerate(numbers):
            pt = etree.SubElement(cache_el, f"{{{ns['c']}}}pt")
            pt.set('idx', str(idx))
            v_el = etree.SubElement(pt, f"{{{ns['c']}}}v")
            # 保留 1 位小数，与 Excel 输出一致
            v_el.text = f"{float(num):.1f}"

    updated_files = []

    for chart_xml in charts_dir.glob('chart*.xml'):
        try:
            tree = etree.parse(str(chart_xml))
            root = tree.getroot()

            # 更新堆叠柱状图（barChart）
            bar_charts = root.findall('.//c:barChart', ns)
            updated = False
            for bar in bar_charts:
                # 更新每个系列的分类（cat）与数值（val）缓存
                for s_idx, ser in enumerate(bar.findall('c:ser', ns)):
                    # 读取系列名以匹配到正确数据（优先根据系列名公式的列字母推断品牌）
                    name_text = None
                    brand_from_tx_f = None
                    tx_f = ser.find('.//c:tx//c:strRef//c:f', ns)
                    if tx_f is not None and tx_f.text and '!' in tx_f.text:
                        # 形如 Sheet1!$D$1 -> 列 D
                        try:
                            addr = tx_f.text.split('!')[1]
                            col_letter = addr.split('$')[1]
                            # Sheet1 列从 B 开始映射到 brands_order[0]
                            from openpyxl.utils import column_index_from_string
                            col_idx_1 = column_index_from_string(col_letter)
                            brand_idx = col_idx_1 - 2
                            if 0 <= brand_idx < len(brands_order):
                                brand_from_tx_f = brands_order[brand_idx]
                        except Exception:
                            brand_from_tx_f = None
                    tx = ser.find('.//c:tx', ns)
                    if tx is not None:
                        # 可能存在 strRef/strCache 或直接 c:v
                        name_el = tx.find('.//c:strCache/c:pt/c:v', ns)
                        if name_el is None:
                            name_el = tx.find('.//c:v', ns)
                        if name_el is not None and name_el.text:
                            name_text = name_el.text.strip()

                    # 分类缓存（横轴标签）—写入统一的 labels
                    cat_cache = ser.find('.//c:cat//c:strCache', ns)
                    if cat_cache is None:
                        # 若不存在则创建结构 c:cat/c:strRef/c:strCache
                        cat = ser.find('.//c:cat', ns)
                        if cat is None:
                            cat = etree.SubElement(ser, f"{{{ns['c']}}}cat")
                        str_ref = cat.find('c:strRef', ns)
                        if str_ref is None:
                            str_ref = etree.SubElement(cat, f"{{{ns['c']}}}strRef")
                        cat_cache = str_ref.find('c:strCache', ns)
                        if cat_cache is None:
                            cat_cache = etree.SubElement(str_ref, f"{{{ns['c']}}}strCache")
                    _write_str_cache(cat_cache, labels)

                    # 数值缓存（垂直数据）—优先用 tx_f 列字母推断的品牌，其次用现有 strCache 名，再次用出现顺序
                    values = None
                    target_brand = brand_from_tx_f or name_text
                    if target_brand and target_brand in series_by_name:
                        values = series_by_name[target_brand]
                    elif 0 <= s_idx < len(series_list):
                        values = series_list[s_idx]['values']
                    else:
                        values = []

                    num_cache = ser.find('.//c:val//c:numCache', ns)
                    if num_cache is None:
                        val = ser.find('.//c:val', ns)
                        if val is None:
                            val = etree.SubElement(ser, f"{{{ns['c']}}}val")
                        num_ref = val.find('c:numRef', ns)
                        if num_ref is None:
                            num_ref = etree.SubElement(val, f"{{{ns['c']}}}numRef")
                        num_cache = num_ref.find('c:numCache', ns)
                        if num_cache is None:
                            num_cache = etree.SubElement(num_ref, f"{{{ns['c']}}}numCache")
                    _write_num_cache(num_cache, values)

                    # 同步 tx 的 strCache 为目标品牌，避免后续基于 name_text 的逻辑错配
                    if target_brand:
                        tx_str_cache = ser.find('.//c:tx//c:strRef//c:strCache', ns)
                        if tx_str_cache is None:
                            tx = ser.find('c:tx', ns)
                            if tx is None:
                                tx = etree.SubElement(ser, f"{{{ns['c']}}}tx")
                            str_ref = tx.find('c:strRef', ns)
                            if str_ref is None:
                                str_ref = etree.SubElement(tx, f"{{{ns['c']}}}strRef")
                            tx_str_cache = str_ref.find('c:strCache', ns)
                            if tx_str_cache is None:
                                tx_str_cache = etree.SubElement(str_ref, f"{{{ns['c']}}}strCache")
                        _write_str_cache(tx_str_cache, [target_brand])
                    updated = True

            # 更新饼图（pieChart）
            pie_charts = root.findall('.//c:pieChart', ns)
            for pie in pie_charts:
                ser = pie.find('c:ser', ns)
                if ser is not None:
                    cat_cache = ser.find('.//c:cat//c:strCache', ns)
                    if cat_cache is None:
                        cat = ser.find('.//c:cat', ns)
                        if cat is None:
                            cat = etree.SubElement(ser, f"{{{ns['c']}}}cat")
                        str_ref = cat.find('c:strRef', ns)
                        if str_ref is None:
                            str_ref = etree.SubElement(cat, f"{{{ns['c']}}}strRef")
                        cat_cache = str_ref.find('c:strCache', ns)
                        if cat_cache is None:
                            cat_cache = etree.SubElement(str_ref, f"{{{ns['c']}}}strCache")
                    _write_str_cache(cat_cache, pie_labels)

                    num_cache = ser.find('.//c:val//c:numCache', ns)
                    if num_cache is None:
                        val = ser.find('.//c:val', ns)
                        if val is None:
                            val = etree.SubElement(ser, f"{{{ns['c']}}}val")
                        num_ref = val.find('c:numRef', ns)
                        if num_ref is None:
                            num_ref = etree.SubElement(val, f"{{{ns['c']}}}numRef")
                        num_cache = num_ref.find('c:numCache', ns)
                        if num_cache is None:
                            num_cache = etree.SubElement(num_ref, f"{{{ns['c']}}}numCache")
                    _write_num_cache(num_cache, pie_values)
                    updated = True

            if updated:
                tree.write(str(chart_xml), xml_declaration=True, encoding='UTF-8', standalone='yes')
                updated_files.append(chart_xml.name)

        except Exception as e:
            print(f"更新 {chart_xml.name} 时出错：{e}")

    # 启用 externalData 的自动更新，打开即刷新
    chart_rels_dir = extract_dir / 'ppt' / 'charts' / '_rels'
    for rel_file in chart_rels_dir.glob('chart*.xml.rels'):
        try:
            rel_tree = etree.parse(str(rel_file))
            rel_root = rel_tree.getroot()
            # 保持 rId1 指向嵌入工作簿，后续可能改为 xlsx
            # 无需改 Id，只需要确保 Target 正确
            rel_tree.write(str(rel_file), xml_declaration=True, encoding='UTF-8', standalone='yes')
        except Exception:
            pass

    if updated_files:
        print("已刷新以下图表缓存：")
        for name in updated_files:
            print(f"  - {name}")
    else:
        print("未发现可更新的图表缓存（可能模板未包含 bar/pie 图表）")

def normalize_bar_chart_data_labels(extract_dir, chart_data, config):
    """
    规范柱状图数据标签：
    - 为每个系列设置 dLbls：显示数值、bestFit 位置、统一百分号格式（0%）。
    - 移除模板中针对特定索引的 <c:dLbl><c:delete/> 标签，避免某些品牌在某些渠道不显示标签。
    - 可选：按配置阈值（fill_policy.min_label_percent）为过小的值添加 delete，从而减少拥挤（默认 0，不隐藏）。

    说明：仅影响显示，不改动数据；遵循不兜底原则，若配置未设阈值则不做隐藏。
    """
    charts_dir = extract_dir / 'ppt' / 'charts'
    if not charts_dir.exists():
        print('未找到 charts 目录，跳过数据标签规范')
        return

    ns = {
        'c': 'http://schemas.openxmlformats.org/drawingml/2006/chart',
        'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
    }

    # 从配置读取最小显示阈值（百分比整数），默认 0 表示不隐藏
    min_label_percent = int(((config.get('fill_policy') or {}).get('min_label_percent') or 0))

    # 建立 series 名 -> 数值序列 的映射，便于阈值判断
    series_by_name = {s['name']: s['values'] for s in chart_data['bar_chart']['series']}

    for chart_xml in charts_dir.glob('chart*.xml'):
        try:
            tree = etree.parse(str(chart_xml))
            root = tree.getroot()
            changed = False

            for bar in root.findall('.//c:barChart', ns):
                for ser in bar.findall('c:ser', ns):
                    # 取系列名（优先 c:tx/c:strCache，其次 c:tx/c:v）
                    brand_name = None
                    tx_cache_v = ser.find('.//c:tx//c:strCache//c:pt//c:v', ns)
                    if tx_cache_v is not None and (tx_cache_v.text or '').strip():
                        brand_name = tx_cache_v.text.strip()
                    else:
                        tx_v = ser.find('.//c:tx//c:v', ns)
                        if tx_v is not None and (tx_v.text or '').strip():
                            brand_name = tx_v.text.strip()

                    # 确保存在 dLbls
                    dLbls = ser.find('c:dLbls', ns)
                    if dLbls is None:
                        dLbls = etree.SubElement(ser, f"{{{ns['c']}}}dLbls")

                    # 清理针对索引的删除项，避免缺失标签
                    for d in list(dLbls.findall('c:dLbl', ns)):
                        del_el = d.find('c:delete', ns)
                        if del_el is not None:
                            dLbls.remove(d)
                            changed = True

                    # 统一设置显示选项与格式
                    def ensure_child(parent, tag):
                        el = parent.find(tag, ns)
                        if el is None:
                            el = etree.SubElement(parent, f"{{{ns['c']}}}{tag.split(':')[1]}")
                        return el

                    # dLblPos bestFit
                    pos = dLbls.find('c:dLblPos', ns)
                    if pos is None:
                        pos = etree.SubElement(dLbls, f"{{{ns['c']}}}dLblPos")
                    pos.set('val', 'bestFit')

                    # showVal=1，其他保持关闭
                    show_val = dLbls.find('c:showVal', ns)
                    if show_val is None:
                        show_val = etree.SubElement(dLbls, f"{{{ns['c']}}}showVal")
                    show_val.set('val', '1')
                    for tag in ['c:showLegendKey','c:showCatName','c:showSerName','c:showPercent','c:showBubbleSize']:
                        el = dLbls.find(tag, ns)
                        if el is None:
                            el = etree.SubElement(dLbls, f"{{{ns['c']}}}{tag.split(':')[1]}")
                        el.set('val', '0')

                    # 百分比显示格式：将整数值按“数字+百分号”显示，而不进行 *100 的缩放。
                    # Excel 的标准百分比格式 0% 会把 36 解释为 3600%，因此这里采用自定义格式 0\%，
                    # 仅追加符号，不改变数值。
                    num_fmt = dLbls.find('c:numFmt', ns)
                    if num_fmt is None:
                        num_fmt = etree.SubElement(dLbls, f"{{{ns['c']}}}numFmt")
                    num_fmt.set('formatCode', '0\\%')
                    num_fmt.set('sourceLinked', '0')
                    changed = True

                    # 统一设置数据标签字体颜色：默认白色；Apple 品牌统一为黑色。
                    # 说明：遵循 MVP 原则保持简单，后续如需动态对比度再扩展。
                    # Apple 设为黑色，其余默认白色
                    label_color = '000000' if (brand_name == 'Apple') else 'FFFFFF'
                    print(f"数据标签颜色: series='{brand_name}' -> {label_color}")
                    txPr = dLbls.find('c:txPr', ns)
                    if txPr is None:
                        txPr = etree.SubElement(dLbls, f"{{{ns['c']}}}txPr")
                        etree.SubElement(txPr, f"{{{ns['a']}}}bodyPr")
                        etree.SubElement(txPr, f"{{{ns['a']}}}lstStyle")
                        p = etree.SubElement(txPr, f"{{{ns['a']}}}p")
                        # 追加段落属性并设置默认字符属性颜色（defRPr）
                        pPr = etree.SubElement(p, f"{{{ns['a']}}}pPr")
                        defRPr = etree.SubElement(pPr, f"{{{ns['a']}}}defRPr")
                        solidFill_def = etree.SubElement(defRPr, f"{{{ns['a']}}}solidFill")
                        srgb_def = etree.SubElement(solidFill_def, f"{{{ns['a']}}}srgbClr")
                        srgb_def.set('val', label_color)
                        r = etree.SubElement(p, f"{{{ns['a']}}}r")
                        rPr = etree.SubElement(r, f"{{{ns['a']}}}rPr")
                        # 设置填充为纯色 label_color（Apple 黑，其它白）
                        solidFill = etree.SubElement(rPr, f"{{{ns['a']}}}solidFill")
                        srgb = etree.SubElement(solidFill, f"{{{ns['a']}}}srgbClr")
                        srgb.set('val', label_color)
                        changed = True
                        # 追加一个空文本节点，避免 r 无内容导致部分渲染器忽略 rPr
                        etree.SubElement(r, f"{{{ns['a']}}}t").text = ''

                        # 同步更新子级数据标签（c:dLbl）上的 txPr（模板可能包含），避免颜色被子级覆盖
                        for d in dLbls.findall('c:dLbl', ns):
                            txPr_child = d.find('c:txPr', ns)
                            if txPr_child is None:
                                continue
                            # 子级 rPr 颜色
                            p_child = txPr_child.find('a:p', ns)
                            if p_child is None:
                                p_child = etree.SubElement(txPr_child, f"{{{ns['a']}}}p")
                            r_child = p_child.find('a:r', ns)
                            if r_child is None:
                                r_child = etree.SubElement(p_child, f"{{{ns['a']}}}r")
                            rPr_child = r_child.find('a:rPr', ns)
                            if rPr_child is None:
                                rPr_child = etree.SubElement(r_child, f"{{{ns['a']}}}rPr")
                            solid_child = rPr_child.find('a:solidFill', ns)
                            if solid_child is None:
                                solid_child = etree.SubElement(rPr_child, f"{{{ns['a']}}}solidFill")
                            srgb_child = solid_child.find('a:srgbClr', ns)
                            if srgb_child is None:
                                srgb_child = etree.SubElement(solid_child, f"{{{ns['a']}}}srgbClr")
                            srgb_child.set('val', label_color)
                            # 子级 defRPr 颜色
                            pPr_child = p_child.find('a:pPr', ns)
                            if pPr_child is None:
                                pPr_child = etree.SubElement(p_child, f"{{{ns['a']}}}pPr")
                            defRPr_child = pPr_child.find('a:defRPr', ns)
                            if defRPr_child is None:
                                defRPr_child = etree.SubElement(pPr_child, f"{{{ns['a']}}}defRPr")
                            solid_def_child = defRPr_child.find('a:solidFill', ns)
                            if solid_def_child is None:
                                solid_def_child = etree.SubElement(defRPr_child, f"{{{ns['a']}}}solidFill")
                            srgb_def_child = solid_def_child.find('a:srgbClr', ns)
                            if srgb_def_child is None:
                                srgb_def_child = etree.SubElement(solid_def_child, f"{{{ns['a']}}}srgbClr")
                            srgb_def_child.set('val', label_color)
                            changed = True
                    else:
                        # 若已存在 txPr，确保存在 rPr 且颜色按规则设置
                        p = txPr.find('a:p', ns)
                        if p is None:
                            p = etree.SubElement(txPr, f"{{{ns['a']}}}p")
                        r = p.find('a:r', ns)
                        if r is None:
                            r = etree.SubElement(p, f"{{{ns['a']}}}r")
                        rPr = r.find('a:rPr', ns)
                        if rPr is None:
                            rPr = etree.SubElement(r, f"{{{ns['a']}}}rPr")
                        solidFill = rPr.find('a:solidFill', ns)
                        if solidFill is None:
                            solidFill = etree.SubElement(rPr, f"{{{ns['a']}}}solidFill")
                        srgb = solidFill.find('a:srgbClr', ns)
                        if srgb is None:
                            srgb = etree.SubElement(solidFill, f"{{{ns['a']}}}srgbClr")
                        srgb.set('val', label_color)
                        changed = True
                        # 同步更新默认字符属性 defRPr 的颜色，避免渲染器优先使用 defRPr
                        pPr = p.find('a:pPr', ns)
                        if pPr is None:
                            pPr = etree.SubElement(p, f"{{{ns['a']}}}pPr")
                        defRPr = pPr.find('a:defRPr', ns)
                        if defRPr is None:
                            defRPr = etree.SubElement(pPr, f"{{{ns['a']}}}defRPr")
                        solidFill_def = defRPr.find('a:solidFill', ns)
                        if solidFill_def is None:
                            solidFill_def = etree.SubElement(defRPr, f"{{{ns['a']}}}solidFill")
                        srgb_def = solidFill_def.find('a:srgbClr', ns)
                        if srgb_def is None:
                            srgb_def = etree.SubElement(solidFill_def, f"{{{ns['a']}}}srgbClr")
                        srgb_def.set('val', label_color)
                        changed = True

                        # 同步更新子级数据标签（c:dLbl）上的 txPr（模板可能包含），避免颜色被子级覆盖
                        for d in dLbls.findall('c:dLbl', ns):
                            txPr_child = d.find('c:txPr', ns)
                            if txPr_child is None:
                                continue
                            # 子级 rPr 颜色
                            p_child = txPr_child.find('a:p', ns)
                            if p_child is None:
                                p_child = etree.SubElement(txPr_child, f"{{{ns['a']}}}p")
                            r_child = p_child.find('a:r', ns)
                            if r_child is None:
                                r_child = etree.SubElement(p_child, f"{{{ns['a']}}}r")
                            rPr_child = r_child.find('a:rPr', ns)
                            if rPr_child is None:
                                rPr_child = etree.SubElement(r_child, f"{{{ns['a']}}}rPr")
                            solid_child = rPr_child.find('a:solidFill', ns)
                            if solid_child is None:
                                solid_child = etree.SubElement(rPr_child, f"{{{ns['a']}}}solidFill")
                            srgb_child = solid_child.find('a:srgbClr', ns)
                            if srgb_child is None:
                                srgb_child = etree.SubElement(solid_child, f"{{{ns['a']}}}srgbClr")
                            srgb_child.set('val', label_color)
                            # 子级 defRPr 颜色
                            pPr_child = p_child.find('a:pPr', ns)
                            if pPr_child is None:
                                pPr_child = etree.SubElement(p_child, f"{{{ns['a']}}}pPr")
                            defRPr_child = pPr_child.find('a:defRPr', ns)
                            if defRPr_child is None:
                                defRPr_child = etree.SubElement(pPr_child, f"{{{ns['a']}}}defRPr")
                            solid_def_child = defRPr_child.find('a:solidFill', ns)
                            if solid_def_child is None:
                                solid_def_child = etree.SubElement(defRPr_child, f"{{{ns['a']}}}solidFill")
                            srgb_def_child = solid_def_child.find('a:srgbClr', ns)
                            if srgb_def_child is None:
                                srgb_def_child = etree.SubElement(solid_def_child, f"{{{ns['a']}}}srgbClr")
                            srgb_def_child.set('val', label_color)
                            changed = True

                    # 可选：按阈值隐藏过小的标签（仅影响显示，不改动数据）
                    if brand_name in series_by_name and min_label_percent > 0:
                        values = series_by_name[brand_name]
                        for idx, v in enumerate(values):
                            try:
                                if float(v) < float(min_label_percent):
                                    d = etree.SubElement(dLbls, f"{{{ns['c']}}}dLbl")
                                    i = etree.SubElement(d, f"{{{ns['c']}}}idx")
                                    i.set('val', str(idx))
                                    dele = etree.SubElement(d, f"{{{ns['c']}}}delete")
                                    dele.set('val', '1')
                                    changed = True
                            except Exception:
                                # 保守处理：解析失败不隐藏
                                pass

            if changed:
                tree.write(str(chart_xml), xml_declaration=True, encoding='UTF-8', standalone='yes')
                print(f"已规范数据标签：{chart_xml.name}")
        except Exception as e:
            print(f"规范数据标签失败 {chart_xml.name}：{e}")

def update_embedded_excel_and_links(extract_dir, excel_path):
    """
    用最新的 p29_data.xlsx 替换嵌入工作簿，并把关系的 Target 指向 .xlsx；
    同时将 chart XML 的 externalData 自动更新开关设为 1。
    """
    embeddings_dir = extract_dir / 'ppt' / 'embeddings'
    charts_rels_dir = extract_dir / 'ppt' / 'charts' / '_rels'
    content_types_path = extract_dir / '[Content_Types].xml'

    if not embeddings_dir.exists():
        print("警告：未找到 embeddings 目录，无法替换嵌入工作簿")
        return

    # 将最新 Excel 复制为 .xlsx，供图表外部数据引用
    target_xlsx = embeddings_dir / 'Workbook1.xlsx'
    shutil.copy2(excel_path, target_xlsx)

    # 如存在旧的 .xlsb，保留以防其他对象引用，但关系将切换到 .xlsx
    old_xlsb = embeddings_dir / 'Workbook1.xlsb'
    if not old_xlsb.exists():
        print("提示：模板未包含 Workbook1.xlsb，直接使用 .xlsx")

    # 更新图表关系，指向 .xlsx
    if charts_rels_dir.exists():
        for rel_file in charts_rels_dir.glob('chart*.xml.rels'):
            try:
                tree = etree.parse(str(rel_file))
                root = tree.getroot()
                for rel in root.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                    target = rel.get('Target')
                    if target and target.endswith('Workbook1.xlsb'):
                        rel.set('Target', '../embeddings/Workbook1.xlsx')
                tree.write(str(rel_file), xml_declaration=True, encoding='UTF-8', standalone='yes')
            except Exception as e:
                print(f"更新关系文件 {rel_file.name} 失败：{e}")

    # 更新 Content_Types，添加 .xlsx 默认类型
    try:
        ct = etree.parse(str(content_types_path))
        ct_root = ct.getroot()
        has_xlsx = False
        for d in ct_root.findall('{http://schemas.openxmlformats.org/package/2006/content-types}Default'):
            if d.get('Extension') == 'xlsx':
                has_xlsx = True
                break
        if not has_xlsx:
            new_def = etree.SubElement(ct_root, '{http://schemas.openxmlformats.org/package/2006/content-types}Default')
            new_def.set('Extension', 'xlsx')
            new_def.set('ContentType', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        ct.write(str(content_types_path), xml_declaration=True, encoding='UTF-8', standalone='yes')
    except Exception as e:
        print(f"更新 Content_Types 失败：{e}")

    # 设置 chart XML 的 externalData autoUpdate=1，让打开 PPT 自动刷新数据
    charts_dir = extract_dir / 'ppt' / 'charts'
    if charts_dir.exists():
        for chart_xml in charts_dir.glob('chart*.xml'):
            try:
                tree = etree.parse(str(chart_xml))
                root = tree.getroot()
                # 查找 externalData 节点
                ext_data = root.find('.//{http://schemas.openxmlformats.org/drawingml/2006/chart}externalData')
                if ext_data is not None:
                    auto = ext_data.find('{http://schemas.openxmlformats.org/drawingml/2006/chart}autoUpdate')
                    if auto is None:
                        auto = etree.SubElement(ext_data, '{http://schemas.openxmlformats.org/drawingml/2006/chart}autoUpdate')
                    auto.set('val', '1')
                    tree.write(str(chart_xml), xml_declaration=True, encoding='UTF-8', standalone='yes')
            except Exception as e:
                print(f"设置 {chart_xml.name} 自动更新失败：{e}")

def update_chart_series_formulas(extract_dir, chart_data, config):
    """
    将图表系列公式（c:ser -> c:cat/c:strRef/c:f、c:val/c:numRef/c:f、以及 c:tx/c:strRef/c:f）
    指向我们在 Sheet1 中的标准布局：
      - 分类：Sheet1!$A$2:$A${N}
      - 每个系列值：Sheet1!$<COL>$2:$<COL>${N}
      - 系列名：Sheet1!$<COL>$1
    这样在 PowerPoint 中“编辑数据”与图表绑定一致，体验更友好。
    """
    charts_dir = extract_dir / 'ppt' / 'charts'
    if not charts_dir.exists():
        print('未找到 charts 目录，跳过系列公式更新')
        return

    ns = {'c': 'http://schemas.openxmlformats.org/drawingml/2006/chart'}
    labels = chart_data['bar_chart']['labels']
    brands_display = config['filters']['brands_display']
    last_row = 1 + len(labels)

    # 目标显示顺序（从上到下），为了匹配 WPS/PowerPoint 的“最后系列在最上”的渲染规则，
    # 我们将 c:ser 的绑定列按反向顺序映射：ser[0] 绑定最底部品牌（Samsung），ser[最后] 绑定最顶部品牌（Lenovo）。
    brands_for_series = list(reversed(brands_display))

    def col_for_brand(brand):
        idx = brands_display.index(brand) if brand in brands_display else -1
        if idx < 0:
            return None
        # B -> 2，C -> 3 ...
        return openpyxl.utils.get_column_letter(idx + 2)

    def col_for_index(s_idx):
        # 序号 0 -> 列 B，1 -> 列 C ...
        return openpyxl.utils.get_column_letter(2 + s_idx)

    for chart_xml in charts_dir.glob('chart*.xml'):
        try:
            tree = etree.parse(str(chart_xml))
            root = tree.getroot()
            updated = False

            # barChart 系列
            for bar in root.findall('.//c:barChart', ns):
                for s_idx, ser in enumerate(bar.findall('c:ser', ns)):
                    # 系列名文本，用来确定列
                    brand_name = None
                    # 优先从 strCache 取值，其次回退到 c:v
                    tx_cache_v = ser.find('.//c:tx//c:strCache//c:pt//c:v', ns)
                    if tx_cache_v is not None and tx_cache_v.text:
                        brand_name = tx_cache_v.text.strip()
                    else:
                        tx_v = ser.find('.//c:tx//c:v', ns)
                        if tx_v is not None and tx_v.text:
                            brand_name = tx_v.text.strip()
                    # 分类公式
                    cat_f = ser.find('.//c:cat//c:strRef//c:f', ns)
                    if cat_f is None:
                        cat_f_parent = ser.find('.//c:cat//c:strRef', ns)
                        if cat_f_parent is None:
                            cat = ser.find('c:cat', ns)
                            if cat is None:
                                cat = etree.SubElement(ser, f"{{{ns['c']}}}cat")
                            str_ref = etree.SubElement(cat, f"{{{ns['c']}}}strRef")
                            cat_f_parent = str_ref
                        cat_f = etree.SubElement(cat_f_parent, f"{{{ns['c']}}}f")
                    cat_f.text = f"Sheet1!$A$2:$A${last_row}"

                    # 数值公式：按“从上到下”的反向品牌顺序绑定列，保证顶部 Lenovo、底部 Samsung
                    target_brand = brands_for_series[s_idx] if s_idx < len(brands_for_series) else None
                    col_letter = col_for_brand(target_brand) if target_brand and col_for_brand(target_brand) else col_for_index(s_idx)
                    val_f = ser.find('.//c:val//c:numRef//c:f', ns)
                    if val_f is None:
                        val = ser.find('c:val', ns)
                        if val is None:
                            val = etree.SubElement(ser, f"{{{ns['c']}}}val")
                        num_ref = val.find('c:numRef', ns)
                        if num_ref is None:
                            num_ref = etree.SubElement(val, f"{{{ns['c']}}}numRef")
                        val_f = etree.SubElement(num_ref, f"{{{ns['c']}}}f")
                    val_f.text = f"Sheet1!${col_letter}$2:${col_letter}${last_row}"

                    # 系列名公式（指向表头单元格）
                    tx_f = ser.find('.//c:tx//c:strRef//c:f', ns)
                    if tx_f is None:
                        tx = ser.find('c:tx', ns)
                        if tx is None:
                            tx = etree.SubElement(ser, f"{{{ns['c']}}}tx")
                        str_ref = tx.find('c:strRef', ns)
                        if str_ref is None:
                            str_ref = etree.SubElement(tx, f"{{{ns['c']}}}strRef")
                        tx_f = etree.SubElement(str_ref, f"{{{ns['c']}}}f")
                    # 系列名直接指向目标品牌所在的表头
                    tx_f.text = f"Sheet1!${col_letter}$1"
                    updated = True

            # 饼图系列（若存在则使用总 SOV）
            for pie in root.findall('.//c:pieChart', ns):
                ser = pie.find('c:ser', ns)
                if ser is not None:
                    cat_f = ser.find('.//c:cat//c:strRef//c:f', ns)
                    if cat_f is None:
                        cat = ser.find('c:cat', ns)
                        if cat is None:
                            cat = etree.SubElement(ser, f"{{{ns['c']}}}cat")
                        str_ref = cat.find('c:strRef', ns)
                        if str_ref is None:
                            str_ref = etree.SubElement(cat, f"{{{ns['c']}}}strRef")
                        cat_f = etree.SubElement(str_ref, f"{{{ns['c']}}}f")
                    cat_f.text = f"Sheet1!$A$2:$A${last_row}"

                    # 取第一个系列（Lenovo）示例，或按需要映射到总 SOV
                    val_f = ser.find('.//c:val//c:numRef//c:f', ns)
                    if val_f is None:
                        val = ser.find('c:val', ns)
                        if val is None:
                            val = etree.SubElement(ser, f"{{{ns['c']}}}val")
                        num_ref = val.find('c:numRef', ns)
                        if num_ref is None:
                            num_ref = etree.SubElement(val, f"{{{ns['c']}}}numRef")
                        val_f = etree.SubElement(num_ref, f"{{{ns['c']}}}f")
                    # 使用所有品牌之和不适用于饼图；此处保持默认或按需求外部数据定义
                    # 这里不强行重写饼图数据来源，仅保证分类公式存在
                    updated = True

            if updated:
                tree.write(str(chart_xml), xml_declaration=True, encoding='UTF-8', standalone='yes')
                print(f"已更新系列公式：{chart_xml.name}")
        except Exception as e:
            print(f"更新系列公式时出错 {chart_xml.name}：{e}")

def apply_brand_colors(extract_dir, config, excel_path):
    """
    为柱状图系列应用固定的品牌颜色映射，确保“颜色 ↔ 品牌 ↔ 数值列”一致。

    颜色映射（从模板反取，满足“图3取色”）：
      - Lenovo:  FF0000
      - Dell:    4E5AFF
      - Apple:   BEBEBE
      - HP:      43A7D0
      - ASUS:    E935A8
      - Acer:    0AA0A9
      - Samsung: 142D5C

    说明：
    - 覆盖所有 7 个品牌颜色；颜色读取优先使用配置 filters.brand_colors。
    - 系列颜色按系列名公式解析到的 Excel 表头品牌应用，确保颜色与数值列一致。
    - 使用 `a:solidFill/a:srgbClr@val` 直接设置 RGB，避免主题色偏移。
    """

    charts_dir = extract_dir / 'ppt' / 'charts'
    if not charts_dir.exists():
        print('未找到 charts 目录，跳过颜色应用')
        return

    ns = {
        'c': 'http://schemas.openxmlformats.org/drawingml/2006/chart',
        'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
    }

    brands_display = config['filters']['brands_display']
    # 优先使用配置文件中的颜色映射（filters.brand_colors）
    color_map = ((config.get('filters') or {}).get('brand_colors') or {
        'Lenovo':  'FF0000',
        'Dell':    '4E5AFF',
        'Apple':   'BEBEBE',
        'HP':      '43A7D0',
        'ASUS':    'E935A8',
        'Acer':    '0AA0A9',
        'Samsung': '142D5C',
    })

    # 预读 Excel，建立列字母 -> 品牌名映射（例如 B -> Lenovo）
    col_to_brand = {}
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb['Sheet1']
        col_index = 2
        while True:
            cell = ws.cell(row=1, column=col_index).value
            if cell is None:
                break
            col_to_brand[openpyxl.utils.get_column_letter(col_index)] = str(cell)
            col_index += 1
    except Exception:
        pass

    for chart_xml in charts_dir.glob('chart*.xml'):
        try:
            tree = etree.parse(str(chart_xml))
            root = tree.getroot()

            changed = False
            for bar in root.findall('.//c:barChart', ns):
                for s_idx, ser in enumerate(bar.findall('c:ser', ns)):
                    # 优先通过系列名公式解析列字母 -> Excel 表头的品牌名
                    tx_f = ser.find('.//c:tx//c:strRef//c:f', ns)
                    brand = None
                    if tx_f is not None and tx_f.text:
                        # 形如 Sheet1!$D$1
                        try:
                            addr = tx_f.text.split('!')[1]
                            col_letter = addr.split('$')[1]
                            brand = col_to_brand.get(col_letter)
                        except Exception:
                            brand = None
                    # 回退：用 brands_display 的序号
                    if not brand and s_idx < len(brands_display):
                        brand = brands_display[s_idx]
                    if not brand or brand not in color_map:
                        continue

                    # 获取/创建 spPr
                    sppr = ser.find('c:spPr', ns)
                    if sppr is None:
                        sppr = etree.SubElement(ser, f"{{{ns['c']}}}spPr")

                    # 清理现有填充设置
                    for e in list(sppr):
                        if e.tag in {f"{{{ns['a']}}}solidFill", f"{{{ns['a']}}}gradFill"}:
                            sppr.remove(e)

                    # 写入 solidFill srgbClr
                    solid = etree.SubElement(sppr, f"{{{ns['a']}}}solidFill")
                    srgb = etree.SubElement(solid, f"{{{ns['a']}}}srgbClr")
                    srgb.set('val', str(color_map[brand]).upper())
                    changed = True

            if changed:
                tree.write(str(chart_xml), xml_declaration=True, encoding='UTF-8', standalone='yes')
                print(f"已应用品牌颜色：{chart_xml.name}")
        except Exception as e:
            print(f"应用品牌颜色失败 {chart_xml.name}：{e}")

def remove_wps_chart_extensions(extract_dir):
    """
    移除 WPS 专用图表扩展（c:chart/c:extLst 下的 web.wps.cn/et 扩展），
    目的：避免在 WPS/PowerPoint 中点击“编辑数据”时，WPS 扩展重新套用内部样式导致系列颜色被覆盖。

    实现策略（MVP）：
    - 仅删除 URI 为 "{0b15fc19-7d7d-44ad-8c2d-2c3a37ce22c3}" 的 c:ext 及其包含的 https://web.wps.cn/et/2018/main 命名空间节点。
    - 若 extLst 清空，则连同父 extLst 一并移除。
    - 保留其他 Office 扩展，避免破坏非相关功能。
    """
    charts_dir = extract_dir / 'ppt' / 'charts'
    if not charts_dir.exists():
        return

    ns_c = 'http://schemas.openxmlformats.org/drawingml/2006/chart'
    wps_ns = 'https://web.wps.cn/et/2018/main'

    for chart_xml in charts_dir.glob('chart*.xml'):
        try:
            tree = etree.parse(str(chart_xml))
            root = tree.getroot()
            chart = root.find(f'.//{{{ns_c}}}chart')
            if chart is None:
                continue
            extLst = chart.find(f'{{{ns_c}}}extLst')
            if extLst is None:
                continue

            changed = False
            for ext in list(extLst):
                uri = ext.get('uri')
                # 匹配 WPS 扩展（GUID 或命名空间判断）
                is_wps = (uri == '{0b15fc19-7d7d-44ad-8c2d-2c3a37ce22c3}') or any(
                    child.tag.startswith(f'{{{wps_ns}}}') for child in ext
                )
                if is_wps:
                    extLst.remove(ext)
                    changed = True

            # 若 extLst 无子节点则移除整个 extLst
            if changed and len(list(extLst)) == 0:
                chart.remove(extLst)

            if changed:
                tree.write(str(chart_xml), xml_declaration=True, encoding='UTF-8', standalone='yes')
                print(f"已移除 WPS 扩展：{chart_xml.name}")
        except Exception as e:
            print(f"移除 WPS 扩展失败 {chart_xml.name}：{e}")

def repack_ppt(extract_dir, output_path):
    """重新打包PPT文件"""
    print(f"正在重新打包PPT到：{output_path}")
    
    # 确保输出目录存在
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # 创建新的ZIP文件
    with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zip_ref:
        for file_path in extract_dir.rglob('*'):
            if file_path.is_file():
                # 计算相对路径
                arcname = file_path.relative_to(extract_dir)
                zip_ref.write(file_path, arcname)
    
    print(f"PPT文件已生成：{output_path}")

def main():
    """主函数"""
    try:
        print("开始从Excel数据生成PPT...")
        
        # 清理之前的临时文件
        temp_dir = ROOT / 'tmp'
        if temp_dir.exists():
            shutil.rmtree(temp_dir)
            print("已清理之前的临时文件")
        
        # 加载配置
        config = load_config()
        
        # 检查Excel文件
        excel_path = ROOT / 'p29_data.xlsx'
        if not excel_path.exists():
            print(f"❌ Excel数据文件不存在：{excel_path}")
            print("请先运行 generate_excel.py 生成Excel数据文件")
            return 1
        
        # 读取Excel数据
        print("正在读取Excel数据...")
        sov_data, pie_data = read_excel_data(excel_path)
        
        # 准备图表数据（柱状图优先从 Sheet1 读取）
        print("正在准备图表数据...")
        chart_data = prepare_chart_data(sov_data, pie_data, config, excel_path)
        
        # 更新JSON数据文件
        print("正在更新图表数据文件...")
        update_chart_data_json(chart_data)
        
        # 解压PPT模板
        print("正在解压PPT模板...")
        extract_dir = extract_ppt_template()

        # 校验 Sheet1 布局（Sheet1 由 generate_excel.py 生成）
        print("正在校验 Excel 的 Sheet1 布局...")
        assert_sheet1_layout(excel_path, config)

        # 替换嵌入的工作簿并修正关系，启用自动更新
        print("正在替换嵌入工作簿并启用自动刷新...")
        update_embedded_excel_and_links(extract_dir, excel_path)

        # 更新系列公式指向 Sheet1 的命名布局
        print("正在更新图表系列公式绑定...")
        update_chart_series_formulas(extract_dir, chart_data, config)

        # 应用品牌颜色映射，保证颜色与品牌和值一致
        print("正在应用品牌颜色映射...")
        apply_brand_colors(extract_dir, config, excel_path)

        # 先刷新 chart XML 缓存，确保系列名（c:tx/strCache）可用
        print("正在刷新图表缓存数据...")
        update_chart_xml_caches(extract_dir, chart_data)

        # 规范柱状图数据标签，统一 bestFit 与百分号格式，并清理模板的删除项
        print("正在规范柱状图数据标签样式...")
        normalize_bar_chart_data_labels(extract_dir, chart_data, config)

        # 填充右侧总体 SOV 文本标签（整数百分比）
        print("正在填充右侧总体SOV文本标签...")
        update_right_total_sov_texts(extract_dir, pie_data, config)

        # 移除 WPS 图表扩展，防止“编辑数据”时颜色被重置
        print("正在移除 WPS 图表扩展...")
        remove_wps_chart_extensions(extract_dir)
        
        # 重新打包PPT
        output_dir = ROOT / 'output'
        output_path = output_dir / 'p29-final.pptx'
        repack_ppt(extract_dir, output_path)
        
        print(f"\n✅ PPT文件生成成功：{output_path}")
        print("\n📊 生成的PPT包含：")
        print("  - 左侧：各渠道品牌声量份额堆叠柱状图")
        print("  - 右侧：品牌总体声量份额饼图")
        print("  - 嵌入的Excel数据已更新，可在PowerPoint中编辑")
        
        # 保留临时文件用于调试（页面级 tmp 下）
        if TMP_DIR.exists():
            print(f"临时文件已保留在：{TMP_DIR}")
            print("  - chart_data.json: 图表数据")
            print("  - ppt_extracted/: 解压的PPT内容")
        
        return 0
        
    except Exception as e:
        print(f"❌ 生成PPT时发生错误：{e}")
        import traceback
        traceback.print_exc()
        return 1

if __name__ == '__main__':
    exit(main())