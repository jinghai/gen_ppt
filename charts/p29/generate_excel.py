#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
P29页面Excel数据生成器
根据数据库数据生成Excel文件，供人工检验和修改
"""

import sqlite3
import pandas as pd
import yaml
from pathlib import Path
from datetime import datetime, timedelta, timezone
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl
import math

# 项目路径配置
ROOT = Path(__file__).resolve().parent

def load_config():
    """加载页面级配置文件"""
    config_path = ROOT / 'config.yaml'
    with open(config_path, 'r', encoding='utf-8') as f:
        config = yaml.safe_load(f)
    return config

def get_database_paths(config):
    """获取数据库路径（相对于P29目录）"""
    neticle_db = ROOT / config['data_sources']['neticle_db']
    metrics_db = ROOT / config['data_sources']['metrics_db']
    return neticle_db, metrics_db

def to_utc_ms(date_str):
    """将日期字符串转换为UTC毫秒时间戳"""
    dt = datetime.strptime(date_str, '%Y-%m-%d').replace(tzinfo=timezone.utc)
    return int(dt.timestamp() * 1000)

def extract_channel_data(config):
    """从neticle数据库提取渠道数据"""
    neticle_db, _ = get_database_paths(config)
    
    # 配置参数
    country_id = config['filters']['countryId']
    start_date = config['update']['start_date']
    end_date = config['update']['end_date']
    brands = config['filters']['brands']
    # 使用配置中的展示名做标准化映射，确保后续聚合与展示一致
    brands_display = config['filters']['brands_display']
    # 构建品牌归一化映射：如 'hp' -> 'HP', 'asus' -> 'ASUS'
    brand_norm_map = {b.lower(): d for b, d in zip(brands, brands_display)}
    # 渠道映射改为严格使用 channel_map；缺失则抛错，避免兜底掩盖错误
    if 'channel_map' not in config:
        raise KeyError('配置缺少 channel_map，请在 charts/p29/config.yaml 中定义渠道映射')
    channel_mapping = config['channel_map']
    
    # 时间范围转换
    start_ms = to_utc_ms(start_date)
    end_ms = to_utc_ms((datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)).strftime('%Y-%m-%d'))
    
    # 创建渠道映射字典
    source_to_channel = {}
    for channel, sources in channel_mapping.items():
        for source in sources:
            source_to_channel[str(source).lower()] = channel
    
    # 查询数据
    with sqlite3.connect(neticle_db) as conn:
        # 获取所有相关数据
        sql = """
        SELECT 
            m.sourceName,
            m.keyword_label,
            COUNT(*) as mention_count,
            SUM(m.sumInteractions) as total_interactions,
            AVG(m.polarity) as avg_sentiment
        FROM mentions_wide m
        WHERE m.countryId = ?
          AND m.createdAtUtcMs >= ? 
          AND m.createdAtUtcMs < ?
          AND m.keyword_label IS NOT NULL
        GROUP BY m.sourceName, m.keyword_label
        ORDER BY m.sourceName, m.keyword_label
        """
        
        df = pd.read_sql_query(sql, conn, params=[country_id, start_ms, end_ms])
    
    # 数据处理
    processed_data = []
    
    for _, row in df.iterrows():
        source_name = row['sourceName'].lower() if row['sourceName'] else ''
        channel = source_to_channel.get(source_name, 'Other')
        
        # 提取并归一化品牌名称
        keyword_label = row['keyword_label']
        brand = None
        for b in brands:
            if keyword_label and b.lower() in keyword_label.lower():
                # 使用映射后的展示名，避免大小写不一致造成聚合漏算
                brand = brand_norm_map.get(b.lower(), b.title())
                break
        
        if brand and channel != 'Other':
            processed_data.append({
                'Channel': channel,
                'Brand': brand,
                'Source': row['sourceName'],
                'Mention_Count': row['mention_count'],
                'Total_Interactions': row['total_interactions'] or 0,
                'Avg_Sentiment': round(row['avg_sentiment'] or 0, 2)
            })
    
    return pd.DataFrame(processed_data)

def calculate_channel_sov(df, config):
    """计算各渠道内的品牌声量份额"""
    brands_display = config['filters']['brands_display']
    channels = config['fill_policy']['channel_order']
    
    # 按渠道和品牌聚合数据
    channel_brand_data = df.groupby(['Channel', 'Brand'])['Mention_Count'].sum().reset_index()
    
    # 计算每个渠道内的SOV
    sov_data = []
    
    for channel in channels:
        channel_data = channel_brand_data[channel_brand_data['Channel'] == channel]
        total_mentions = channel_data['Mention_Count'].sum()
        
        for brand in brands_display:
            brand_mentions = channel_data[channel_data['Brand'] == brand]['Mention_Count'].sum()
            sov = round((brand_mentions / total_mentions * 100), 1) if total_mentions > 0 else 0.0
            
            sov_data.append({
                'Channel': channel,
                'Brand': brand,
                'Mentions': brand_mentions,
                'Total_Channel_Mentions': total_mentions,
                'SOV_Percentage': sov
            })
    
    return pd.DataFrame(sov_data)

def calculate_brand_total_sov(df, config):
    """计算品牌总体声量份额（饼图数据）"""
    brands_display = config['filters']['brands_display']

    # 按品牌聚合总提及数
    brand_totals = df.groupby('Brand')['Mention_Count'].sum().reset_index()
    total_mentions = brand_totals['Mention_Count'].sum()
    
    # 使用最大余数法（Largest Remainder Method）将百分比整数化并保证总和为100。
    # 说明：相比逐个四舍五入，该方法先按精确配额取地板值，再按余数大小分配剩余点数，
    #       可严格保证总和为100，避免出现超过/低于100的情况。
    def allocate_percentages_lrm(weights, target=100):
        """
        最大余数法分配：给定非负权重列表，分配整数百分比使总和严格等于 target。
        - weights: 各品牌的权重（这里使用提及数）
        - target: 目标总和（默认100）
        返回：与 weights 同长度的整数列表，和为 target（若总权重为0则全为0）。
        """
        total = float(sum(weights))
        if total <= 0 or target <= 0:
            return [0] * len(weights)

        quotas = [target * (w / total) for w in weights]
        floors = [int(math.floor(q)) for q in quotas]
        remainder = [q - f for q, f in zip(quotas, floors)]
        base_sum = sum(floors)
        need = target - base_sum

        if need > 0:
            # 余数从大到小分配剩余点数；稳定排序保证品牌顺序可重复。
            order = sorted(range(len(weights)), key=lambda i: (-remainder[i], i))
            for i in order[:need]:
                floors[i] += 1
        elif need < 0:
            # 极罕见的浮点异常保护：如果地板和超过目标，则按余数从小到大回收。
            order = sorted(range(len(weights)), key=lambda i: (remainder[i], i))
            for i in order[:abs(need)]:
                # 避免负数，最少为0
                floors[i] = max(0, floors[i] - 1)

        # 最终校验；若仍有偏差（理论不应发生），按顺序微调到匹配目标。
        diff = target - sum(floors)
        if diff != 0 and len(floors) > 0:
            seq = list(range(len(floors)))
            if diff > 0:
                for i in seq[:diff]:
                    floors[i] += 1
            else:
                for i in seq[:abs(diff)]:
                    floors[i] = max(0, floors[i] - 1)
        return floors

    # 构造权重（按展示名顺序），并执行最大余数法分配
    weights = [int(brand_totals[brand_totals['Brand'] == b]['Mention_Count'].sum()) for b in brands_display]
    allocations = allocate_percentages_lrm(weights, target=100)

    # 组装输出数据框
    pie_rows = []
    for b, w, p in zip(brands_display, weights, allocations):
        pie_rows.append({
            'Brand': b,
            'Total_Mentions': w,
            'Percentage': int(p)
        })

    return pd.DataFrame(pie_rows)

def style_worksheet(ws, title):
    """设置工作表样式"""
    # 标题样式
    title_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    title_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    # 表头样式
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    
    # 数据样式
    data_font = Font(name='Arial', size=10)
    
    # 边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 设置标题
    ws['A1'] = title
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 合并标题单元格
    if ws.max_column > 1:
        ws.merge_cells(f'A1:{chr(64 + ws.max_column)}1')
    
    # 设置表头样式
    for cell in ws[2]:
        if cell.value:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
    
    # 设置数据样式
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        for cell in row:
            if cell.value is not None:
                cell.font = data_font
                cell.border = thin_border
                if isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # 自动调整列宽
    for col_num in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(col_num)
        
        for row_num in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_num, column=col_num)
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width

def build_sheet1(workbook, sov_data: pd.DataFrame, config):
    """
    构建标准 Sheet1：
    - 第1行：B.. 为品牌名；A1 留空
    - 第2..N行：A列为渠道名；B.. 为对应品牌在该渠道的数值（整数百分比）

    说明：此 Sheet1 专供 PowerPoint 图表的系列绑定使用，保持最小必要结构。
    """
    brands_display = config['filters']['brands_display']
    channels = config['fill_policy']['channel_order']

    # 若已有 Sheet1，先移除以保证结构一致
    if 'Sheet1' in workbook.sheetnames:
        ws_old = workbook['Sheet1']
        workbook.remove(ws_old)
    ws = workbook.create_sheet('Sheet1', 0)

    # 表头：品牌名（B..）
    ws.cell(row=1, column=1, value=None)
    for col_idx, brand in enumerate(brands_display, start=2):
        ws.cell(row=1, column=col_idx, value=brand)

    # 采用最大余数法为每行分配整数百分比，使行总和严格为100。
    # 为避免浮点传播误差，权重直接使用提及数 Mentions。
    def allocate_percentages_lrm(weights, target=100):
        """
        行级百分比分配（最大余数法）。
        - weights: 当前渠道下各品牌的提及数列表
        - target: 目标总和（固定为100）
        返回：与 weights 同长度的整数列表；若总权重为0则全为0（总和=0）。
        注：当某渠道总提及为0时，保持全0以如实反映无数据，不强制造100。
        """
        total = float(sum(weights))
        if total <= 0:
            return [0] * len(weights)
        quotas = [target * (w / total) for w in weights]
        floors = [int(math.floor(q)) for q in quotas]
        remainder = [q - f for q, f in zip(quotas, floors)]
        base_sum = sum(floors)
        need = target - base_sum
        if need > 0:
            order = sorted(range(len(weights)), key=lambda i: (-remainder[i], i))
            for i in order[:need]:
                floors[i] += 1
        elif need < 0:
            order = sorted(range(len(weights)), key=lambda i: (remainder[i], i))
            for i in order[:abs(need)]:
                floors[i] = max(0, floors[i] - 1)
        return floors

    # 填充各渠道数据（整数，行和=100 或 0（无数据））
    for r_idx, channel in enumerate(channels, start=2):
        ws.cell(row=r_idx, column=1, value=channel)

        # 收集该渠道下各品牌的权重（提及数）
        channel_rows = sov_data[sov_data['Channel'] == channel]
        weights = []
        for brand in brands_display:
            row_sel = channel_rows[channel_rows['Brand'] == brand]
            w = int(row_sel['Mentions'].iloc[0]) if not row_sel.empty else 0
            weights.append(w)

        allocations = allocate_percentages_lrm(weights, target=100)

        for c_idx, val in enumerate(allocations, start=2):
            ws.cell(row=r_idx, column=c_idx, value=int(val))

    # 简单对齐与列宽（避免过度设计）
    for col_num in range(1, ws.max_column + 1):
        column_letter = openpyxl.utils.get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 14

    return ws

def create_excel_file(config):
    """创建Excel文件"""
    print("正在提取数据...")
    
    # 提取原始数据
    raw_data = extract_channel_data(config)
    
    if raw_data.empty:
        print("警告：未找到符合条件的数据")
        return
    
    print(f"提取到 {len(raw_data)} 条原始数据记录")
    
    # 计算SOV数据
    sov_data = calculate_channel_sov(raw_data, config)
    pie_data = calculate_brand_total_sov(raw_data, config)
    
    # 创建Excel文件
    output_path = ROOT / 'p29_data.xlsx'
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # 写入各个工作表
        raw_data.to_excel(writer, sheet_name='原始数据', index=False, startrow=1)
        sov_data.to_excel(writer, sheet_name='渠道SOV数据', index=False, startrow=1)
        # 品牌总体SOV百分比保留整数
        pie_data_int = pie_data.copy()
        if 'Percentage' in pie_data_int.columns:
            pie_data_int['Percentage'] = pie_data_int['Percentage'].astype(int)
        pie_data_int.to_excel(writer, sheet_name='品牌总体SOV', index=False, startrow=1)
        
        # 获取工作簿对象
        workbook = writer.book
        
        # 设置样式
        style_worksheet(workbook['原始数据'], 'P29 - 原始提及数据')
        style_worksheet(workbook['渠道SOV数据'], 'P29 - 各渠道品牌声量份额')
        style_worksheet(workbook['品牌总体SOV'], 'P29 - 品牌总体声量份额（饼图）')

        # 生成 Sheet1（供 PPT 图表绑定使用），按“列为品牌、行为渠道”的布局，数据为整数
        build_sheet1(workbook, sov_data, config)
    
    print(f"Excel文件已生成：{output_path}")
    print(f"- 原始数据：{len(raw_data)} 条记录")
    print(f"- 渠道SOV数据：{len(sov_data)} 条记录")
    print(f"- 品牌总体SOV：{len(pie_data)} 条记录")
    
    return output_path

def main():
    """主函数"""
    try:
        print("开始生成P29页面Excel数据文件...")
        
        # 加载配置
        config = load_config()
        
        # 创建Excel文件
        excel_path = create_excel_file(config)
        
        if excel_path:
            print(f"\n✅ Excel数据文件生成成功：{excel_path}")
            print("\n📋 文件包含以下工作表：")
            print("  1. 原始数据 - 从数据库提取的原始提及数据")
            print("  2. 渠道SOV数据 - 各渠道内品牌声量份额（用于左侧堆叠柱状图）")
            print("  3. 品牌总体SOV - 品牌总体声量份额（用于右侧饼图）")
            print("\n💡 您可以在Excel中检验和修改数据，然后使用fill_from_excel.py脚本将数据填充到PPT中")
        else:
            print("❌ Excel文件生成失败")
            return 1
            
    except Exception as e:
        print(f"❌ 生成Excel文件时发生错误：{e}")
        import traceback
        traceback.print_exc()
        return 1
    
    return 0

if __name__ == '__main__':
    exit(main())