#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
P13 页面：Engagement breakdown 数据生成脚本（简化版）

职责（按用户要求简化）：
- 读取页面级配置 config.yaml（仅引用本页配置，不依赖全局配置）；
- 从 Neticle 宽表中按国家与关键词过滤数据；
- 计算“参与度”（Engagement）= 行级候选互动字段求和（如 likes/shares/comments 及平台变体）；
- 基于参与度进行百分比归一化：
  - 饼图：各情绪的参与度占整体参与度的百分比；
  - 折线：每一天各情绪的参与度占当日总参与度的百分比；
- 不再计算或输出曝光（impressions）与 IPM；
- 将结果写入同一个 Excel：PieData（百分比）、LineData（百分比）、MetaData（元信息）。

备注（严格报错，不兜底）：
- 自动探测 mentions_wide 表与字段，若候选互动字段一个都不存在，直接报错；
- 运行时仅按过滤条件读取必要列；
- 所有临时文件放置在页面级 tmp 目录。
"""

import os
import sys
import json
import math
import sqlite3
import datetime as dt
from typing import Dict, List, Tuple, Optional

try:
    import yaml
except Exception as e:
    print("缺少依赖：pyyaml。请安装后重试。")
    raise

try:
    from openpyxl import Workbook
except Exception:
    print("缺少依赖：openpyxl。请安装后重试。")
    raise


class P13DataGenerator:
    def __init__(self, config_path: str):
        with open(config_path, 'r', encoding='utf-8') as f:
            self.cfg = yaml.safe_load(f)

        self.page_root = os.path.dirname(os.path.abspath(config_path))
        self.project = self.cfg.get('project', {})
        self.data_sources = self.cfg.get('data_sources', {})
        self.time_range = self.cfg.get('time_range', {})
        self.filters = self.cfg.get('filters', {})
        self.sentiment = self.cfg.get('sentiment', {})
        self.charts = self.cfg.get('charts', {})
        self.engagement_cfg = self.cfg.get('engagement', {})
        self.output_cfg = self.cfg.get('output', {})
        self.fill_policy = self.cfg.get('fill_policy', {})
        self.logging_cfg = self.cfg.get('logging', {})

        # 路径
        self.tmp_dir = os.path.join(self.page_root, self.project.get('tmp_dir', 'tmp'))
        # 输出固定到页面目录，避免 output 子目录；要求配置提供 excel_file_name
        if 'excel_file_name' not in self.output_cfg:
            raise KeyError('config.output.excel_file_name 未配置')
        self.excel_file_name = self.output_cfg['excel_file_name']
        self.excel_file_path = os.path.join(self.page_root, self.excel_file_name)

        # 颜色与阈值
        self.labels = self.sentiment.get('labels', ['Positive', 'Neutral', 'Negative'])
        thresholds = self.sentiment.get('thresholds', {})
        self.pos_min = float(thresholds.get('positive_min', 0.05))
        self.neg_max = float(thresholds.get('negative_max', -0.05))

        # 互动字段候选（仅用于计算 Engagement；不再使用曝光/不计算 IPM）
        self.eng_fields_candidates: List[str] = self.engagement_cfg.get('fields_candidates', [])

        # 运行时信息
        self.wide_table_name: Optional[str] = None
        self.wide_columns: List[str] = []
        # 审计附加列（动态匹配宽表候选列）
        self.audit_post_id_col: Optional[str] = None
        self.audit_channel_col: Optional[str] = None

    # ---------- 基础工具 ----------
    def _ensure_dirs(self):
        # 仅确保 tmp 目录存在；不再创建 output 子目录
        os.makedirs(self.tmp_dir, exist_ok=True)

    def _to_date(self, s) -> dt.date:
        """将配置中的日期值统一为 date。
        - 支持 `datetime.date`/`datetime.datetime` 直接返回/取 date；
        - 支持字符串格式 `YYYY-MM-DD`；
        - 其余情况抛出 TypeError 以便尽早暴露配置问题。
        """
        if s is None:
            raise TypeError('start_date/end_date 未设置')
        if isinstance(s, dt.date):
            return s
        if isinstance(s, dt.datetime):
            return s.date()
        if isinstance(s, str):
            return dt.datetime.strptime(s.strip(), '%Y-%m-%d').date()
        raise TypeError(f'不支持的日期类型: {type(s)}')

    def _day_range(self) -> List[dt.date]:
        start = self._to_date(self.time_range.get('start_date'))
        end = self._to_date(self.time_range.get('end_date'))
        days = []
        cur = start
        while cur <= end:
            days.append(cur)
            cur += dt.timedelta(days=1)
        return days

    def _date_to_utc_ms_bounds(self, d: dt.date) -> Tuple[int, int]:
        start_dt = dt.datetime(d.year, d.month, d.day, 0, 0, 0)
        end_dt = start_dt + dt.timedelta(days=1) - dt.timedelta(milliseconds=1)
        epoch = dt.datetime(1970, 1, 1)
        start_ms = int((start_dt - epoch).total_seconds() * 1000)
        end_ms = int((end_dt - epoch).total_seconds() * 1000)
        return start_ms, end_ms

    # ---------- 数据库探测 ----------
    def _connect_neticle(self) -> sqlite3.Connection:
        db_path = os.path.join(self.page_root, self.data_sources.get('neticle_db'))
        if not os.path.isabs(db_path):
            db_path = os.path.abspath(db_path)
        if not os.path.exists(db_path):
            raise FileNotFoundError(f"未找到 Neticle DB: {db_path}")
        return sqlite3.connect(db_path)

    def _detect_wide_table(self, con: sqlite3.Connection) -> str:
        cur = con.cursor()
        cur.execute("SELECT name FROM sqlite_master WHERE type='table'")
        names = [r[0] for r in cur.fetchall()]
        candidates = [n for n in names if 'mentions_wide' in n.lower()]
        if not candidates:
            raise RuntimeError("未找到 mentions_wide 表，请检查数据库结构或更新配置。")
        # 优先 neticle_mentions_wide
        for n in candidates:
            if 'neticle' in n.lower():
                return n
        return candidates[0]

    def _load_columns(self, con: sqlite3.Connection, table: str) -> List[str]:
        cur = con.cursor()
        cur.execute(f"PRAGMA table_info('{table}')")
        cols = [r[1] for r in cur.fetchall()]
        return cols

    # ---------- 动态构造查询 ----------
    def _pick_first_existing(self, candidates: List[str]) -> Optional[str]:
        for c in candidates:
            if c in self.wide_columns:
                return c
        return None

    def _build_where(self) -> Tuple[str, List]:
        params: List = []

        # 国家过滤字段候选
        country_col = self._pick_first_existing(['countryId', 'country_id', 'country'])
        where_clauses = []
        if country_col:
            where_clauses.append(f"{country_col} = ?")
            params.append(int(self.filters.get('country_id')))

        # 关键词过滤（品牌）字段候选
        keyword_col = self._pick_first_existing(['keyword', 'keyword_label', 'keywordLabel'])
        if keyword_col and self.filters.get('keyword_like'):
            where_clauses.append(f"{keyword_col} LIKE ?")
            params.append(self.filters.get('keyword_like'))

        # 时间范围
        created_col = self._pick_first_existing(['createdAtUtcMs', 'createdAtMs', 'createdAt'])
        if created_col:
            # 整体区间过滤，减少 IO
            start_d = self._to_date(self.time_range.get('start_date'))
            end_d = self._to_date(self.time_range.get('end_date'))
            epoch = dt.datetime(1970, 1, 1)
            start_ms = int((dt.datetime.combine(start_d, dt.time.min) - epoch).total_seconds() * 1000)
            end_ms = int((dt.datetime.combine(end_d, dt.time.max) - epoch).total_seconds() * 1000)
            where_clauses.append(f"{created_col} BETWEEN ? AND ?")
            params.extend([start_ms, end_ms])

        where_sql = ''
        if where_clauses:
            where_sql = 'WHERE ' + ' AND '.join(where_clauses)
        return where_sql, params

    def _select_columns(self) -> List[str]:
        cols = []
        # 必要列
        for c in ['polarity', 'createdAtUtcMs', 'createdAtMs', 'createdAt']:
            if c in self.wide_columns:
                cols.append(c)

        # 互动候选列（存在才加入）
        for c in self.eng_fields_candidates:
            if c in self.wide_columns:
                cols.append(c)

        # 国家与关键词列（便于调试与后续校验）
        for c in ['countryId', 'country_id', 'country', 'keyword', 'keyword_label', 'keywordLabel']:
            if c in self.wide_columns:
                cols.append(c)

        # 审计附加列（若存在则选取）
        if self.audit_post_id_col and self.audit_post_id_col in self.wide_columns:
            cols.append(self.audit_post_id_col)
        if self.audit_channel_col and self.audit_channel_col in self.wide_columns:
            cols.append(self.audit_channel_col)

        # 去重保持顺序
        dedup = []
        seen = set()
        for c in cols:
            if c not in seen:
                dedup.append(c)
                seen.add(c)
        return dedup

    # ---------- 计算逻辑 ----------
    def _classify_sentiment(self, polarity: Optional[float]) -> str:
        if polarity is None:
            return 'Neutral'
        try:
            p = float(polarity)
        except Exception:
            return 'Neutral'
        if p > self.pos_min:
            return 'Positive'
        if p < self.neg_max:
            return 'Negative'
        return 'Neutral'

    def _calc_row_engagement(self, row: Dict[str, Optional[float]]) -> float:
        """
        计算单行参与度：将配置中的候选互动字段求和。
        - 若某字段不存在或值非数值，忽略该字段；
        - 全部候选字段均不存在时，上层会统一进行严格校验并报错。
        """
        total = 0.0
        for c in self.eng_fields_candidates:
            v = row.get(c)
            if v is None:
                continue
            try:
                total += float(v)
            except Exception:
                # 数据类型异常时跳过该字段，避免污染总和
                continue
        return total

    def _pick_created_ms(self, row: Dict) -> Optional[int]:
        for c in ['createdAtUtcMs', 'createdAtMs', 'createdAt']:
            v = row.get(c)
            if v is None:
                continue
            try:
                return int(v)
            except Exception:
                pass
        return None

    def _ms_to_date(self, ms: int) -> dt.date:
        return dt.datetime.utcfromtimestamp(ms / 1000.0).date()

    # ---------- 主流程 ----------
    def run(self) -> None:
        self._ensure_dirs()

        # 探测表和字段
        con = self._connect_neticle()
        try:
            self.wide_table_name = self._detect_wide_table(con)
            self.wide_columns = self._load_columns(con, self.wide_table_name)

            # 动态匹配审计附加列（仅用于展示，不参与计算；存在则纳入查询）
            self.audit_post_id_col = self._pick_first_existing([
                'postId', 'post_id', 'mentionId', 'id', 'postID'
            ])
            self.audit_channel_col = self._pick_first_existing([
                'channel', 'source', 'platform', 'network', 'site'
            ])

            # 严格校验：至少存在一个互动候选字段，否则直接报错，避免退回“计数占比”造成误导
            if not any(c in self.wide_columns for c in self.eng_fields_candidates):
                raise RuntimeError(
                    f"互动候选字段不存在：{self.eng_fields_candidates}。请检查数据库列名或更新配置。")

            where_sql, params = self._build_where()
            select_cols = self._select_columns()
            if not select_cols:
                raise RuntimeError("没有可选取的列，无法生成数据。")

            sql = f"SELECT {', '.join(select_cols)} FROM {self.wide_table_name} {where_sql}"
            cur = con.cursor()
            cur.execute(sql, params)

            # 聚合容器（按参与度进行聚合与归一化）
            pie_eng_sums = {lbl: 0.0 for lbl in self.labels}  # 饼图：各情绪参与度总和
            day_eng_buckets: Dict[dt.date, Dict[str, float]] = {}  # 折线：每日各情绪参与度
            # 审计示例行（最多 200 行，按时间顺序采样）
            audit_row_max = 200
            audit_rows: List[Dict] = []

            for row_t in cur.fetchall():
                row = {select_cols[i]: row_t[i] for i in range(len(select_cols))}

                # 时间
                ms = self._pick_created_ms(row)
                if ms is None:
                    continue
                d = self._ms_to_date(ms)

                # 过滤到配置区间（保险）
                start_d = self._to_date(self.time_range.get('start_date'))
                end_d = self._to_date(self.time_range.get('end_date'))
                if not (start_d <= d <= end_d):
                    continue

                # 情绪分类
                sentiment = self._classify_sentiment(row.get('polarity'))
                # 行级参与度
                eng = self._calc_row_engagement(row)
                # 饼图累计（按情绪累加参与度）
                pie_eng_sums[sentiment] = pie_eng_sums.get(sentiment, 0.0) + eng
                # 折线桶累计（每日情绪参与度）
                if d not in day_eng_buckets:
                    day_eng_buckets[d] = {lbl: 0.0 for lbl in self.labels}
                day_eng_buckets[d][sentiment] += eng

                # 审计示例行：保留前 audit_row_max 行（含关键原始列与派生值）
                if len(audit_rows) < audit_row_max:
                    # 收集候选互动列的原始值
                    cand_vals = {c: row.get(c) for c in self.eng_fields_candidates}
                    # 审计附加列值（若存在）
                    post_id_val = row.get(self.audit_post_id_col) if self.audit_post_id_col else None
                    channel_val = row.get(self.audit_channel_col) if self.audit_channel_col else None
                    audit_rows.append({
                        'Date': d.strftime('%Y-%m-%d'),
                        'Polarity': row.get('polarity'),
                        'Sentiment': sentiment,
                        'PostId': post_id_val,
                        'Channel': channel_val,
                        **cand_vals,
                        'RowEngagement': eng,
                    })

        finally:
            con.close()

        # 生成 Excel
        wb = Workbook()
        ws_pie = wb.active
        ws_pie.title = self.charts.get('pie', {}).get('sheet_name', 'PieData')
        ws_line = wb.create_sheet(self.charts.get('line', {}).get('sheet_name', 'LineData'))
        ws_meta = wb.create_sheet('MetaData')
        # 审计工作表
        ws_audit_day = wb.create_sheet('AuditDay')
        ws_audit_row = wb.create_sheet('AuditRow')
        ws_audit_month = wb.create_sheet('AuditMonth')

        # PieData：按总体参与度计算百分比（严格归一化）
        total_engagement = sum(pie_eng_sums.values())
        ws_pie.append(['Sentiment', 'Percentage'])
        for lbl in self.labels:
            part = pie_eng_sums.get(lbl, 0.0)
            pct = (part / total_engagement * 100.0) if total_engagement > 0 else 0.0
            ws_pie.append([lbl, round(pct, 2)])

        # LineData：改为“全局最大值归一化（IndexGlobal）”
        # 中文说明：
        #   口径与附件示例一致，纵轴为指数(0-100)，而非“当日占比”。
        #   公式：index(lbl, day) = 100 * eng(lbl, day) / max_{d,l}(eng(l, d))
        #   特性：各日三条线不再相加为100；最大峰值约为100。
        #   严格报错与保护：若全局最大值<=0，则指数全记为0。
        ws_line.append(['Date', 'Positive', 'Neutral', 'Negative'])
        # 计算全局最大值
        all_values = []
        for b in day_eng_buckets.values():
            all_values.extend(list(b.values()))
        global_max = max(all_values) if all_values else 0.0
        for d in sorted(day_eng_buckets.keys()):
            bucket = day_eng_buckets[d]
            pos_val = bucket.get('Positive', 0.0)
            neu_val = bucket.get('Neutral', 0.0)
            neg_val = bucket.get('Negative', 0.0)
            if global_max > 0:
                pos_idx = 100.0 * pos_val / global_max
                neu_idx = 100.0 * neu_val / global_max
                neg_idx = 100.0 * neg_val / global_max
            else:
                pos_idx = neu_idx = neg_idx = 0.0

            ws_line.append([
                d.strftime('%Y-%m-%d'),
                round(pos_idx, 2),
                round(neu_idx, 2),
                round(neg_idx, 2),
            ])

        # MetaData：写入基础信息
        meta_pairs = [
            ('Page', self.project.get('page')),
            ('Country', self.filters.get('country_name')),
            ('CountryId', self.filters.get('country_id')),
            ('BrandKey', self.filters.get('brand_key')),
            ('KeywordLike', self.filters.get('keyword_like')),
            ('StartDate', self.time_range.get('start_date')),
            ('EndDate', self.time_range.get('end_date')),
            ('TotalEngagement', total_engagement),
            ('LineNormalization', 'IndexGlobal'),
            ('WideTableName', self.wide_table_name),
            ('SelectedColumns', ','.join(self._select_columns())),
            ('AuditRowCount', len(audit_rows)),
        ]
        ws_meta.append(['Key', 'Value'])
        for k, v in meta_pairs:
            ws_meta.append([k, v])

        # AuditDay：每日原始参与度与百分比（便于人工核对 LineData）
        ws_audit_day.append(['Date', 'PositiveEng', 'NeutralEng', 'NegativeEng', 'TotalEng', 'Positive%', 'Neutral%', 'Negative%'])
        for d in sorted(day_eng_buckets.keys()):
            bucket = day_eng_buckets[d]
            day_total = sum(bucket.values())
            pos = bucket.get('Positive', 0.0)
            neu = bucket.get('Neutral', 0.0)
            neg = bucket.get('Negative', 0.0)
            pos_pct = (pos / day_total * 100.0) if day_total > 0 else 0.0
            neu_pct = (neu / day_total * 100.0) if day_total > 0 else 0.0
            neg_pct = (neg / day_total * 100.0) if day_total > 0 else 0.0
            ws_audit_day.append([
                d.strftime('%Y-%m-%d'),
                round(pos, 2),
                round(neu, 2),
                round(neg, 2),
                round(day_total, 2),
                round(pos_pct, 2),
                round(neu_pct, 2),
                round(neg_pct, 2),
            ])

        # AuditRow：示例原始行及派生值（便于人工核对行级参与度与情绪分类）
        row_header = ['Date', 'Polarity', 'Sentiment']
        if self.audit_post_id_col:
            row_header.append('PostId')
        if self.audit_channel_col:
            row_header.append('Channel')
        row_header += self.eng_fields_candidates + ['RowEngagement']
        ws_audit_row.append(row_header)
        for rec in audit_rows:
            row_vals = [
                rec.get('Date'),
                rec.get('Polarity'),
                rec.get('Sentiment'),
            ]
            if self.audit_post_id_col:
                row_vals.append(rec.get('PostId'))
            if self.audit_channel_col:
                row_vals.append(rec.get('Channel'))
            for c in self.eng_fields_candidates:
                v = rec.get(c)
                try:
                    row_vals.append(float(v) if v is not None else 0.0)
                except Exception:
                    row_vals.append(0.0)
            row_vals.append(rec.get('RowEngagement'))
            ws_audit_row.append(row_vals)

        # AuditMonth：当月各情绪参与度及百分比（便于人工核对 PieData）
        ws_audit_month.append(['Sentiment', 'Engagement', 'Percentage'])
        for lbl in self.labels:
            part = pie_eng_sums.get(lbl, 0.0)
            pct = (part / total_engagement * 100.0) if total_engagement > 0 else 0.0
            ws_audit_month.append([lbl, round(part, 2), round(pct, 2)])

        # 保存 Excel
        wb.save(self.excel_file_path)
        print(f"数据已生成：{self.excel_file_path}")


def main():
    page_dir = os.path.dirname(os.path.abspath(__file__))
    cfg_path = os.path.join(page_dir, 'config.yaml')
    gen = P13DataGenerator(cfg_path)
    gen.run()


if __name__ == '__main__':
    main()