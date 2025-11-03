#!/usr/bin/env python3
"""
fill_from_excel.py

职责：
- 读取 `output/p18_data.xlsx`，将数值写入页面级 `template/ppt/charts/chart1..4.xml` 的 `numCache`；
- 设置 `externalData/autoUpdate=1`；
- 写入 `_rels/chartX.xml.rels` 指向页面级嵌入 xlsx；

说明：
- 不修改模板文字或样式；仅替换数据缓存与关系目标；
- 颜色与顺序保持模板配置；
- 临时文件与处理均在页面级目录内完成。
"""

from pathlib import Path
from lxml import etree as ET
import zipfile
import shutil
import subprocess
import sys

try:
    from openpyxl import load_workbook
except Exception:
    raise RuntimeError('需要 openpyxl 读取 output/p18_data.xlsx')

PAGE_DIR = Path(__file__).resolve().parent
OUT_XLSX = PAGE_DIR / 'p18_data.xlsx'
# 使用页面级 tmp 目录，统一使用 tmp/ppt 结构
TMP_DIR = PAGE_DIR / 'tmp'

# 路径解析：兼容两种解压结构（tmp/ppt/ppt 与 tmp/ppt）
def _resolve_tmp_paths():
    """解析 tmp 下的关键目录路径。
    返回 (ppt_root, charts_dir, rels_dir, embeddings_dir, slides_dir, content_types_path)
    """
    base = TMP_DIR / 'ppt'
    ppt_root = base / 'ppt' if (base / 'ppt').exists() else base
    charts_dir = ppt_root / 'charts'
    rels_dir = charts_dir / '_rels'
    embeddings_dir = ppt_root / 'embeddings'
    slides_dir = ppt_root / 'slides'
    content_types_path = base / '[Content_Types].xml'
    return ppt_root, charts_dir, rels_dir, embeddings_dir, slides_dir, content_types_path

NS = {
    'c': 'http://schemas.openxmlformats.org/drawingml/2006/chart',
    'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
}


def _ensure_templates():
    """确保图表模板文件存在于 tmp 解压目录中。"""
    ppt_root, charts_dir, rels_dir, _, _, _ = _resolve_tmp_paths()
    charts_dir.mkdir(parents=True, exist_ok=True)
    rels_dir.mkdir(parents=True, exist_ok=True)
    for i in range(1, 5):
        chart_file = charts_dir / f'chart{i}.xml'
        if not chart_file.exists():
            print(f'[p18] 警告：图表文件不存在 {chart_file}，请先运行 generate_excel.py')
        rels_file = rels_dir / f'chart{i}.xml.rels'
        if not rels_file.exists():
            print(f'[p18] 警告：关系文件不存在 {rels_file}，请先运行 generate_excel.py')


def _set_external_auto_update(tree: ET.ElementTree):
    ed = tree.find('.//{%(c)s}externalData' % NS)
    if ed is None:
        return
    au = ed.find('{%(c)s}autoUpdate' % NS)
    if au is None:
        au = ET.SubElement(ed, '{%(c)s}autoUpdate' % NS)
    au.set('val', '1')


def _write_num_cache_points(num_ref: ET.Element, values):
    """用给定 values 重写 numCache/pt 节点序列。"""
    cache = num_ref.find('{%(c)s}numCache' % NS)
    if cache is None:
        cache = ET.SubElement(num_ref, '{%(c)s}numCache' % NS)
    # formatCode 保持模板原样
    pts = list(cache.findall('{%(c)s}pt' % NS))
    for p in pts:
        cache.remove(p)
    # 重新写 ptCount 与 pt
    pt_count = cache.find('{%(c)s}ptCount' % NS)
    if pt_count is None:
        pt_count = ET.SubElement(cache, '{%(c)s}ptCount' % NS)
    pt_count.set('val', str(len(values)))
    for idx, v in enumerate(values):
        pt = ET.SubElement(cache, '{%(c)s}pt' % NS, idx=str(idx))
        vv = ET.SubElement(pt, '{%(c)s}v' % NS)
        vv.text = str(int(round(v)))


def _update_chart_values(chart_xml_path: Path, values):
    tree = ET.fromstring(chart_xml_path.read_bytes())
    # 找到第一条系列的数值引用
    num_ref = tree.find('.//{%(c)s}ser/{%(c)s}val/{%(c)s}numRef' % NS)
    if num_ref is None:
        raise RuntimeError(f'未找到数值引用: {chart_xml_path.name}')
    _write_num_cache_points(num_ref, values)
    # 关键修复：移除固定的Y轴范围(<c:min>/<c:max>)，改为自动缩放，避免最后一个值超出轴范围被裁剪
    for scaling in list(tree.findall('.//{%(c)s}valAx/{%(c)s}scaling' % NS)):
        mn = scaling.find('{%(c)s}min' % NS)
        mx = scaling.find('{%(c)s}max' % NS)
        if mn is not None:
            scaling.remove(mn)
        if mx is not None:
            scaling.remove(mx)
    # 启用图表数据标签，使其随数据点移动并显示当前数值
    line_chart = tree.find('.//{%(c)s}lineChart' % NS)
    if line_chart is None:
        raise RuntimeError(f'未找到折线图节点: {chart_xml_path.name}')
    dLbls = line_chart.find('{%(c)s}dLbls' % NS)
    if dLbls is None:
        dLbls = ET.SubElement(line_chart, '{%(c)s}dLbls' % NS)
    # 统一关闭不需要的标签项，开启数值标签
    def _set_flag(parent, tag, val):
        # 正确拼接命名空间前缀，避免格式化错误
        el = parent.find(f'{{{NS["c"]}}}{tag}')
        if el is None:
            el = ET.SubElement(parent, f'{{{NS["c"]}}}{tag}')
        el.set('val', str(val))
    _set_flag(dLbls, 'showLegendKey', 0)
    _set_flag(dLbls, 'showCatName', 0)
    _set_flag(dLbls, 'showSerName', 0)
    _set_flag(dLbls, 'showPercent', 0)
    _set_flag(dLbls, 'showBubbleSize', 0)
    _set_flag(dLbls, 'showLeaderLines', 0)
    _set_flag(dLbls, 'showVal', 1)
    # 标签位置：顶部（随点位变化），并设置数值格式为 0"%"（整数后加%号，不进行百分比缩放）
    pos = dLbls.find('{%(c)s}dLblPos' % NS)
    if pos is None:
        pos = ET.SubElement(dLbls, '{%(c)s}dLblPos' % NS)
    pos.set('val', 't')
    numFmt = dLbls.find('{%(c)s}numFmt' % NS)
    if numFmt is None:
        numFmt = ET.SubElement(dLbls, '{%(c)s}numFmt' % NS)
    numFmt.set('formatCode', '0"%"')
    numFmt.set('sourceLinked', '0')
    # 系列级也显式开启标签，避免模板默认关闭
    for ser in list(line_chart.findall('{%(c)s}ser' % NS)):
        dLbls_ser = ser.find('{%(c)s}dLbls' % NS)
        if dLbls_ser is None:
            dLbls_ser = ET.SubElement(ser, '{%(c)s}dLbls' % NS)
        _set_flag(dLbls_ser, 'showVal', 1)
        _set_flag(dLbls_ser, 'showLeaderLines', 0)
        pos_ser = dLbls_ser.find('{%(c)s}dLblPos' % NS)
        if pos_ser is None:
            pos_ser = ET.SubElement(dLbls_ser, '{%(c)s}dLblPos' % NS)
        pos_ser.set('val', 't')
        numFmt_ser = dLbls_ser.find('{%(c)s}numFmt' % NS)
        if numFmt_ser is None:
            numFmt_ser = ET.SubElement(dLbls_ser, '{%(c)s}numFmt' % NS)
        numFmt_ser.set('formatCode', '0"%"')
        numFmt_ser.set('sourceLinked', '0')
    _set_external_auto_update(ET.ElementTree(tree))
    chart_xml_path.write_bytes(ET.tostring(tree, xml_declaration=True, encoding='UTF-8', standalone='yes'))


def _write_rels(chart_idx: int, rels_dir: Path):
    rels = ET.Element('Relationships', xmlns='http://schemas.openxmlformats.org/package/2006/relationships')
    r = ET.SubElement(rels, 'Relationship')
    r.set('Id', 'rId1')
    r.set('Type', 'http://schemas.openxmlformats.org/officeDocument/2006/relationships/package')
    r.set('Target', f'../embeddings/Microsoft_Office_Excel_Binary_Worksheet{chart_idx}.xlsx')
    path = rels_dir / f'chart{chart_idx}.xml.rels'
    path.write_bytes(ET.tostring(rels, xml_declaration=True, encoding='UTF-8', standalone='yes'))


def _clean_tmp_extracted():
    """清理 tmp 中已解压的模板，防止污染。"""
    extract_dir = TMP_DIR / 'ppt'
    if extract_dir.exists():
        shutil.rmtree(extract_dir)


def _run_generate_excel():
    """运行页面级 generate_excel.py，内含解压模板与数据生成逻辑。"""
    script = PAGE_DIR / 'generate_excel.py'
    if not script.exists():
        raise FileNotFoundError(f'缺少数据生成脚本: {script}')
    subprocess.run([sys.executable, str(script)], check=True, cwd=str(PAGE_DIR))


def _update_content_types_for_xlsx(ct_path: Path, embed_names: list[str]):
    """更新 Content_Types，使嵌入引用 xlsx，而非 xlsb。"""
    if not ct_path.exists():
        raise FileNotFoundError(f'缺少 Content_Types 文件: {ct_path}')
    ns_ct = 'http://schemas.openxmlformats.org/package/2006/content-types'
    root = ET.fromstring(ct_path.read_bytes())
    # 移除 xlsb 的覆盖项
    for o in list(root.findall('{%s}Override' % ns_ct)):
        part = o.get('PartName') or ''
        if part.startswith('/ppt/embeddings/') and part.endswith('.xlsb'):
            root.remove(o)
    existing = {o.get('PartName') for o in root.findall('{%s}Override' % ns_ct)}
    for nm in embed_names:
        part = f'/ppt/embeddings/{nm}'
        if part not in existing:
            ET.SubElement(root, '{%s}Override' % ns_ct, PartName=part,
                          ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    ct_path.write_bytes(ET.tostring(root, xml_declaration=True, encoding='UTF-8', standalone='yes'))


def _read_ranking_values() -> list[int]:
    """读取 Ranking 第二行（Position），不允许空值。"""
    try:
        from openpyxl import load_workbook
    except Exception:
        raise RuntimeError('需要 openpyxl 读取 p18_data.xlsx')
    xlsx = OUT_XLSX
    if not xlsx.exists():
        raise FileNotFoundError(f'[p18] 缺少排名数据文件: {xlsx}，请先生成')
    wb = load_workbook(str(xlsx), data_only=True)
    if 'Ranking' not in wb.sheetnames:
        raise RuntimeError('[p18] p18_data.xlsx 缺少 Ranking 工作表')
    ws = wb['Ranking']
    row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    vals = []
    for v in row[1:]:
        if v is None:
            raise RuntimeError('[p18] Ranking 存在空值，禁止兜底')
        vals.append(int(v))
    if not vals:
        raise RuntimeError('[p18] Ranking 为空，禁止兜底')
    return vals


def _update_ranking_in_slide(slide_xml_path: Path, ranks: list[int]):
    """将 slide1.xml 中的 #数字 文本替换为给定排名。"""
    root = ET.fromstring(slide_xml_path.read_bytes())
    ns_a = 'http://schemas.openxmlformats.org/drawingml/2006/main'
    nodes = []
    for t in root.findall('.//{%s}t' % ns_a):
        text = (t.text or '')
        if text.startswith('#') and text[1:].isdigit():
            nodes.append(t)
    if len(nodes) != len(ranks):
        raise RuntimeError(f'[p18] Ranking 文本节点数量({len(nodes)})与数据数量({len(ranks)})不一致')
    for i, t in enumerate(nodes):
        t.text = f'#{int(ranks[i])}'
    slide_xml_path.write_bytes(ET.tostring(root, xml_declaration=True, encoding='UTF-8', standalone='yes'))


def _read_main_values() -> list[int]:
    """读取 MainTrend 第二行的 6 个值，严格校验数量与空值。"""
    try:
        from openpyxl import load_workbook
    except Exception:
        raise RuntimeError('需要 openpyxl 读取 p18_data.xlsx')
    xlsx = OUT_XLSX
    if not xlsx.exists():
        raise FileNotFoundError(f'[p18] 缺少主趋势数据文件: {xlsx}，请先生成')
    wb = load_workbook(str(xlsx), data_only=True)
    if 'MainTrend' not in wb.sheetnames:
        raise RuntimeError('[p18] p18_data.xlsx 缺少 MainTrend 工作表')
    ws = wb['MainTrend']
    row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    vals = []
    for v in row[1:]:
        if v is None:
            raise RuntimeError('[p18] MainTrend 存在空值，禁止兜底')
        vals.append(int(round(float(v))))
    if len(vals) != 6:
        raise RuntimeError(f'[p18] MainTrend 期望 6 个值，实际 {len(vals)}')
    return vals


def _update_main_percent_texts(slide_xml_path: Path, values: list[int]):
    """已弃用：使用图表数据标签替代静态文本。
    为避免位置与数值不同步，这里改为清空这些静态百分比文本。
    """
    import re as _re
    root = ET.fromstring(slide_xml_path.read_bytes())
    ns_a = 'http://schemas.openxmlformats.org/drawingml/2006/main'
    cleared = 0
    for t in root.findall('.//{%s}t' % ns_a):
        tx = (t.text or '').strip()
        if _re.fullmatch(r'\d+%$', tx):
            t.text = ''
            cleared += 1
    if cleared == 0:
        raise RuntimeError('[p18] 未发现可清理的静态百分比文本节点')
    slide_xml_path.write_bytes(ET.tostring(root, xml_declaration=True, encoding='UTF-8', standalone='yes'))


def _package_final(output_path: Path):
    """将 tmp/ppt 目录打包为目标 PPTX。"""
    base = TMP_DIR / 'ppt'
    if not base.exists():
        raise FileNotFoundError('打包失败：tmp/ppt 不存在，请确保已解压模板')
    with zipfile.ZipFile(str(output_path), 'w', compression=zipfile.ZIP_DEFLATED) as z:
        for p in base.rglob('*'):
            if p.is_file():
                arc = p.relative_to(base)
                z.write(p, arcname=str(arc))
    print(f'[p18] 已生成: {output_path}')


def main():
    """端到端流程：清理 → 数据生成与解压 → 填充 → 文本更新 → 内容类型修正 → 打包。"""
    # 1) 清理 tmp 已解压模板
    _clean_tmp_extracted()

    # 2) 生成数据并解压模板到 tmp（generate_excel 负责解压）
    _run_generate_excel()

    # 3) 路径解析与存在性校验
    ppt_root, charts_dir, rels_dir, embeddings_dir, slides_dir, ct_path = _resolve_tmp_paths()
    if not charts_dir.exists() or not slides_dir.exists():
        raise RuntimeError('[p18] 模板解压不完整，缺少 charts 或 slides 目录')

    # 4) 加载数据，写 chart1..4.xml 的 numCache
    wb = load_workbook(str(OUT_XLSX), data_only=True)
    def row_values(sheet_name):
        ws = wb[sheet_name]
        return [cell.value for cell in list(ws.iter_rows(min_row=2, max_row=2))[0][1:]]
    main_values = row_values('MainTrend')
    lovers_values = row_values('Lovers')
    neutral_values = row_values('Neutral')
    haters_values = row_values('Haters')
    _update_chart_values(charts_dir / 'chart1.xml', main_values)
    _update_chart_values(charts_dir / 'chart2.xml', lovers_values)
    _update_chart_values(charts_dir / 'chart3.xml', neutral_values)
    _update_chart_values(charts_dir / 'chart4.xml', haters_values)

    # 5) 更新 chart rels 到嵌入 xlsx，并校验嵌入文件存在
    for i in range(1, 5):
        _write_rels(i, rels_dir)
        embed_name = f'Microsoft_Office_Excel_Binary_Worksheet{i}.xlsx'
        if not (embeddings_dir / embed_name).exists():
            raise FileNotFoundError(f'[p18] 缺少嵌入工作簿: {embeddings_dir / embed_name}')

    # 6) 更新 slide1 文本：仅更新 Ranking，并清空主趋势静态百分比文本（改用图表数据标签）
    slide1 = slides_dir / 'slide1.xml'
    if not slide1.exists():
        raise FileNotFoundError(f'[p18] 模板缺少 slide1.xml: {slide1}')
    ranks = _read_ranking_values()
    _update_ranking_in_slide(slide1, ranks)
    main_vals = _read_main_values()
    _update_main_percent_texts(slide1, main_vals)

    # 7) 修正 Content_Types 为 xlsx 嵌入
    embed_files = [f.name for f in embeddings_dir.glob('Microsoft_Office_Excel_Binary_Worksheet*.xlsx')]
    if len(embed_files) != 4:
        raise RuntimeError(f'[p18] 嵌入 xlsx 数量异常：{len(embed_files)}，期望 4')
    _update_content_types_for_xlsx(ct_path, embed_files)

    # 8) 打包为最终 PPTX到页面级 output 目录
    output_dir = PAGE_DIR / 'output'
    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / 'p18-final.pptx'
    _package_final(out_path)
    print('[p18] 填充完成：chartXML 与 rels 已更新，autoUpdate=1；文本与内容类型已修正；已打包。')


if __name__ == '__main__':
    main()