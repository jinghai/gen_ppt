#!/usr/bin/env python3
"""
generate_excel.py

职责：
- 解压模板PPT到tmp目录，读取模板中的Excel数据作为基础；
- 从 `input/neticle-v4-08.sqlite` 的宽表 `mentions_wide` 聚合生成法国市场的月度情感百分比；
- 计算 Net Lovers%（Lovers%-Haters%）以及 Lenovo 的排名；
- 用最新数据替换模板数据中对应的部分；
- 写出一份可检验的汇总 Excel：`p18_data.xlsx`；
- 同时生成 4 份嵌入用 xlsx（Sheet1），用于替换模板中的 xlsb：
  - Microsoft_Office_Excel_Binary_Worksheet1.xlsx：主趋势 A1:F1
  - Microsoft_Office_Excel_Binary_Worksheet2.xlsx：Lovers A1:D1
  - Microsoft_Office_Excel_Binary_Worksheet3.xlsx：Neutral A1:D1
  - Microsoft_Office_Excel_Binary_Worksheet4.xlsx：Haters A1:D1

注意：
- 代码仅在页面级目录读写文件；无跨页副作用。
- 缺失的数据使用模板数据兜底，计算的新数据及时更新。
"""

from pathlib import Path
import sqlite3
from datetime import datetime
import yaml
import zipfile
import shutil
import tempfile

# 仅依赖 openpyxl 写 xlsx；若无法导入，给出降级提示并退出
try:
    from openpyxl import Workbook, load_workbook
except Exception as e:
    raise RuntimeError("需要 openpyxl 才能生成 xlsx，请安装后重试") from e

PAGE_DIR = Path(__file__).resolve().parent
# 使用页面级tmp目录，统一使用tmp/ppt结构
TMP_DIR = PAGE_DIR / 'tmp'
TMP_DIR.mkdir(parents=True, exist_ok=True)
# 统一使用tmp/ppt/ppt/embeddings目录
TEMPLATE_EMBED_DIR = TMP_DIR / 'ppt' / 'ppt' / 'embeddings'
TEMPLATE_EMBED_DIR.mkdir(parents=True, exist_ok=True)

# 从页面级配置加载所需参数
CFG_PATH = PAGE_DIR / 'config.yaml'
if not CFG_PATH.exists():
    raise FileNotFoundError('缺少页面配置文件 charts/p18/config.yaml')
with CFG_PATH.open('r', encoding='utf-8') as f:
    CFG = yaml.safe_load(f) or {}

def _resolve_path(p: str) -> Path:
    q = Path(p)
    if not q.is_absolute():
        q = (PAGE_DIR / q).resolve()
    return q

NETICLE_DB = _resolve_path(CFG.get('data_sources', {}).get('neticle_db', '../../input/neticle-v4-08.sqlite'))
TABLE_NAME = CFG.get('data_sources', {}).get('neticle_table', 'mentions_wide')
CFG_DATE_COL = CFG.get('data_sources', {}).get('date_column')
CFG_COUNTRY_COL = CFG.get('data_sources', {}).get('country_column')
CFG_SENTIMENT_COL = CFG.get('sentiment', {}).get('thresholds', {}).get('column', 'polarity')

MONTHS_MAIN = CFG.get('months', {}).get('main', ['2025-07','2025-08','2025-09','2025-10','2025-11','2025-12'])
MONTHS_SPLIT = CFG.get('months', {}).get('split', ['2025-09','2025-10','2025-11','2025-12'])
AUTO_MONTHS = bool(CFG.get('months', {}).get('auto_discover', False))
BRANDS_FOR_RANK = CFG.get('ranking', {}).get('brands', ['lenovo','dell','hp','asus','acer','apple','samsung'])
COUNTRY_ID = CFG.get('filters', {}).get('country_id', 39)
COUNTRY_NAME = CFG.get('filters', {}).get('country_name', 'France')
BRAND_KEY = CFG.get('filters', {}).get('brand_key', 'Lenovo')
KEYWORD_LIKE = CFG.get('filters', {}).get('keyword_like', '%lenovo%')
POS_MIN = float(CFG.get('sentiment', {}).get('thresholds', {}).get('positive_min', 0.33))
NEG_MAX = float(CFG.get('sentiment', {}).get('thresholds', {}).get('negative_max', -0.33))


def _get_available_date_columns(conn: sqlite3.Connection):
    """检测 mentions_wide 中可用的日期列（白名单），返回存在的列名列表。"""
    whitelist = {
        'publishedAt', 'createdAt', 'created_at', 'postDate', 'post_date',
        'date', 'timestamp', 'created_ts', 'createdAtUtcMs', 'discoveredAtUtcMs'
    }
    cur = conn.execute(f"PRAGMA table_info({TABLE_NAME})")
    cols = [r[1] for r in cur.fetchall() if r and len(r) >= 2]
    if not cols:
        return []
    if CFG_DATE_COL and CFG_DATE_COL in cols:
        return [CFG_DATE_COL]
    candidates = []
    for c in cols:
        cl = c.lower()
        if c in whitelist or any(k in cl for k in ['date','time','timestamp','published','created','post']):
            candidates.append(c)
    return candidates


def _month_match_clause_sql(date_cols: list[str]) -> str:
    """生成按月份匹配的 SQL 片段，支持文本/epoch 两种存储。"""
    parts = []
    for c in date_cols:
        parts.append(f"strftime('%Y-%m', {c}) = :ym")
        parts.append(f"strftime('%Y-%m', datetime({c}, 'unixepoch')) = :ym")
        parts.append(f"substr(COALESCE({c},''),1,7) = :ym")
    return ' OR '.join(parts) if parts else "1=0"  # 无日期列则不匹配


def _get_available_brand_columns(conn: sqlite3.Connection):
    """检测用于品牌匹配的列，仅允许 `keyword_label`。若不存在则返回空列表。"""
    cur = conn.execute(f"PRAGMA table_info({TABLE_NAME})")
    cols = [r[1] for r in cur.fetchall() if r and len(r) >= 2]
    return ['keyword_label'] if 'keyword_label' in cols else []


def _brand_where_and_params(brand_columns: list[str], brand: str):
    """仅使用 `keyword_label` 的 LIKE 模式进行品牌匹配。
    - 若缺少 `keyword_label` 列，调用方应在更高层抛错。
    - Lenovo（配置中的 `filters.brand_key`）使用 `filters.keyword_like` 模式；
      其他品牌使用 `'%{brand.lower()}%'`。
    """
    if 'keyword_label' not in brand_columns:
        return '1=0', {}
    brand_key_cfg = str(CFG.get('filters', {}).get('brand_key', '')).lower()
    keyword_like_cfg = CFG.get('filters', {}).get('keyword_like', '%lenovo%')
    pat = keyword_like_cfg if brand.lower() == brand_key_cfg else f"%{brand.lower()}%"
    return "LOWER(keyword_label) LIKE LOWER(:brand_pat)", {'brand_pat': pat}


def _brand_match_clause_sql(brand_columns: list[str]) -> str:
    """仅返回 `LOWER(keyword_label) LIKE LOWER(:keyword_like)`。
    若缺少 `keyword_label`，调用方应抛错。"""
    if 'keyword_label' not in brand_columns:
        return '1=0'
    return '(LOWER(keyword_label) LIKE LOWER(:keyword_like))'


def _get_country_column(conn: sqlite3.Connection):
    """返回国家列名，仅允许使用配置中的 `country_column`（默认 `countryId`）。
    若列不存在则抛出异常，避免兜底。"""
    cur = conn.execute(f"PRAGMA table_info({TABLE_NAME})")
    cols = {r[1] for r in cur.fetchall() if r and len(r) >= 2}
    if CFG_COUNTRY_COL and CFG_COUNTRY_COL in cols:
        return CFG_COUNTRY_COL
    raise RuntimeError(f"mentions_wide 缺少国家列: {CFG_COUNTRY_COL}")


def _query_monthly_sentiment(conn: sqlite3.Connection, brand: str, ym: str):
    """在 mentions_wide 中聚合某品牌某月的情感计数。
    识别规则：文本包含品牌关键字；国家=指定country_id。
    返回 (pos, neu, neg, total) 或 None。
    """
    date_cols = _get_available_date_columns(conn)
    if not date_cols:
        print(f"警告：未找到可用的日期列")
        return None
    
    brand_cols = _get_available_brand_columns(conn)
    if not brand_cols:
        raise RuntimeError('mentions_wide 缺少 keyword_label 列，无法进行品牌匹配')
    
    country_col = _get_country_column(conn)
    # 如缺少国家列，_get_country_column 已抛异常
    
    # 构建时间过滤条件 - 支持时间戳格式
    date_col = date_cols[0]  # 使用第一个可用的日期列
    if 'UtcMs' in date_col:
        # 时间戳格式（毫秒）
        year, month = ym.split('-')
        start_ts = int(datetime(int(year), int(month), 1).timestamp() * 1000)
        if int(month) == 12:
            end_ts = int(datetime(int(year) + 1, 1, 1).timestamp() * 1000)
        else:
            end_ts = int(datetime(int(year), int(month) + 1, 1).timestamp() * 1000)
        month_clause = f"{date_col} >= {start_ts} AND {date_col} < {end_ts}"
    else:
        # 字符串格式
        month_clause = _month_match_clause_sql(date_cols)
    
    brand_clause, brand_params = _brand_where_and_params(brand_cols, brand)
    
    # 构建WHERE条件
    where_conditions = []
    if country_col:
        where_conditions.append(f"{country_col} = :country")
    where_conditions.append(f"( {month_clause} )")
    where_conditions.append(f"( {brand_clause} )")
    
    where_clause = " AND ".join(where_conditions)
    
    # 使用sentiment列进行情感分类
    sentiment_col = CFG_SENTIMENT_COL
    
    sql = f'''
    SELECT
      SUM(CASE WHEN {sentiment_col} > :pos_min THEN 1 ELSE 0 END) AS pos,
      SUM(CASE WHEN {sentiment_col} BETWEEN :neg_max AND :pos_min THEN 1 ELSE 0 END) AS neu,
      SUM(CASE WHEN {sentiment_col} < :neg_max THEN 1 ELSE 0 END) AS neg,
      COUNT(*) AS total
    FROM {TABLE_NAME}
    WHERE {where_clause}
    '''
    
    # 构建参数字典 - 使用配置中的keyword_like而不是传入的brand参数
    params = {
        'pos_min': POS_MIN,
        'neg_max': NEG_MAX,
        'country': COUNTRY_ID,
    }
    # 合并品牌匹配参数（keyword_label 或 LIKE 模式）
    params.update(brand_params)
    
    # 只有在使用字符串格式时才添加ym参数
    if 'UtcMs' not in date_col:
        params['ym'] = ym
    
    cur = conn.execute(sql, params)
    row = cur.fetchone()
    if not row or row[3] is None or row[3] == 0:
        return None
    return int(row[0] or 0), int(row[1] or 0), int(row[2] or 0), int(row[3] or 0)


def _pct(a: int, b: int) -> int:
    """安全百分比，四舍五入为整数。"""
    if b <= 0:
        return 0
    return round(100.0 * a / b)


def compute_for_brand(conn: sqlite3.Connection, brand: str, months: list[str]):
    """返回指定品牌的月度百分比：lovers/neutral/haters 以及 net_lovers。"""
    lovers = {}
    neutral = {}
    haters = {}
    net_lovers = {}
    for ym in months:
        q = _query_monthly_sentiment(conn, brand, ym)
        if q is None:
            continue
        pos, neu, neg, total = q
        lovers[ym] = _pct(pos, total)
        neutral[ym] = _pct(neu, total)
        haters[ym] = _pct(neg, total)
        net_lovers[ym] = lovers[ym] - haters[ym]
    return lovers, neutral, haters, net_lovers


def discover_months(conn: sqlite3.Connection, brand: str, need_main: int, need_split: int):
    """从数据库发现有数据的月份，返回升序列表。若不足则报错。"""
    # 动态联合查询，基于存在的日期列生成月份
    date_cols = _get_available_date_columns(conn)
    brand_cols = _get_available_brand_columns(conn)
    if not date_cols:
        raise ValueError('无法识别日期列，mentions_wide 不包含已知日期字段')
    if 'keyword_label' not in brand_cols:
        raise ValueError('mentions_wide 缺少 keyword_label 列，无法进行品牌匹配')
    country_col = _get_country_column(conn)
    country_filter = f"{country_col} = :country AND "
    selects = []
    brand_where = "LOWER(keyword_label) LIKE LOWER(:keyword_like)"
    base_where = f"{country_filter}( {brand_where} )"
    for c in date_cols:
        # 处理毫秒时间戳（createdAtUtcMs）
        if 'ms' in c.lower() or c == 'createdAtUtcMs':
            selects.append(f"SELECT strftime('%Y-%m', datetime({c}/1000, 'unixepoch')) AS ym FROM {TABLE_NAME} WHERE {base_where}")
        else:
            # 处理其他日期格式
            selects.append(f"SELECT strftime('%Y-%m', {c}) AS ym FROM {TABLE_NAME} WHERE {base_where}")
            selects.append(f"SELECT strftime('%Y-%m', datetime({c}, 'unixepoch')) AS ym FROM {TABLE_NAME} WHERE {base_where}")
            selects.append(f"SELECT substr(COALESCE({c},''),1,7) AS ym FROM {TABLE_NAME} WHERE {base_where}")
    union_sql = ' UNION '.join(selects)
    sql = f"SELECT DISTINCT ym FROM ( {union_sql} ) WHERE ym IS NOT NULL ORDER BY ym ASC;"
    # 使用keyword_like参数而不是brand参数
    cur = conn.execute(sql, {'country': COUNTRY_ID, 'keyword_like': KEYWORD_LIKE})
    months = [r[0] for r in cur.fetchall() if r and r[0]]
    if len(months) < need_main:
        raise ValueError(f'可用月份不足 {need_main} 个，仅发现: {months}')
    main = months[-need_main:]
    split = main[-need_split:]
    return main, split


def extract_template_data():
    """解压模板PPT，读取其中的Excel或chart XML数据作为基础数据。
    实现策略：
    1) 首选读取嵌入的 xlsx（tmp/ppt/ppt/embeddings）。
    2) 若无法读取或为空，回退解析 chartXML 的 numCache（tmp/ppt/ppt/charts/chart1..4.xml）。
    3) 若两种方式均失败，抛出异常（不使用默认兜底以避免掩盖问题）。
    """
    template_pptx = PAGE_DIR / 'p18.pptx'
    if not template_pptx.exists():
        raise FileNotFoundError(f'[p18] 模板文件不存在：{template_pptx}')

    # 创建临时目录解压PPT - 统一使用tmp/ppt目录
    extract_dir = TMP_DIR / 'ppt'
    if extract_dir.exists():
        shutil.rmtree(extract_dir)
    extract_dir.mkdir(parents=True, exist_ok=True)

    # 内部辅助：解析 chart XML 的数值
    def _parse_chart_values(xml_path: Path) -> list[int]:
        import xml.etree.ElementTree as ET
        ns = {'c': 'http://schemas.openxmlformats.org/drawingml/2006/chart'}
        try:
            tree = ET.parse(str(xml_path))
            root = tree.getroot()
            # 查找第一个系列的数值缓存
            # 路径：c:chart/c:plotArea//c:ser//c:val//c:numCache//c:pt/c:v
            ser = root.find('.//c:chart/c:plotArea//c:ser', ns)
            if ser is None:
                # 有些模板结构不同，尝试直接在 plotArea 下找
                pass
            vals = []
            for v in root.findall('.//c:chart/c:plotArea//c:ser//c:val//c:numCache//c:pt/c:v', ns):
                try:
                    # 将数值转为 int，保留四舍五入
                    vals.append(int(round(float(v.text))))
                except Exception:
                    continue
            return vals
        except Exception as e:
            print(f'[p18] 解析 chartXML 失败 {xml_path.name}: {e}')
            return []

    try:
        # 解压PPT文件
        with zipfile.ZipFile(template_pptx, 'r') as zip_ref:
            zip_ref.extractall(extract_dir)

        # 路径准备
        embeddings_dir = extract_dir / 'ppt' / 'ppt' / 'embeddings'
        charts_dir = extract_dir / 'ppt' / 'ppt' / 'charts'
        if not embeddings_dir.exists():
            # 尝试备用路径
            embeddings_dir = extract_dir / 'ppt' / 'embeddings'
        if not charts_dir.exists():
            charts_dir = extract_dir / 'ppt' / 'charts'

        template_data: dict[str, list[int]] = {}

        # 先尝试读取嵌入xlsx
        excel_files = {
            'main': 'Microsoft_Office_Excel_Binary_Worksheet1.xlsx',
            'lovers': 'Microsoft_Office_Excel_Binary_Worksheet2.xlsx',
            'neutral': 'Microsoft_Office_Excel_Binary_Worksheet3.xlsx',
            'haters': 'Microsoft_Office_Excel_Binary_Worksheet4.xlsx',
        }

        for key, filename in excel_files.items():
            values: list[int] = []
            excel_path = embeddings_dir / filename if embeddings_dir.exists() else None
            if excel_path and excel_path.exists():
                try:
                    wb = load_workbook(str(excel_path), data_only=True)
                    ws = wb.active
                    # 读取第一行数据（A1:... 一个行向量）
                    row = list(ws.iter_rows(min_row=1, max_row=1))[0]
                    for cell in row:
                        if cell.value is None:
                            continue
                        try:
                            values.append(int(round(float(cell.value))))
                        except Exception:
                            # 忽略非数值单元格
                            pass
                except Exception as e:
                    print(f'[p18] 读取嵌入xlsx失败 {filename}: {e}，将尝试解析chartXML')
            else:
                print(f'[p18] 未找到嵌入xlsx {filename}，将尝试解析chartXML')

            # 若嵌入读取失败或为空，回退解析 chartXML
            if not values:
                chart_map = {
                    'main': 'chart1.xml',
                    'lovers': 'chart2.xml',
                    'neutral': 'chart3.xml',
                    'haters': 'chart4.xml',
                }
                chart_file = charts_dir / chart_map[key]
                if chart_file.exists():
                    values = _parse_chart_values(chart_file)
                else:
                    print(f'[p18] 未找到 chartXML {chart_map[key]}')

            template_data[key] = values
            print(f'[p18] 模板 {key} 数据: {values}')

        # 严格校验：若任一数据为空，则抛错，避免默认兜底掩盖问题
        missing = [k for k, v in template_data.items() if not v]
        if missing:
            raise RuntimeError(f'[p18] 模板数据提取失败，缺少: {missing}，请检查模板嵌入或chartXML')

        return template_data

    except Exception as e:
        # 显式抛错，避免使用默认值掩盖问题
        raise


from typing import Optional

def _normalize_month_label_to_ym(label: str) -> Optional[str]:
    """将类似 "Mar '25" / "Aug '25" 的模板标签转为 "YYYY-MM"。
    仅处理英文月份缩写：Jan..Dec。
    """
    if not isinstance(label, str):
        return None
    label = label.strip()
    months = {
        'Jan': '01', 'Feb': '02', 'Mar': '03', 'Apr': '04', 'May': '05', 'Jun': '06',
        'Jul': '07', 'Aug': '08', 'Sep': '09', 'Oct': '10', 'Nov': '11', 'Dec': '12',
    }
    try:
        # 形如 "Mar '25" 或 "Mar 25"
        parts = label.replace("\u2019", "'").replace("\u2018", "'").split()
        if len(parts) >= 2:
            mon = parts[0][:3]
            yr = parts[1].replace("'", "")
            if len(yr) == 2:
                yr = '20' + yr
            mm = months.get(mon)
            if mm:
                return f"{yr}-{mm}"
    except Exception:
        pass
    return None


def read_template_months(chart_idx: int) -> list[str]:
    """读取指定图表的类别标签并转为 YYYY-MM 列表。
    chart_idx: 1=主趋势, 2=Lovers, 3=Neutral, 4=Haters
    依赖：extract_template_data 已经将模板解压到 tmp/ppt。
    """
    charts_dir = TMP_DIR / 'ppt' / 'ppt' / 'charts'
    if not charts_dir.exists():
        charts_dir = TMP_DIR / 'ppt' / 'charts'
    chart1 = charts_dir / f'chart{chart_idx}.xml'
    import xml.etree.ElementTree as ET
    ns = {'c': 'http://schemas.openxmlformats.org/drawingml/2006/chart'}
    labels: list[str] = []
    try:
        tree = ET.parse(str(chart1))
        root = tree.getroot()
        # 优先读取字符串缓存 strCache
        for v in root.findall('.//c:chart/c:plotArea//c:ser//c:cat//c:strRef//c:strCache//c:pt/c:v', ns):
            ym = _normalize_month_label_to_ym(v.text)
            if ym:
                labels.append(ym)
        # 若没读到，尝试 numCache（轴为数字型日期）
        if not labels:
            for v in root.findall('.//c:chart/c:plotArea//c:ser//c:cat//c:numRef//c:numCache//c:pt/c:v', ns):
                try:
                    # 数字日期转为 Excel 序列日，仍需通过 config.axis_day_base 还原，这里不处理
                    labels.append(str(v.text))
                except Exception:
                    pass
    except Exception as e:
        print(f'[p18] 读取模板类别失败 chart{chart_idx}: {e}')
    return labels


def read_template_ranking_values() -> list[int]:
    """从模板 slide1.xml 中读取 Ranking 文本（形如 #3），返回整数列表。
    依赖：extract_template_data 已经将模板解压到 tmp/ppt。
    """
    slides_dir = TMP_DIR / 'ppt' / 'ppt' / 'slides'
    if not slides_dir.exists():
        slides_dir = TMP_DIR / 'ppt' / 'slides'
    slide1 = slides_dir / 'slide1.xml'
    import xml.etree.ElementTree as ET
    ns_a = {'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'}
    values: list[int] = []
    try:
        root = ET.parse(str(slide1)).getroot()
        for t in root.findall('.//a:t', ns_a):
            txt = (t.text or '').strip()
            if txt.startswith('#') and txt[1:].isdigit():
                values.append(int(txt[1:]))
    except Exception as e:
        print(f"[p18] 读取模板 Ranking 失败: {e}")
    return values


def create_default_template_data():
    """创建默认的模板数据"""
    return {
        'main': [45, 48, 52, 55, 58, 60],  # 6个月的Net Lovers%
        'lovers': [65, 68, 70, 72],        # 4个月的Lovers%
        'neutral': [25, 23, 22, 20],       # 4个月的Neutral%
        'haters': [10, 9, 8, 8]            # 4个月的Haters%
    }


def load_fallback_data():
    """加载示例数据作为后备"""
    fallback_file = PAGE_DIR / 'p18_data.xlsx'
    if not fallback_file.exists():
        return {}
    
    try:
        import pandas as pd
        df = pd.read_excel(fallback_file)
        
        # 按月份和品牌组织数据
        data = {}
        for _, row in df.iterrows():
            month = row['Month']
            brand = row['Brand']
            if month not in data:
                data[month] = {}
            data[month][brand] = {
                'net_lovers': row['Net Lovers%'],
                'positive': row['Positive'],
                'negative': row['Negative'],
                'total': row['Total']
            }
        return data
    except Exception as e:
        print(f'[p18] 警告：无法读取示例数据 {fallback_file}: {e}')
        return {}


def main():
    print('[p18] 开始生成 Excel 数据文件')
    
    # 1) 解压模板，读取模板数据作为基础
    print('[p18] 解压模板PPT，读取基础数据...')
    template_data = extract_template_data()
    
    # 2) 取消示例后备，严格以模板为基线、数据库为覆盖来源。
    
    # 3) 初始化数据（使用模板数据，并与配置月份对齐）
    months_main = MONTHS_MAIN
    months_split = MONTHS_SPLIT

    def _align(values: list[int], tmpl_months: list[str], target_months: list[str]) -> list[int]:
        # 将模板月份与数值映射后按目标月份重排；若模板未提供类别标签，则按末尾对齐。
        if not tmpl_months:
            # 模板未包含类别标签（常见于仅缓存数值的图表），按末尾N个值对齐
            n = len(target_months)
            if len(values) < n:
                raise RuntimeError(f'[p18] 模板数值不足以对齐到目标月份，需要 {n}，实际 {len(values)}')
            tail = values[-n:]
            print(f'[p18] 模板未含月份标签，按末尾对齐: {tail}')
            return tail
        if len(values) != len(tmpl_months):
            print(f'[p18] 警告：模板值与月份长度不一致: values={len(values)} tmpl={len(tmpl_months)}')
        m = {ym: values[i] for i, ym in enumerate(tmpl_months)}
        aligned = []
        missing = []
        for ym in target_months:
            if ym in m:
                aligned.append(m[ym])
            else:
                missing.append(ym)
        if missing:
            raise RuntimeError(f'[p18] 目标月份在模板类别中不存在: {missing}')
        return aligned

    tmpl_main_months = read_template_months(1)
    tmpl_lovers_months = read_template_months(2)
    tmpl_neutral_months = read_template_months(3)
    tmpl_haters_months = read_template_months(4)

    main_values = _align(template_data.get('main', []), tmpl_main_months, months_main)
    lovers_values = _align(template_data.get('lovers', []), tmpl_lovers_months, months_split)
    neutral_values = _align(template_data.get('neutral', []), tmpl_neutral_months, months_split)
    haters_values = _align(template_data.get('haters', []), tmpl_haters_months, months_split)
    
    main_sources = ['template'] * len(main_values)
    lovers_sources = ['template'] * len(lovers_values)
    neutral_sources = ['template'] * len(neutral_values)
    haters_sources = ['template'] * len(haters_values)
    
    # 4) 尝试连接数据库，用最新数据替换模板数据
    if NETICLE_DB.exists():
        try:
            print('[p18] 连接数据库，计算最新数据...')
            with sqlite3.connect(str(NETICLE_DB)) as conn:
                if AUTO_MONTHS:
                    try:
                        # 先查询所有可用月份
                        date_cols = _get_available_date_columns(conn)
                        brand_cols = _get_available_brand_columns(conn)
                        country_col = _get_country_column(conn)
                        country_filter = f"{country_col} = :country AND " if country_col else ""
                        brand_where = _brand_match_clause_sql(brand_cols)
                        base_where = f"{country_filter}( {brand_where} )"
                        
                        selects = []
                        for c in date_cols:
                            if 'ms' in c.lower() or c == 'createdAtUtcMs':
                                selects.append(f"SELECT strftime('%Y-%m', datetime({c}/1000, 'unixepoch')) AS ym FROM {TABLE_NAME} WHERE {base_where}")
                        
                        union_sql = ' UNION '.join(selects)
                        sql = f"SELECT DISTINCT ym FROM ( {union_sql} ) WHERE ym IS NOT NULL ORDER BY ym ASC;"
                        cur = conn.execute(sql, {'country': COUNTRY_ID, 'keyword_like': KEYWORD_LIKE})
                        available_months = [r[0] for r in cur.fetchall() if r and r[0]]
                        
                        if len(available_months) >= 2:
                            # 使用实际可用的月份
                            months_main = available_months
                            months_split = available_months
                            print(f'[p18] 自动发现月份 - 主趋势: {months_main}, 分割: {months_split}')
                        else:
                            print(f'[p18] 可用月份不足，使用配置月份: {available_months}')
                    except Exception as e:
                        # 自动发现失败直接抛出，避免静默兜底
                        raise
                
                # 计算品牌的数据
                lv_l, lv_n, lv_h, lv_net = compute_for_brand(conn, BRAND_KEY, months_main)
                
                # 更新主趋势数据
                for i, ym in enumerate(months_main):
                    if i < len(main_values):
                        v = lv_net.get(ym)
                        if v is not None:
                            main_values[i] = v
                            main_sources[i] = 'computed'
                            print(f'[p18] 更新 Net Lovers% 数据：{ym} = {v}%')
                        # 无数据库数据则保持模板值（来源为 template），不做示例兜底。
                
                # 更新三小图数据
                for i, ym in enumerate(months_split):
                    if i < len(lovers_values):
                        lv = lv_l.get(ym)
                        ne = lv_n.get(ym)
                        ha = lv_h.get(ym)
                        
                        if lv is not None:
                            lovers_values[i] = lv
                            lovers_sources[i] = 'computed'
                            print(f'[p18] 更新 Lovers% 数据：{ym} = {lv}%')
                        
                        if ne is not None:
                            neutral_values[i] = ne
                            neutral_sources[i] = 'computed'
                            print(f'[p18] 更新 Neutral% 数据：{ym} = {ne}%')
                        
                        if ha is not None:
                            haters_values[i] = ha
                            haters_sources[i] = 'computed'
                            print(f'[p18] 更新 Haters% 数据：{ym} = {ha}%')
                
        except Exception as e:
            # 严格模式：数据库连接或查询失败直接抛出
            raise
    else:
        print(f'[p18] 数据库文件不存在，使用模板数据: {NETICLE_DB}')
    
    # 5) 计算排名：以模板为基线，DB 有数据则覆盖，无默认常量兜底
    tmpl_ranks = read_template_ranking_values()
    if not tmpl_ranks:
        raise RuntimeError('[p18] 模板缺少 Ranking 文本，无法建立基线')
    # 对齐到主趋势月份数量（通常模板即为6个）
    if len(tmpl_ranks) != len(months_main):
        if len(tmpl_ranks) < len(months_main):
            raise RuntimeError(f"[p18] 模板 Ranking 数量不足：{len(tmpl_ranks)} < {len(months_main)}")
        ranking_values = tmpl_ranks[-len(months_main):]
    else:
        ranking_values = tmpl_ranks[:]
    ranking_sources = ['template'] * len(ranking_values)

    computed_ranking_count = 0
    if NETICLE_DB.exists():
        with sqlite3.connect(str(NETICLE_DB)) as conn:
            for i, ym in enumerate(months_main):
                scores = []
                for b in BRANDS_FOR_RANK:
                    _, _, _, net = compute_for_brand(conn, b, [ym])
                    s = net.get(ym)
                    if s is not None:
                        scores.append((b, s))
                if not scores:
                    # 保持模板值，记录来源
                    print(f"[p18] {ym} 排名使用模板值: #{ranking_values[i]}")
                    continue
                scores.sort(key=lambda x: x[1], reverse=True)
                pos = None
                for j, (b, _) in enumerate(scores, start=1):
                    if b.lower() == BRAND_KEY.lower():
                        pos = j
                        break
                if pos is not None:
                    ranking_values[i] = pos
                    ranking_sources[i] = 'computed'
                    computed_ranking_count += 1
                else:
                    print(f"[p18] {ym} 排名未找到 {BRAND_KEY}，沿用模板值: #{ranking_values[i]}")
    else:
        print('[p18] 数据库不可用，Ranking 保持模板值')
    
    # 6) 写出汇总 Excel（直接到 p18 目录）
    wb = Workbook()

    # MainTrend
    ws = wb.active
    ws.title = 'MainTrend'
    ws.append(['Month'] + months_main)
    ws.append(['NetLovers%'] + main_values)
    ws.append(['Source'] + main_sources)

    # Ranking
    ws = wb.create_sheet('Ranking')
    ws.append(['Month'] + months_main)
    ws.append(['Position'] + ranking_values)
    ws.append(['Source'] + ranking_sources)

    # Lovers/Neutral/Haters
    ws = wb.create_sheet('Lovers')
    ws.append(['Month'] + months_split)
    ws.append(['Lovers%'] + lovers_values)
    ws.append(['Source'] + lovers_sources)

    ws = wb.create_sheet('Neutral')
    ws.append(['Month'] + months_split)
    ws.append(['Neutral%'] + neutral_values)
    ws.append(['Source'] + neutral_sources)

    ws = wb.create_sheet('Haters')
    ws.append(['Month'] + months_split)
    ws.append(['Haters%'] + haters_values)
    ws.append(['Source'] + haters_sources)

    # === 新增可解释性工作表 ===
    
    # 1. 原始数据详情工作表
    ws_raw = wb.create_sheet('RawData')
    ws_raw.append(['说明', '原始数据详情 - 用于人工检查和解释'])
    ws_raw.append([])  # 空行
    
    # 收集所有月份的原始数据
    all_months = sorted(set(months_main + months_split))
    raw_data_rows = []
    
    if NETICLE_DB.exists():
        with sqlite3.connect(str(NETICLE_DB)) as conn:
            for ym in all_months:
                try:
                    result = _query_monthly_sentiment(conn, BRAND_KEY, ym)
                    if result and len(result) == 4:
                        pos_count, neu_count, neg_count, total_count = result
                        
                        pos_pct = _pct(pos_count, total_count)
                        neu_pct = _pct(neu_count, total_count)
                        neg_pct = _pct(neg_count, total_count)
                        net_pct = pos_pct - neg_pct
                        
                        raw_data_rows.append([
                            ym,
                            pos_count,
                            neu_count, 
                            neg_count,
                            total_count,
                            f"{pos_count}/{total_count}*100 = {pos_pct}%",
                            f"{neu_count}/{total_count}*100 = {neu_pct}%", 
                            f"{neg_count}/{total_count}*100 = {neg_pct}%",
                            f"{pos_pct} - {neg_pct} = {net_pct}%"
                        ])
                    else:
                        raw_data_rows.append([ym, 0, 0, 0, 0, '无数据', '无数据', '无数据', '无数据'])
                except Exception as e:
                    raw_data_rows.append([ym, 'ERROR', 'ERROR', 'ERROR', 'ERROR', str(e), '', '', ''])
    
    ws_raw.append(['月份', '正向提及数', '中性提及数', '负向提及数', '总提及数', 'Lovers%计算', 'Neutral%计算', 'Haters%计算', 'NetLovers%计算'])
    for row in raw_data_rows:
        ws_raw.append(row)
    
    # 2. 计算过程详情工作表
    ws_calc = wb.create_sheet('Calculation')
    ws_calc.append(['说明', '计算过程详情 - 展示具体的计算逻辑和公式'])
    ws_calc.append([])
    
    ws_calc.append(['计算公式说明'])
    ws_calc.append(['指标', '计算公式', '阈值条件'])
    ws_calc.append(['Lovers%', '正向提及数 / 总提及数 × 100', f'polarity >= {POS_MIN}'])
    ws_calc.append(['Neutral%', '中性提及数 / 总提及数 × 100', f'{NEG_MAX} < polarity < {POS_MIN}'])
    ws_calc.append(['Haters%', '负向提及数 / 总提及数 × 100', f'polarity <= {NEG_MAX}'])
    ws_calc.append(['NetLovers%', 'Lovers% - Haters%', ''])
    ws_calc.append([])
    
    ws_calc.append(['数据处理说明'])
    ws_calc.append(['步骤', '说明'])
    ws_calc.append(['1. 数据过滤', f'国家ID = {COUNTRY_ID} ({COUNTRY_NAME})'])
    ws_calc.append(['2. 品牌匹配', f'关键词包含: {KEYWORD_LIKE}'])
    ws_calc.append(['3. 情感分类', f'使用 {CFG_SENTIMENT_COL} 字段进行情感分类'])
    ws_calc.append(['4. 月度聚合', '按月份聚合各情感类别的提及数'])
    ws_calc.append(['5. 百分比计算', '各类别数量除以总数乘以100，四舍五入到整数'])
    
    # 3. 数据源说明工作表
    ws_source = wb.create_sheet('DataSources')
    ws_source.append(['说明', '数据源详细信息 - 便于追溯和验证'])
    ws_source.append([])
    
    ws_source.append(['配置信息'])
    ws_source.append(['项目', '值', '说明'])
    ws_source.append(['数据库路径', str(NETICLE_DB), '情感分析数据来源'])
    ws_source.append(['数据表名', TABLE_NAME, '使用的数据表'])
    ws_source.append(['国家ID', COUNTRY_ID, f'目标国家: {COUNTRY_NAME}'])
    ws_source.append(['品牌关键词', BRAND_KEY, '目标品牌'])
    ws_source.append(['关键词匹配', KEYWORD_LIKE, '数据库查询条件'])
    ws_source.append(['情感字段', CFG_SENTIMENT_COL, '用于情感分类的字段'])
    ws_source.append([])
    
    ws_source.append(['月份配置'])
    ws_source.append(['类型', '月份列表', '用途'])
    ws_source.append(['主趋势月份', ', '.join(months_main), 'NetLovers%趋势图和排名'])
    ws_source.append(['分割月份', ', '.join(months_split), 'Lovers/Neutral/Haters三小图'])
    ws_source.append(['自动发现', str(AUTO_MONTHS), '是否自动发现可用月份'])
    ws_source.append([])
    
    ws_source.append(['数据更新状态'])
    ws_source.append(['数据类型', '更新数量', '总数量', '更新率'])
    updated_main = sum(1 for s in main_sources if s == 'computed')
    updated_lovers = sum(1 for s in lovers_sources if s == 'computed') 
    updated_neutral = sum(1 for s in neutral_sources if s == 'computed')
    updated_haters = sum(1 for s in haters_sources if s == 'computed')
    ws_source.append(['主趋势', updated_main, len(main_sources), f'{updated_main/len(main_sources)*100:.1f}%'])
    ws_source.append(['Lovers', updated_lovers, len(lovers_sources), f'{updated_lovers/len(lovers_sources)*100:.1f}%'])
    ws_source.append(['Neutral', updated_neutral, len(neutral_sources), f'{updated_neutral/len(neutral_sources)*100:.1f}%'])
    ws_source.append(['Haters', updated_haters, len(haters_sources), f'{updated_haters/len(haters_sources)*100:.1f}%'])
    ws_source.append(['排名', computed_ranking_count, len(ranking_sources), f'{computed_ranking_count/len(ranking_sources)*100:.1f}%'])
    
    # 4. 阈值配置工作表
    ws_thresh = wb.create_sheet('Thresholds')
    ws_thresh.append(['说明', '情感分析阈值配置 - 用于理解分类标准'])
    ws_thresh.append([])
    
    ws_thresh.append(['情感阈值设置'])
    ws_thresh.append(['情感类别', '阈值条件', '数值范围', '说明'])
    ws_thresh.append(['正向 (Lovers)', f'polarity >= {POS_MIN}', f'[{POS_MIN}, +∞)', '积极正面的提及'])
    ws_thresh.append(['中性 (Neutral)', f'{NEG_MAX} < polarity < {POS_MIN}', f'({NEG_MAX}, {POS_MIN})', '中性客观的提及'])
    ws_thresh.append(['负向 (Haters)', f'polarity <= {NEG_MAX}', f'(-∞, {NEG_MAX}]', '消极负面的提及'])
    ws_thresh.append([])
    
    ws_thresh.append(['排名计算'])
    ws_thresh.append(['参与品牌', ', '.join(BRANDS_FOR_RANK)])
    ws_thresh.append(['排名依据', 'NetLovers% (Lovers% - Haters%)'])
    ws_thresh.append(['排序方式', '降序 (NetLovers%越高排名越靠前)'])
    ws_thresh.append([])
    
    ws_thresh.append(['数据质量检查'])
    ws_thresh.append(['检查项', '状态', '说明'])
    db_status = '正常' if NETICLE_DB.exists() else '缺失'
    ws_thresh.append(['数据库文件', db_status, str(NETICLE_DB)])
    
    total_computed = updated_main + updated_lovers + updated_neutral + updated_haters + computed_ranking_count
    data_status = '正常' if total_computed > 0 else '异常'
    ws_thresh.append(['计算数据', data_status, f'共更新{total_computed}个数据点'])
    
    template_status = '正常' if any(s == 'template' for s in main_sources + lovers_sources + neutral_sources + haters_sources + ranking_sources) else '缺失'
    ws_thresh.append(['模板数据', template_status, '作为基线数据使用'])

    # 输出到 p18 目录而不是 output 目录
    out_xlsx = PAGE_DIR / 'p18_data.xlsx'
    wb.save(str(out_xlsx))
    print(f'[p18] 已生成汇总 Excel：{out_xlsx}')

    # 严格校验：如果没有任何一个值来源为 computed，则报错，防止“全模板生成”。
    updated_main = sum(1 for s in main_sources if s == 'computed')
    updated_lovers = sum(1 for s in lovers_sources if s == 'computed')
    updated_neutral = sum(1 for s in neutral_sources if s == 'computed')
    updated_haters = sum(1 for s in haters_sources if s == 'computed')
    updated_total = updated_main + updated_lovers + updated_neutral + updated_haters
    if updated_total == 0 and computed_ranking_count == 0:
        raise RuntimeError('[p18] 未从数据库得到任何覆盖数据，当前输出等同模板。请检查品牌匹配与月份配置。')

    # 7) 写出 4 份嵌入 xlsx（Sheet1）
    # 注意：首行必须是纯数值，保持与 chartX.xml 的引用范围一致（如 Sheet1!$A$1:$D$1）。
    # 不在嵌入工作簿中写入月份标签，避免 PPT 编辑时因数据类型不匹配导致系列消失。
    def write_embed(name: str, values: list[int]):
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = 'Sheet1'
        ws2.append(values)
        fp = TEMPLATE_EMBED_DIR / name
        wb2.save(str(fp))

    # Chart1 (主趋势) - 6个月 Mar-Aug
    write_embed('Microsoft_Office_Excel_Binary_Worksheet1.xlsx', main_values)
    # Chart2-4 (情感分析) - 4个月 May-Aug
    write_embed('Microsoft_Office_Excel_Binary_Worksheet2.xlsx', lovers_values)
    write_embed('Microsoft_Office_Excel_Binary_Worksheet3.xlsx', neutral_values)
    write_embed('Microsoft_Office_Excel_Binary_Worksheet4.xlsx', haters_values)

    print(f'[p18] 生成数据: {out_xlsx}')
    print(f'[p18] 生成嵌入: {TEMPLATE_EMBED_DIR}')
    print(f'[p18] 主趋势数据源: {main_sources}')
    print(f'[p18] 三小图数据源: Lovers={lovers_sources}, Neutral={neutral_sources}, Haters={haters_sources}')
    print(f'[p18] 排名数据源: {ranking_sources}, 值: {ranking_values}')


if __name__ == '__main__':
    main()