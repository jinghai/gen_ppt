#!/usr/bin/env python3
# 页内日志：stdout/stderr 双写到 logs/build.log，利于后续排障
import sys, atexit
from datetime import datetime
from pathlib import Path as _PathForLogging

def _init_file_logging():
    page_dir = _PathForLogging(__file__).resolve().parent
    logs_dir = page_dir / 'logs'
    logs_dir.mkdir(parents=True, exist_ok=True)
    fh = open(logs_dir / 'build.log', 'a', encoding='utf-8')
    class _Tee:
        def __init__(self, *streams): self.streams = streams
        def write(self, data):
            for s in self.streams:
                try: s.write(data); s.flush()
                except Exception: pass
        def flush(self):
            for s in self.streams:
                try: s.flush()
                except Exception: pass
    sys.stdout = _Tee(sys.__stdout__, fh)
    sys.stderr = _Tee(sys.__stderr__, fh)
    print(f"[log] start {datetime.utcnow().isoformat()}Z")
    def _close():
        print(f"[log] end {datetime.utcnow().isoformat()}Z")
        try: fh.flush(); fh.close()
        except Exception: pass
    atexit.register(_close)

_init_file_logging()
import zipfile
from pathlib import Path
import yaml
from lxml import etree as ET
from typing import Optional

SLIDE_NO = 18
BASE = Path(__file__).resolve().parents[2]
def _resolve_tpl() -> Path:
    try:
        cfg = yaml.safe_load((BASE / 'config.yaml').read_text(encoding='utf-8')) or {}
    except Exception:
        cfg = {}
    op = (cfg.get('project') or {}).get('original_ppt', 'input/LRTBH.pptx')
    p = Path(op)
    if not p.is_absolute():
        p = BASE / op
    return p

TPL = _resolve_tpl()

def _resolve_unz_root() -> Path:
    try:
        cfg = yaml.safe_load((BASE / 'config.yaml').read_text(encoding='utf-8')) or {}
    except Exception:
        cfg = {}
    tr = (cfg.get('project') or {}).get('template_root', 'input/LRTBH-unzip')
    p = Path(tr)
    if not p.is_absolute():
        p = BASE / tr
    return p

UNZ = _resolve_unz_root()
# 标准化输出目录至页面内，避免跨页共享输出目录导致混淆
PAGE_DIR = Path(__file__).resolve().parent
OUT = PAGE_DIR / 'output' / f'p{SLIDE_NO}-final.pptx'

# 嵌入工作簿目录：统一使用tmp/ppt/ppt/embeddings目录
TMP_DIR = PAGE_DIR / 'tmp'
EMBED_DIR = TMP_DIR / 'ppt' / 'ppt' / 'embeddings'
if not EMBED_DIR.exists():
    EMBED_DIR = UNZ / 'ppt' / 'embeddings'

# 图表目录：统一使用tmp/ppt/ppt/charts目录
# 优先级：tmp/ppt/ppt/charts → UNZ/ppt/charts → UNZ/charts
PAGE_CHARTS_DIR = TMP_DIR / 'ppt' / 'ppt' / 'charts'
PAGE_CHARTS_RELS_DIR = PAGE_CHARTS_DIR / '_rels'
if PAGE_CHARTS_DIR.exists():
    CHARTS_DIR = PAGE_CHARTS_DIR
    CHARTS_RELS_DIR = PAGE_CHARTS_RELS_DIR
else:
    CHARTS_DIR = (UNZ / 'ppt' / 'charts') if (UNZ / 'ppt' / 'charts').exists() else (UNZ / 'charts')
    CHARTS_RELS_DIR = CHARTS_DIR / '_rels'

APP_XML = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties" xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes">
<Application>Microsoft Office PowerPoint</Application>
<PresentationFormat>Widescreen</PresentationFormat>
<TotalTime>0</TotalTime>
<Words>0</Words>
<Slides>1</Slides>
<Notes>0</Notes>
<HiddenSlides>0</HiddenSlides>
<MMClips>0</MMClips>
<ScaleCrop>false</ScaleCrop>
<LinksUpToDate>false</LinksUpToDate>
<SharedDoc>false</SharedDoc>
<HyperlinksChanged>false</HyperlinksChanged>
<AppVersion>16.0000</AppVersion>
</Properties>'''

NS = {
    'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
    'c': 'http://schemas.openxmlformats.org/drawingml/2006/chart',
    'p': 'http://schemas.openxmlformats.org/presentationml/2006/main',
}


def build():
    OUT.parent.mkdir(parents=True, exist_ok=True)

    # 读取模板中的 presentation 关系、结构与内容类型
    with zipfile.ZipFile(TPL, 'r') as tpl:
        pres_rels_bytes = tpl.read('ppt/_rels/presentation.xml.rels')
        pres_xml_bytes = tpl.read('ppt/presentation.xml')
        ct_bytes = tpl.read('[Content_Types].xml')

    # 预加载图表及关系作为回退（页面级 tmp 优先）
    charts_fallback = {}
    if CHARTS_DIR.exists():
        for chart_file in CHARTS_DIR.glob('chart*.xml'):
            charts_fallback[chart_file.name] = chart_file.read_bytes()
        for rels_file in CHARTS_RELS_DIR.glob('chart*.xml.rels'):
            charts_fallback[f'rels:{rels_file.name}'] = rels_file.read_bytes()

    # 过滤 presentation.xml.rels，仅保留指向指定slide的关系（重定向到 slide1）以及非 slide 的其他关系
    def filter_pres_rels(rels_bytes: bytes):
        root = ET.fromstring(rels_bytes)
        new_root = ET.Element('Relationships', xmlns='http://schemas.openxmlformats.org/package/2006/relationships')
        slide_rel_id = None
        for rel in root.findall('{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
            t = rel.get('Target') or ''
            typ = rel.get('Type') or ''
            if typ.endswith('/slide') and t == f'slides/slide{SLIDE_NO}.xml':
                slide_rel_id = rel.get('Id')
                e = ET.SubElement(new_root, 'Relationship')
                for k, v in rel.attrib.items():
                    e.set(k, v)
                # 重定向为 slide1
                e.set('Target', 'slides/slide1.xml')
            elif not typ.endswith('/slide'):
                e = ET.SubElement(new_root, 'Relationship')
                for k, v in rel.attrib.items():
                    e.set(k, v)
        return ET.tostring(new_root, xml_declaration=True, encoding='UTF-8', standalone="yes"), slide_rel_id

    # 过滤 presentation.xml，仅保留对应 slide_rel_id 的 sldId
    def filter_pres_xml(pres_bytes: bytes, slide_rel_id: Optional[str]):
        tree = ET.fromstring(pres_bytes)
        sldIdLst = tree.find('{%s}sldIdLst' % NS['p'])
        if sldIdLst is None:
            sldIdLst = ET.SubElement(tree, '{%s}sldIdLst' % NS['p'])
        for ch in list(sldIdLst):
            sldIdLst.remove(ch)
        ET.SubElement(sldIdLst, '{%s}sldId' % NS['p'], attrib={'id': '256', '{%s}id' % NS['r']: slide_rel_id or 'rId1'})
        return ET.tostring(tree, xml_declaration=True, encoding='UTF-8', standalone="yes")

    # 过滤 Content_Types：移除全部 slide 覆盖项后添加 slide1
    def filter_content_types(ct_bytes: bytes, embed_files: list[str]):
        ct = ET.fromstring(ct_bytes)
        ns_ct = 'http://schemas.openxmlformats.org/package/2006/content-types'
        for o in list(ct.findall('{%s}Override' % ns_ct)):
            part = o.get('PartName') or ''
            if part.startswith('/ppt/slides/slide'):
                ct.remove(o)
        ET.SubElement(ct, '{%s}Override' % ns_ct, PartName='/ppt/slides/slide1.xml', ContentType='application/vnd.openxmlformats-officedocument.presentationml.slide+xml')
        # 确保存在 docProps/app.xml 覆盖项
        has_app = any((o.get('PartName') == '/docProps/app.xml') for o in ct.findall('{%s}Override' % ns_ct))
        if not has_app:
            ET.SubElement(ct, '{%s}Override' % ns_ct, PartName='/docProps/app.xml', ContentType='application/vnd.openxmlformats-officedocument.extended-properties+xml')
        # 添加嵌入式 xlsx 覆盖项
        for nm in embed_files:
            ET.SubElement(ct, '{%s}Override' % ns_ct, PartName=f'/ppt/embeddings/{nm}', ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        return ET.tostring(ct, xml_declaration=True, encoding='UTF-8', standalone="yes")

    # 生成新文件：复制模板的所有项，删除除指定slide之外的其它幻灯片及其 rels，替换 presentation 与 Content_Types；并重命名指定slide→slide1
    with zipfile.ZipFile(TPL, 'r') as tpl:
        new_pres_rels_xml, slide_rel_id = filter_pres_rels(pres_rels_bytes)
        new_pres_xml = filter_pres_xml(pres_xml_bytes, slide_rel_id)
        # 收集嵌入式 xlsx（跳过 xlsb）
        embed_files = []
        try:
            if EMBED_DIR.exists():
                for f in EMBED_DIR.glob('Microsoft_Office_Excel_Binary_Worksheet*.xlsx'):
                    embed_files.append(f.name)
        except Exception:
            embed_files = []
        new_ct_xml = filter_content_types(ct_bytes, embed_files)
        
        # 读取指定slide的关系文件以获取图表信息
        slide_rels_bytes = None
        try:
            slide_rels_bytes = tpl.read(f'ppt/slides/_rels/slide{SLIDE_NO}.xml.rels')
        except KeyError:
            pass

        # 将 tmp/ppt/ppt/charts 的 chart1..4.xml 映射到当前页的目标图表名（例如 chart25..28.xml）
        def _remap_charts_for_slide(slide_rels_bytes: Optional[bytes]):
            if not slide_rels_bytes:
                return
            try:
                root = ET.fromstring(slide_rels_bytes)
                ns_rel = 'http://schemas.openxmlformats.org/package/2006/relationships'
                targets = []
                for rel in root.findall('{%s}Relationship' % ns_rel):
                    typ = rel.get('Type') or ''
                    if typ.endswith('/chart'):
                        tgt = rel.get('Target') or ''
                        nm = tgt.split('/')[-1]  # e.g. '../charts/chart25.xml' -> 'chart25.xml'
                        targets.append(nm)
                # 若页面级存在 chart1..4，则按出现顺序映射到目标名
                for idx, nm in enumerate(targets, start=1):
                    src_nm = f'chart{idx}.xml'
                    src_rels_nm = f'chart{idx}.xml.rels'
                    if src_nm in charts_fallback:
                        charts_fallback[nm] = charts_fallback[src_nm]
                    if f'rels:{src_rels_nm}' in charts_fallback:
                        charts_fallback[f"rels:{nm.replace('.xml', '.xml.rels')}"] = charts_fallback[f'rels:{src_rels_nm}']
            except Exception:
                # 映射失败不兜底，保持原模板图表
                pass

        _remap_charts_for_slide(slide_rels_bytes)
        
        # 读取 Ranking 值（用于更新 slide 文本）
        def _read_ranking_values() -> list[int]:
            from openpyxl import load_workbook
            xlsx = PAGE_DIR / 'p18_data.xlsx'
            if not xlsx.exists():
                raise FileNotFoundError(f'[p{SLIDE_NO}] 缺少排名数据文件: {xlsx}，请先生成')
            wb = load_workbook(str(xlsx), data_only=True)
            if 'Ranking' not in wb.sheetnames:
                raise RuntimeError('[p18] p18_data.xlsx 缺少 Ranking 工作表')
            ws = wb['Ranking']
            # 第二行：Position
            row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
            vals = []
            for v in row[1:]:
                if v is None:
                    raise RuntimeError('[p18] Ranking 存在空值，禁止兜底')
                vals.append(int(v))
            if not vals:
                raise RuntimeError('[p18] Ranking 为空，禁止兜底')
            return vals

        def _update_ranking_in_slide(slide_xml_bytes: bytes, ranks: list[int]) -> bytes:
            # 目标：将形如 <a:t>#3</a:t> 的文本按顺序替换为 ranks
            root = ET.fromstring(slide_xml_bytes)
            ns_a = 'http://schemas.openxmlformats.org/drawingml/2006/main'
            nodes = []
            for t in root.findall('.//{%s}t' % ns_a):
                text = (t.text or '')
                if text.startswith('#') and text[1:].isdigit():
                    nodes.append(t)
            if len(nodes) != len(ranks):
                raise RuntimeError(f'[p18] Ranking 文本节点数量({len(nodes)})与数据数量({len(ranks)})不一致')
            for i, t in enumerate(nodes):
                t.text = f'#{int(ranks[i])}'
            return ET.tostring(root, xml_declaration=True, encoding='UTF-8', standalone='yes')

        ranks = None
        try:
            ranks = _read_ranking_values()
        except Exception as e:
            # 按用户要求：不做兜底，直接抛错
            raise

        with zipfile.ZipFile(OUT, 'w', compression=zipfile.ZIP_DEFLATED) as z:
            for name in tpl.namelist():
                # 去除非指定页的幻灯片
                if name.startswith('ppt/slides/slide') and name != f'ppt/slides/slide{SLIDE_NO}.xml':
                    continue
                if name.startswith('ppt/slides/_rels/slide') and name != f'ppt/slides/_rels/slide{SLIDE_NO}.xml.rels':
                    continue
                if name == 'ppt/_rels/presentation.xml.rels':
                    z.writestr(name, new_pres_rels_xml)
                elif name == 'ppt/presentation.xml':
                    z.writestr(name, new_pres_xml)
                elif name == '[Content_Types].xml':
                    z.writestr(name, new_ct_xml)
                elif name == f'ppt/slides/slide{SLIDE_NO}.xml':
                    # 重命名为 slide1.xml，并更新 Ranking；主趋势百分比改为使用图表数据标签
                    original = tpl.read(name)
                    updated = _update_ranking_in_slide(original, ranks)
                    # 读取 MainTrend 数值并覆盖前 6 个百分比文本（形如 "35%"）
                    def _clear_main_percent_texts(slide_xml_bytes: bytes) -> bytes:
                        import re as _re
                        root = ET.fromstring(slide_xml_bytes)
                        ns_a = 'http://schemas.openxmlformats.org/drawingml/2006/main'
                        cleared = 0
                        for t in root.findall('.//{%s}t' % ns_a):
                            tx = (t.text or '').strip()
                            if _re.fullmatch(r'\d+%$', tx):
                                t.text = ''
                                cleared += 1
                        if cleared == 0:
                            raise RuntimeError('[p18] 未发现可清理的静态百分比文本节点')
                        return ET.tostring(root, xml_declaration=True, encoding='UTF-8', standalone='yes')

                    updated = _clear_main_percent_texts(updated)
                    z.writestr('ppt/slides/slide1.xml', updated)
                elif name == f'ppt/slides/_rels/slide{SLIDE_NO}.xml.rels':
                    # 重命名为 slide1.xml.rels
                    z.writestr('ppt/slides/_rels/slide1.xml.rels', tpl.read(name))
                elif name.startswith('ppt/embeddings/Microsoft_Office_Excel_Binary_Worksheet') and name.endswith('.xlsb'):
                    # 跳过模板中的 xlsb，避免与新生成的 xlsx 冲突
                    continue
                elif name.startswith('ppt/charts/chart') and name.endswith('.xml'):
                    # 使用回退图表数据（如果有）
                    chart_name = name.split('/')[-1]
                    fallback_data = charts_fallback.get(chart_name)
                    z.writestr(name, fallback_data if fallback_data is not None else tpl.read(name))
                elif name.startswith('ppt/charts/_rels/chart') and name.endswith('.xml.rels'):
                    # 使用回退关系数据（如果有）
                    rels_name = name.split('/')[-1]
                    fallback_data = charts_fallback.get(f'rels:{rels_name}')
                    z.writestr(name, fallback_data if fallback_data is not None else tpl.read(name))
                else:
                    z.writestr(name, tpl.read(name))
            # Write docProps/app.xml if it doesn't exist
            if 'docProps/app.xml' not in tpl.namelist():
                z.writestr('docProps/app.xml', APP_XML)
            # 打包嵌入式 xlsx（若存在）
            for nm in embed_files:
                src = EMBED_DIR / nm
                if src.exists():
                    z.writestr(f'ppt/embeddings/{nm}', src.read_bytes())
    print(f'Built {OUT}')


import subprocess, sys, re

def run_make_data():
    """运行页面级数据生成脚本 generate_excel.py，产出 output/p18_data.xlsx 与嵌入 xlsx。"""
    page_dir = Path(__file__).resolve().parent
    script = page_dir / 'generate_excel.py'
    if script.exists():
        print(f'[p{SLIDE_NO}] run generate_excel.py')
        subprocess.run([sys.executable, str(script)], check=True, cwd=str(page_dir))
    else:
        print(f'[p{SLIDE_NO}] skip generate_excel.py (not found)')

def run_fillers():
    """运行页面级填充脚本 fill_from_excel.py，更新 chartXML 与 rels。"""
    page_dir = Path(__file__).resolve().parent
    script = page_dir / 'fill_from_excel.py'
    if script.exists():
        print(f'[p{SLIDE_NO}] run fill_from_excel.py')
        subprocess.run([sys.executable, str(script)], check=True, cwd=str(page_dir))
    else:
        print(f'[p{SLIDE_NO}] skip fill_from_excel.py (not found)')

if __name__ == '__main__':
    # 委托到页面级填充与打包脚本，避免重复打包与样式偏差
    page_dir = Path(__file__).resolve().parent
    script = page_dir / 'fill_from_excel.py'
    if not script.exists():
        raise FileNotFoundError(f'缺少脚本: {script}')
    subprocess.run([sys.executable, str(script)], check=True, cwd=str(page_dir))
