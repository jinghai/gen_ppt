#!/usr/bin/env python3
import argparse
import json
from pathlib import Path
from lxml import etree
import yaml
try:
    # 仅在需要写入嵌入式 xlsx 时使用；未安装时提示但不阻断其他流程
    from openpyxl import Workbook  # type: ignore
except Exception:
    Workbook = None

NS = {
  'c': 'http://schemas.openxmlformats.org/drawingml/2006/chart',
  'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
  'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
}

def replace_num_cache(val_el, values):
    # 通用数值填充，适配 numRef/numLit 两种缓存形式
    if val_el is None:
        return
    num_ref = val_el.find('c:numRef', namespaces=NS)
    if num_ref is not None:
        num_cache = num_ref.find('c:numCache', namespaces=NS)
        if num_cache is None:
            num_cache = etree.SubElement(num_ref, f'{{{NS["c"]}}}numCache')
        for pt in num_cache.findall('c:pt', namespaces=NS):
            num_cache.remove(pt)
        cnt = num_cache.find('c:ptCount', namespaces=NS)
        if cnt is None:
            cnt = etree.SubElement(num_cache, f'{{{NS["c"]}}}ptCount')
        cnt.set('val', str(len(values)))
        for i, v in enumerate(values):
            pt = etree.SubElement(num_cache, f'{{{NS["c"]}}}pt')
            pt.set('idx', str(i))
            f = etree.SubElement(pt, f'{{{NS["c"]}}}v')
            f.text = str(v)
    else:
        num_lit = val_el.find('c:numLit', namespaces=NS)
        if num_lit is None:
            num_lit = etree.SubElement(val_el, f'{{{NS["c"]}}}numLit')
        for pt in num_lit.findall('c:pt', namespaces=NS):
            num_lit.remove(pt)
        cnt = num_lit.find('c:ptCount', namespaces=NS)
        if cnt is None:
            cnt = etree.SubElement(num_lit, f'{{{NS["c"]}}}ptCount')
        cnt.set('val', str(len(values)))
        for i, v in enumerate(values):
            pt = etree.SubElement(num_lit, f'{{{NS["c"]}}}pt')
            pt.set('idx', str(i))
            f = etree.SubElement(pt, f'{{{NS["c"]}}}v')
            f.text = str(v)

def replace_str_lit(cat_el, labels):
    str_ref = cat_el.find('c:strRef', namespaces=NS)
    if str_ref is not None:
        str_cache = str_ref.find('c:strCache', namespaces=NS)
        if str_cache is None:
            str_cache = etree.SubElement(str_ref, f'{{{NS["c"]}}}strCache')
        for pt in str_cache.findall('c:pt', namespaces=NS):
            str_cache.remove(pt)
        cnt = str_cache.find('c:ptCount', namespaces=NS)
        if cnt is None:
            cnt = etree.SubElement(str_cache, f'{{{NS["c"]}}}ptCount')
        cnt.set('val', str(len(labels)))
        for i, v in enumerate(labels):
            pt = etree.SubElement(str_cache, f'{{{NS["c"]}}}pt')
            pt.set('idx', str(i))
            f = etree.SubElement(pt, f'{{{NS["c"]}}}v')
            f.text = str(v)
    else:
        str_lit = cat_el.find('c:strLit', namespaces=NS)
        if str_lit is None:
            str_lit = etree.SubElement(cat_el, f'{{{NS["c"]}}}strLit')
        for pt in str_lit.findall('c:pt', namespaces=NS):
            str_lit.remove(pt)
        cnt = str_lit.find('c:ptCount', namespaces=NS)
        if cnt is None:
            cnt = etree.SubElement(str_lit, f'{{{NS["c"]}}}ptCount')
        cnt.set('val', str(len(labels)))
        for i, v in enumerate(labels):
            pt = etree.SubElement(str_lit, f'{{{NS["c"]}}}pt')
            pt.set('idx', str(i))
            f = etree.SubElement(pt, f'{{{NS["c"]}}}v')
            f.text = str(v)


def _parse_num_cache(val_el):
    if val_el is None:
        return []
    num_ref = val_el.find('c:numRef', namespaces=NS)
    pts = []
    if num_ref is not None:
        num_cache = num_ref.find('c:numCache', namespaces=NS)
        if num_cache is not None:
            pts = num_cache.findall('c:pt', namespaces=NS)
    else:
        num_lit = val_el.find('c:numLit', namespaces=NS)
        if num_lit is not None:
            pts = num_lit.findall('c:pt', namespaces=NS)
    values = []
    for pt in pts:
        v_el = pt.find('c:v', namespaces=NS)
        txt = v_el.text if v_el is not None else ''
        try:
            values.append(float(txt))
        except (TypeError, ValueError):
            values.append(txt)
    return values


def _parse_str_cache(cat_el):
    if cat_el is None:
        return []
    str_ref = cat_el.find('c:strRef', namespaces=NS)
    pts = []
    if str_ref is not None:
        str_cache = str_ref.find('c:strCache', namespaces=NS)
        if str_cache is not None:
            pts = str_cache.findall('c:pt', namespaces=NS)
    else:
        str_lit = cat_el.find('c:strLit', namespaces=NS)
        if str_lit is not None:
            pts = str_lit.findall('c:pt', namespaces=NS)
    labels = []
    for pt in pts:
        v_el = pt.find('c:v', namespaces=NS)
        labels.append(v_el.text if v_el is not None else '')
    return labels


def _load_config():
    """加载全局配置，并在存在页面上下文时叠加该页的局部配置与微模板根。

    覆盖策略：
    - fill_policy: 局部覆盖同名键；
    - project.template_root: 若当前工作目录位于 charts/pXX 且存在 pXX/template，则将其作为覆盖；
      若局部 config.yaml 中显式提供 project.template_root，则优先使用该值。
    """
    root = Path(__file__).resolve().parents[1]
    cfg = {}
    gp = root / 'config.yaml'
    if gp.exists():
        try:
            cfg = yaml.safe_load(gp.read_text(encoding='utf-8')) or {}
        except Exception:
            cfg = {}

    # 尝试发现页面目录（基于当前工作目录）
    cwd = Path.cwd()
    page_dir = None
    for anc in [cwd] + list(cwd.parents):
        if anc.name.startswith('p') and anc.parent.name == 'charts' and (anc / 'config.yaml').exists():
            page_dir = anc
            break
    # 兼容旧结构 charts/p10 与顶层 p10
    if page_dir is None:
        cand = root / 'p10'
        if cand.exists() and (cand / 'config.yaml').exists():
            page_dir = cand

    # 叠加局部配置与微模板根
    if page_dir is not None:
        try:
            local_cfg = yaml.safe_load((page_dir / 'config.yaml').read_text(encoding='utf-8')) or {}
        except Exception:
            local_cfg = {}
        # fill_policy 合并
        if isinstance(local_cfg.get('fill_policy'), dict):
            base = cfg.get('fill_policy', {}) or {}
            merged = {**base, **local_cfg['fill_policy']}
            cfg['fill_policy'] = merged
        # project.template_root 覆盖：优先局部配置；其次检测 page_dir/template 存在
        proj = cfg.get('project') or {}
        local_proj = local_cfg.get('project') or {}
        tpl_override = local_proj.get('template_root')
        if not tpl_override:
            if (page_dir / 'template').exists():
                tpl_override = str(page_dir / 'template')
        if tpl_override:
            if 'project' not in cfg:
                cfg['project'] = {}
            cfg['project']['template_root'] = tpl_override
    return cfg


def _get(cfg, keys, default=None):
    cur = cfg
    for k in keys:
        if not isinstance(cur, dict) or k not in cur:
            return default
        cur = cur[k]
    return cur if cur is not None else default


def _charts_dir_from_config(cfg: dict) -> Path:
    root = Path(__file__).resolve().parents[1]
    tr = ((cfg.get('project') or {}).get('template_root') or 'input/LRTBH-unzip')
    trp = Path(tr)
    if not trp.is_absolute():
        trp = root / trp
    d = trp / 'ppt' / 'charts'
    return d if d.exists() else (trp / 'charts')


def resolve_chart_xml_path(raw: Path) -> Path:
    """根据 config.yaml 解析 chart*.xml 绝对路径（兼容相对/绝对/只含文件名）。"""
    if raw.exists():
        return raw
    cfg = _load_config()
    charts_dir = _charts_dir_from_config(cfg)
    # 仅文件名
    cand = charts_dir / raw.name
    if cand.exists():
        return cand
    # 相对路径：尝试以仓库根与 template_root 解析
    if not raw.is_absolute():
        repo_root = Path(__file__).resolve().parents[1]
        cand2 = repo_root / raw
        if cand2.exists():
            return cand2
        cand3 = charts_dir.parent.parent / raw  # template_root / <raw>
        if cand3.exists():
            return cand3
    # 绝对路径失效：回退 basename
    fallback = charts_dir / raw.name
    if fallback.exists():
        return fallback
    raise FileNotFoundError(f"chart xml not found. tried: {raw}, {cand}, {fallback}")


def fill_chart(chart_xml_path: Path, series_data: list, category_labels: list):
    chart_xml_path = resolve_chart_xml_path(chart_xml_path)
    tree = etree.parse(str(chart_xml_path))
    root = tree.getroot()
    plot = root.find('c:chart/c:plotArea', namespaces=NS)
    if plot is None:
        raise RuntimeError('No plotArea found')
    chart_types = ['lineChart','barChart','pieChart','scatterChart','areaChart','doughnutChart']
    target_chart = None
    for t in chart_types:
        el = plot.find(f'c:{t}', namespaces=NS)
        if el is not None:
            target_chart = el
            break
    if target_chart is None:
        raise RuntimeError('Unsupported chart type')

    cfg = _load_config()
    keep_original = bool(_get(cfg, ['fill_policy','keep_original_for_missing'], False))
    axis_base = int(_get(cfg, ['fill_policy','axis_day_base'], 20300))
    external_fmt = str(_get(cfg, ['fill_policy','external_data_format'], 'xlsb')).lower()
    if external_fmt not in ('xlsb','xlsx'):
        external_fmt = 'xlsb'
    # 快照用于在 external_data_format=xlsx 时生成嵌入工作簿
    snapshot = {
        'type': None,  # 'scatter' | 'category'
        'x': [],       # list[float]
        'ys': [],      # list[list[float]]
    }

    # scatterChart: 使用 xVal/yVal，支持“缺失数据回退到原始图表”
    if target_chart.tag.endswith('scatterChart'):
        series_els = target_chart.findall('c:ser', namespaces=NS)
        snapshot['type'] = 'scatter'
        for idx, ser in enumerate(series_els):
            new_y = series_data[idx] if (series_data and idx < len(series_data)) else []
            y_el = ser.find('c:yVal', namespaces=NS)
            x_el = ser.find('c:xVal', namespaces=NS)
            if keep_original:
                orig_y = _parse_num_cache(y_el)
                orig_x = _parse_num_cache(x_el)
                final_len = max(len(orig_y), len(new_y))
                final_y = [
                    new_y[i] if i < len(new_y) else (orig_y[i] if i < len(orig_y) else '')
                    for i in range(final_len)
                ]
                final_x = []
                for i in range(final_len):
                    if i < len(new_y):
                        final_x.append(axis_base + 1 + i)
                    else:
                        final_x.append(orig_x[i] if i < len(orig_x) else (axis_base + 1 + i))
            else:
                final_y = new_y
                final_x = [axis_base + 1 + i for i in range(len(new_y))]
            replace_num_cache(y_el, final_y)
            if x_el is not None:
                replace_num_cache(x_el, final_x)
            # 记录快照：第一条 series 的 X，所有 series 的 Y
            if idx == 0:
                snapshot['x'] = final_x
            snapshot['ys'].append(final_y)
    else:
        # 类别型图表：按标签对齐，缺失标签用原始缓存补齐
        cat_el = target_chart.find('c:cat', namespaces=NS)
        orig_labels = _parse_str_cache(cat_el) if cat_el is not None else []
        series_els = target_chart.findall('c:ser', namespaces=NS)
        n_series = len(series_els)
        snapshot['type'] = 'category'

        if keep_original:
            # 原始映射：label -> [s0,s1,...]
            orig_series_vals = [
                _parse_num_cache(ser.find('c:val', namespaces=NS)) for ser in series_els
            ]
            orig_map = {}
            for i, lbl in enumerate(orig_labels):
                orig_map[lbl] = [
                    (sv[i] if i < len(sv) else '') for sv in orig_series_vals
                ]
            # 新映射
            new_map = {}
            if category_labels is not None and series_data:
                for i, lbl in enumerate(category_labels):
                    new_map[lbl] = [
                        (series_data[s_idx][i] if s_idx < len(series_data) and i < len(series_data[s_idx]) else '')
                        for s_idx in range(n_series)
                    ]
            # 最终标签：以原始顺序为主，追加新出现的标签
            final_labels = list(orig_labels)
            if category_labels:
                for lbl in category_labels:
                    if lbl not in final_labels:
                        final_labels.append(lbl)
            # 组装最终系列数据
            final_series = [[ ] for _ in range(n_series)]
            for lbl in final_labels:
                src = new_map[lbl] if lbl in new_map else orig_map.get(lbl, ['']*n_series)
                for s_idx in range(n_series):
                    final_series[s_idx].append(src[s_idx])
            # 写入
            if cat_el is not None:
                replace_str_lit(cat_el, final_labels)
            for idx, ser in enumerate(series_els):
                val_el = ser.find('c:val', namespaces=NS)
                values = final_series[idx] if idx < len(final_series) else []
                replace_num_cache(val_el, values)
                snapshot['ys'].append(values)
        else:
            # 原有行为：完全用新 labels 和 series 覆盖
            if cat_el is not None and category_labels is not None:
                replace_str_lit(cat_el, category_labels)
            for idx, ser in enumerate(series_els):
                values = series_data[idx] if (series_data and idx < len(series_data)) else []
                val_el = ser.find('c:val', namespaces=NS)
                replace_num_cache(val_el, values)
                snapshot['ys'].append(values)

    # 配置：是否保留外部工作簿链接（允许在 PowerPoint 中“编辑数据”）
    cfg_keep_external = bool(_get(cfg, ['fill_policy','keep_external_data'], False))

    # 外部工作簿链接处理：默认移除；如开启 keep_external_data，则确保 externalData 与 rels 存在
    if not cfg_keep_external:
        # 移除与嵌入式工作簿的链接，防止 PowerPoint 在编辑/保存时用 xlsb 回写覆盖
        # 1) 删除 chart XML 中的 <c:externalData r:id="..."> 节点
        ext = root.find('c:externalData', namespaces=NS)
        if ext is not None:
            parent = root
            try:
                parent.remove(ext)
            except Exception:
                # 容错：若删除失败，继续写入（不会影响后续 rels 清理）
                pass
    else:
        # 开启保留时：若 externalData 缺失则补回，并设置 r:id="rId1" 与 <c:autoUpdate val="0"/>
        ext = root.find('c:externalData', namespaces=NS)
        if ext is None:
            ext = etree.SubElement(root, f'{{{NS["c"]}}}externalData')
            ext.set(f'{{{NS["r"]}}}id', 'rId1')
            au = etree.SubElement(ext, f'{{{NS["c"]}}}autoUpdate')
            au.set('val', '0')

    tree.write(str(chart_xml_path), encoding='utf-8', xml_declaration=True)

    rels_dir = chart_xml_path.parent / '_rels'
    rels_path = rels_dir / f'{chart_xml_path.name}.rels'
    if not cfg_keep_external:
        # 2) 清理图表 rels 中指向 embeddings/*.xlsb|*.xlsx 的关系
        if rels_path.exists():
            try:
                rel_tree = etree.parse(str(rels_path))
                rel_root = rel_tree.getroot()
                NS_REL = {'rel': 'http://schemas.openxmlformats.org/package/2006/relationships'}
                removed = False
                for rel_el in list(rel_root.findall('rel:Relationship', namespaces=NS_REL)):
                    tgt = rel_el.get('Target', '')
                    if '/embeddings/' in tgt or tgt.endswith('.xlsb') or tgt.endswith('.xlsx'):
                        rel_root.remove(rel_el)
                        removed = True
                if removed:
                    # 若所有关系已清空，删除 .rels；否则回写
                    has_child = len(list(rel_root)) > 0
                    if has_child:
                        rel_tree.write(str(rels_path), encoding='utf-8', xml_declaration=True)
                    else:
                        try:
                            rels_path.unlink()
                        except Exception:
                            # 容错：无法删除则继续保留空文件
                            pass
            except Exception:
                # 容错：rels 解析失败时跳过
                pass
    else:
        # 2) 保留/补齐 rels：确保存在指向 ppt/embeddings/Microsoft_Office_Excel_Binary_WorksheetN.xlsb/xlsx 的关系
        try:
            rels_dir.mkdir(exist_ok=True)
            # 解析编号 N
            name = chart_xml_path.name  # chart8.xml
            try:
                n = int(''.join([ch for ch in name if ch.isdigit()]))
            except Exception:
                n = 1
            target_ext = 'xlsx' if external_fmt == 'xlsx' else 'xlsb'
            target = f"../embeddings/Microsoft_Office_Excel_Binary_Worksheet{n}.{target_ext}"
            NS_REL_URI = 'http://schemas.openxmlformats.org/package/2006/relationships'
            if rels_path.exists():
                rel_tree = etree.parse(str(rels_path))
                rel_root = rel_tree.getroot()
            else:
                rel_root = etree.Element('Relationships')
                rel_root.set('xmlns', NS_REL_URI)
                rel_tree = etree.ElementTree(rel_root)
            # 移除旧的 embeddings 关系，统一为 rId1
            for rel_el in list(rel_root):
                tgt = rel_el.get('Target', '')
                if '/embeddings/' in tgt or tgt.endswith('.xlsb') or tgt.endswith('.xlsx'):
                    rel_root.remove(rel_el)
            new_rel = etree.SubElement(rel_root, 'Relationship')
            new_rel.set('Id', 'rId1')
            new_rel.set('Type', 'http://schemas.openxmlformats.org/officeDocument/2006/relationships/package')
            new_rel.set('Target', target)
            rel_tree.write(str(rels_path), encoding='utf-8', xml_declaration=True)
        except Exception:
            # 容错：补齐失败时忽略（不影响图表数据填充）
            pass

    # 3) 如需生成 xlsx 嵌入工作簿：写入到 ppt/embeddings/Microsoft_Office_Excel_Binary_WorksheetN.xlsx
    if cfg_keep_external and external_fmt == 'xlsx':
        if Workbook is None:
            print('[warn] openpyxl 未安装，无法生成嵌入式 xlsx 工作簿。请在 requirements.txt 添加 openpyxl 并安装。')
        else:
            try:
                n = int(''.join([ch for ch in chart_xml_path.name if ch.isdigit()]))
            except Exception:
                n = 1
            embeddings_dir = chart_xml_path.parent.parent / 'embeddings'
            embeddings_dir.mkdir(exist_ok=True)
            xlsx_path = embeddings_dir / f'Microsoft_Office_Excel_Binary_Worksheet{n}.xlsx'
            wb = Workbook()
            ws = wb.active
            ws.title = 'Sheet1'
            if snapshot['type'] == 'scatter':
                xs = snapshot.get('x') or []
                for j, val in enumerate(xs, start=1):
                    ws.cell(row=1, column=j, value=val)
                for i, ys in enumerate(snapshot.get('ys') or [], start=2):
                    for j, val in enumerate(ys, start=1):
                        ws.cell(row=i, column=j, value=val)
            else:
                first = (snapshot.get('ys') or [[]])[0]
                for r, val in enumerate(first, start=1):
                    ws.cell(row=r, column=1, value=val)
            wb.save(str(xlsx_path))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--chart', required=True, help='chart*.xml 路径')
    ap.add_argument('--series_json', required=True, help='系列数据 JSON 文件路径')
    ap.add_argument('--labels_json', required=False, help='分类标签 JSON 文件路径')
    args = ap.parse_args()
    chart = Path(args.chart)
    series = json.loads(Path(args.series_json).read_text(encoding='utf-8'))
    labels = None
    if args.labels_json:
        labels = json.loads(Path(args.labels_json).read_text(encoding='utf-8'))
    fill_chart(chart, series, labels)
    print(f'filled: {chart}')

if __name__ == '__main__':
    main()